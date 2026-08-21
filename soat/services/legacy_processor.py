"""Genera el informe consolidado SOAT para revisión y carga en Zoho.

Este proceso es independiente de ``informe_renovaciones.py``.

Flujo aplicado
--------------
1. Lee el reporte consolidado exportado desde Zoho.
2. Separa los registros de SOAT y Movilidad.
3. Enriquece el campo ``Motivo cancelación`` desde la hoja ``Movilidad`` del
   libro de referencia, cuando el reporte de Zoho no incluye ese campo.
4. Selecciona un único registro SOAT y un único registro de Movilidad por placa.
5. Construye el informe según el mapa de origen definido en ``Formato informe``.
6. Aplica las siete políticas de ``Criterios asignacion``.
7. Calcula ``ID de registro (Asegurado) CARGA``:
   primero ID SOAT y, si no existe, ID Movilidad.

Estructura recomendada::

    ays_soat/
    ├── informe_renovaciones.py
    ├── informe_soat_analistas.py
    ├── Informe Zoho/
    │   └── SOAT_prueba_4.xlsx
    ├── Formato referencia/
    │   └── SOAT_prueba_3_Def.xlsx
    └── Salidas/

Uso automático::

    python informe_soat_analistas.py

Uso indicando rutas::

    python informe_soat_analistas.py ^
        --entrada "Informe Zoho/SOAT_prueba_4.xlsx" ^
        --referencia "Formato referencia/SOAT_prueba_3_Def.xlsx" ^
        --salida "Salidas/Informe_SOAT_Analistas_Agosto.xlsx" ^
        --verbose

Selección por placa
-------------------
SOAT y Movilidad se seleccionan independientemente. Cuando existen varias
filas de la misma clase para una placa, se prioriza:

1. Estado de la póliza ``Vigente``.
2. Estado del asegurado que contiene ``Activo``.
3. Fecha fin de vigencia más reciente.
4. Última fila del archivo como desempate estable.

La fila elegida se mantiene como una unidad. No se mezclan póliza, estado,
motivo de cancelación o ID provenientes de registros distintos.

Políticas de Gestión SOAT A&S
-----------------------------
La precedencia efectiva es:

1. Vendedor contiene ``Fonconstruimos`` y estado asegurado Movilidad es
   ``Excluido``: se deja en blanco. Corresponde al criterio 8; es la única
   excepción al criterio 1.
2. Estado de la póliza Movilidad es ``Vigente`` y estado asegurado
   Movilidad es ``Excluido``: ``No gestión``. Corresponde al criterio 9.
3. Vendedor contiene ``Fonconstruimos``:
   ``Gestión A&S``. Corresponde al criterio 1; inamovible salvo los
   criterios 8 y 9.
4. Estado asegurado Movilidad es ``Cancelado`` y el motivo es
   ``Por venta`` o ``Por cambio de intermediario``:
   ``No gestión``. Corresponde al criterio 5.
5. Vendedor contiene ``Fabio Arango``:
   ``No gestión``. Corresponde al criterio 4.
6. Existe estado asegurado Movilidad y no contiene ``Activo``:
   ``No gestión``. Corresponde al criterio 6.
7. Líder Comercial Movilidad contiene ``Angelina``:
   ``Autogestión cliente``. Corresponde al criterio 3.
8. Estado asegurado Movilidad contiene ``Activo``:
   ``Gestión A&S``. Corresponde al criterio 2.
9. Si no aplica ningún criterio:
   se deja en blanco. Corresponde al criterio 7.

Los criterios 6, 7 y 8 conservan el fondo durazno del archivo de referencia.
"""

from __future__ import annotations

import argparse
import logging
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


COLUMNAS_FUENTE = [
    "Placa",
    "Líder Comercial",
    "Analista",
    "Vendedor",
    "Empresa",
    "Gestión SOAT A&S",
    "Fecha renovación SOAT (Riesgo)",
    "Póliza - Fecha fin vigencia",
    "Estado asegurado",
    "Estado de la póliza",
    "Motivo cancelación",
    "Póliza (Póliza)",
    "Tomador (Póliza)",
    "Nombre completo",
    "Número ID",
    "Afiliado",
    "Ramo (Póliza)",
    "Aseguradora (Póliza)",
    "Correo electrónico",
    "Correo electrónico (Asegurado)",
    "Correo electrónico afiliado",
    "Correo comercial",
    "Responsable SOAT",
    "ID de registro (Asegurado)",
]

COLUMNAS_SALIDA = [
    "Placa",
    "Líder Comercial Movilidad",
    "Líder Comercial SOAT",
    "Vendedor",
    "Empresa",
    "Gestión SOAT A&S",
    "Fecha renovación SOAT (Riesgo)",
    "Póliza - Fecha fin vigencia SOAT",
    "Póliza SOAT",
    "Estado de la póliza SOAT",
    "Tomador Póliza SOAT",
    "Asegurado Póliza SOAT",
    "Ramo Movilidad",
    "Póliza Movilidad",
    "Estado de la póliza Movilidad",
    "Motivo cancelación  póliza Movilidad",
    "Estado asegurado Movilidad",
    "Tomador Póliza Movilidad",
    "Asegurado Póliza Movilidad",
    "Afiliado Póliza Movilidad",
    "Aseguradora Movilidad",
    "Correo electrónico",
    "Correo electrónico (Asegurado)",
    "Correo electrónico afiliado",
    "Correo comercial",
    "ID de registro (Asegurado) SOAT",
    "ID de registro (Asegurado) Movilidad",
    "ID de registro (Asegurado) CARGA",
]

COLUMNA_CRITERIO_INTERNO = "_criterio_gestion"
COLOR_CRITERIOS_6_7_8 = "FBE2D5"


@dataclass(frozen=True)
class Config:
    base_dir: Path
    carpeta_entrada: str = "Informe Zoho"
    carpeta_referencia: str = "Formato referencia"
    carpeta_salida: str = "Salidas"
    hoja_preferida: str = "Sheet0"
    hoja_movilidad_referencia: str = "Movilidad"
    ramo_soat: str = "SOAT"

    @property
    def dir_entrada(self) -> Path:
        return self.base_dir / self.carpeta_entrada

    @property
    def dir_referencia(self) -> Path:
        return self.base_dir / self.carpeta_referencia

    @property
    def dir_salida(self) -> Path:
        return self.base_dir / self.carpeta_salida


def texto_limpio(valor: object) -> str | None:
    if pd.isna(valor):
        return None
    texto = str(valor).strip()
    return texto or None


def clave_texto(valor: object) -> str:
    texto = texto_limpio(valor) or ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().upper()


def normalizar_placa(valor: object) -> str | None:
    texto = clave_texto(valor)
    if not texto:
        return None
    placa = re.sub(r"[^A-Z0-9]", "", texto)
    return placa or None


def normalizar_identificador(valor: object) -> str | None:
    if pd.isna(valor):
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return texto_limpio(valor)


def listar_excels(carpeta: Path) -> list[Path]:
    if not carpeta.is_dir():
        return []

    return sorted(
        (
            archivo
            for archivo in carpeta.iterdir()
            if archivo.is_file()
            and archivo.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
            and not archivo.name.startswith(("~$", "."))
        ),
        key=lambda archivo: archivo.stat().st_mtime,
        reverse=True,
    )


def obtener_excel_mas_reciente(carpeta: Path, descripcion: str) -> Path:
    archivos = listar_excels(carpeta)
    if not archivos:
        raise FileNotFoundError(
            f"No hay archivos Excel para {descripcion} en: {carpeta}"
        )
    return archivos[0]


def detectar_fila_encabezado(
    path: Path,
    hoja: str | int,
    *,
    requiere_ramo: bool = True,
    max_filas: int = 40,
) -> int:
    vista = pd.read_excel(path, sheet_name=hoja, header=None, nrows=max_filas)

    for indice, fila in vista.iterrows():
        valores = {
            clave_texto(valor)
            for valor in fila.tolist()
            if texto_limpio(valor)
        }

        tiene_placa = "PLACA" in valores
        tiene_ramo = any("RAMO" in valor for valor in valores)

        if tiene_placa and (tiene_ramo or not requiere_ramo):
            return int(indice)

    requisito = "'Placa' y 'Ramo'" if requiere_ramo else "'Placa'"
    raise ValueError(
        f"No se encontró una fila de encabezados con {requisito} "
        f"en la hoja {hoja!r} de {path.name}."
    )


def resolver_hoja(path: Path, hoja_preferida: str) -> str | int:
    with pd.ExcelFile(path) as excel:
        nombres = list(excel.sheet_names)
    if hoja_preferida in nombres:
        return hoja_preferida
    return nombres[0]


def homologar_columnas(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "MOTIVO CANCELACION POLIZA MOVILIDAD": "Motivo cancelación",
        "MOTIVO CANCELACION": "Motivo cancelación",
        "MOTIVO DE CANCELACION": "Motivo cancelación",
        "ID DE REGISTRO (ASEGURADO)": "ID de registro (Asegurado)",
    }

    renombrar: dict[str, str] = {}
    existentes = {clave_texto(columna) for columna in df.columns}

    for columna in df.columns:
        destino = aliases.get(clave_texto(columna))
        if destino and clave_texto(destino) not in existentes:
            renombrar[columna] = destino

    return df.rename(columns=renombrar)


def preparar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = homologar_columnas(df.copy())

    for columna in COLUMNAS_FUENTE:
        if columna not in df.columns:
            df[columna] = None

    df = df[COLUMNAS_FUENTE].copy()
    df["_orden_origen"] = range(len(df))

    df["Placa"] = df["Placa"].map(normalizar_placa)
    df = df[df["Placa"].notna()].copy()

    for columna in (
        "Fecha renovación SOAT (Riesgo)",
        "Póliza - Fecha fin vigencia",
    ):
        df[columna] = pd.to_datetime(df[columna], errors="coerce")

    for columna in (
        "Póliza (Póliza)",
        "Número ID",
        "ID de registro (Asegurado)",
    ):
        df[columna] = df[columna].map(normalizar_identificador)

    return df.reset_index(drop=True)


def leer_fuente(path: Path, hoja_preferida: str) -> pd.DataFrame:
    hoja = resolver_hoja(path, hoja_preferida)
    header = detectar_fila_encabezado(path, hoja, requiere_ramo=True)

    logger.info(
        "Leyendo fuente %s | hoja=%s | encabezado Excel=%d",
        path.name,
        hoja,
        header + 1,
    )

    df = pd.read_excel(
        path,
        sheet_name=hoja,
        header=header,
        dtype=object,
    )
    df.columns = [str(columna).strip() for columna in df.columns]
    return preparar_dataframe(df)


def leer_movilidad_referencia(
    path: Path,
    hoja_movilidad: str,
) -> pd.DataFrame:
    """Lee la hoja Movilidad usada para complementar motivo de cancelación."""
    with pd.ExcelFile(path) as excel:
        nombres = list(excel.sheet_names)

    if hoja_movilidad not in nombres:
        raise ValueError(
            f"El archivo de referencia {path.name} no contiene la hoja "
            f"{hoja_movilidad!r}."
        )

    header = detectar_fila_encabezado(
        path,
        hoja_movilidad,
        requiere_ramo=True,
        max_filas=15,
    )

    logger.info(
        "Leyendo referencia %s | hoja=%s | encabezado Excel=%d",
        path.name,
        hoja_movilidad,
        header + 1,
    )

    df = pd.read_excel(
        path,
        sheet_name=hoja_movilidad,
        header=header,
        dtype=object,
    )
    df.columns = [str(columna).strip() for columna in df.columns]
    return preparar_dataframe(df)


def enriquecer_motivo_cancelacion(
    fuente: pd.DataFrame,
    movilidad_referencia: pd.DataFrame | None,
) -> pd.DataFrame:
    """Complementa el motivo sin mezclar pólizas o asegurados distintos.

    Prioridad de cruce:
    1. ID de registro (Asegurado).
    2. Placa + póliza.
    3. Placa, solo cuando en la referencia existe una única fila para esa placa.

    Los valores existentes en la fuente nunca se reemplazan.
    """
    if movilidad_referencia is None or movilidad_referencia.empty:
        return fuente.copy()

    resultado = fuente.copy()
    referencia = movilidad_referencia.copy()

    referencia = referencia[
        referencia["Motivo cancelación"].map(texto_limpio).notna()
    ].copy()

    if referencia.empty:
        logger.warning(
            "La hoja Movilidad de referencia no contiene motivos de cancelación."
        )
        return resultado

    # 1. Índice exacto por ID.
    por_id: dict[str, object] = {}
    for _, fila in referencia.iterrows():
        identificador = normalizar_identificador(
            fila["ID de registro (Asegurado)"]
        )
        if identificador and identificador not in por_id:
            por_id[identificador] = fila["Motivo cancelación"]

    # 2. Índice exacto por placa y póliza.
    por_placa_poliza: dict[tuple[str, str], object] = {}
    for _, fila in referencia.iterrows():
        placa = normalizar_placa(fila["Placa"])
        poliza = normalizar_identificador(fila["Póliza (Póliza)"])
        if placa and poliza:
            clave = (placa, poliza)
            if clave not in por_placa_poliza:
                por_placa_poliza[clave] = fila["Motivo cancelación"]

    # 3. Índice por placa únicamente si la referencia es inequívoca.
    motivos_por_placa: dict[str, list[object]] = {}
    for _, fila in referencia.iterrows():
        placa = normalizar_placa(fila["Placa"])
        if placa:
            motivos_por_placa.setdefault(placa, []).append(
                fila["Motivo cancelación"]
            )

    por_placa_unica: dict[str, object] = {}
    for placa, motivos in motivos_por_placa.items():
        valores = []
        for motivo in motivos:
            clave = clave_texto(motivo)
            if clave and clave not in {clave_texto(v) for v in valores}:
                valores.append(motivo)
        if len(valores) == 1:
            por_placa_unica[placa] = valores[0]

    completados_id = 0
    completados_poliza = 0
    completados_placa = 0

    for indice, fila in resultado.iterrows():
        if texto_limpio(fila["Motivo cancelación"]) is not None:
            continue

        identificador = normalizar_identificador(
            fila["ID de registro (Asegurado)"]
        )
        placa = normalizar_placa(fila["Placa"])
        poliza = normalizar_identificador(fila["Póliza (Póliza)"])

        motivo = None

        if identificador and identificador in por_id:
            motivo = por_id[identificador]
            completados_id += 1
        elif placa and poliza and (placa, poliza) in por_placa_poliza:
            motivo = por_placa_poliza[(placa, poliza)]
            completados_poliza += 1
        elif placa and placa in por_placa_unica:
            motivo = por_placa_unica[placa]
            completados_placa += 1

        if texto_limpio(motivo) is not None:
            resultado.at[indice, "Motivo cancelación"] = motivo

    logger.info(
        "Motivos complementados: por ID=%d | placa+póliza=%d | placa única=%d",
        completados_id,
        completados_poliza,
        completados_placa,
    )

    return resultado


def seleccionar_por_placa(df: pd.DataFrame) -> pd.DataFrame:
    """Selecciona una fila completa por placa mediante prioridad explícita."""
    if df.empty:
        return df.copy()

    trabajo = df.copy()

    trabajo["_poliza_vigente"] = (
        trabajo["Estado de la póliza"].map(clave_texto) == "VIGENTE"
    ).astype(int)

    trabajo["_asegurado_activo"] = trabajo["Estado asegurado"].map(
        lambda valor: int("ACTIVO" in clave_texto(valor))
    )

    trabajo = trabajo.sort_values(
        by=[
            "Placa",
            "_poliza_vigente",
            "_asegurado_activo",
            "Póliza - Fecha fin vigencia",
            "_orden_origen",
        ],
        ascending=[True, False, False, False, False],
        na_position="last",
        kind="stable",
    )

    return (
        trabajo.drop_duplicates("Placa", keep="first")
        .drop(columns=["_poliza_vigente", "_asegurado_activo"])
        .reset_index(drop=True)
    )


def primer_no_vacio(*valores: object) -> object:
    for valor in valores:
        if texto_limpio(valor) is not None:
            return valor
    return None


def obtener(fila: pd.Series | None, columna: str) -> object:
    if fila is None or columna not in fila.index:
        return None

    valor = fila[columna]
    return None if pd.isna(valor) else valor


def determinar_gestion(
    vendedor: object,
    lider_movilidad: object,
    estado_movilidad: object,
    motivo_cancelacion: object,
    estado_poliza_movilidad: object,
) -> tuple[str | None, int]:
    """Retorna la gestión y el número del criterio formal aplicado."""
    vendedor_key = clave_texto(vendedor)
    lider_key = clave_texto(lider_movilidad)
    estado_key = clave_texto(estado_movilidad)
    motivo_key = clave_texto(motivo_cancelacion)
    estado_poliza_key = clave_texto(estado_poliza_movilidad)

    # Criterio 8: única excepción al criterio 1 (que es inamovible en todo lo demás).
    # Se deja en blanco (solo color, sin texto), igual que el criterio 7.
    if "FONCONSTRUIMOS" in vendedor_key and estado_key == "EXCLUIDO":
        return None, 8

    # Criterio 9: la póliza de Movilidad vigente no gestiona asegurados excluidos.
    if estado_poliza_key == "VIGENTE" and estado_key == "EXCLUIDO":
        return "No gestión", 9

    # Criterio 1: inamovible salvo los criterios 8 y 9.
    if "FONCONSTRUIMOS" in vendedor_key:
        return "Gestión A&S", 1

    # Criterio 5: reemplaza el resultado de Angelina.
    if (
        estado_key == "CANCELADO"
        and motivo_key in {
            "POR VENTA",
            "POR CAMBIO DE INTERMEDIARIO",
        }
    ):
        return "No gestión", 5

    # Criterio 4: puede reemplazar Gestión A&S del criterio 2.
    if "FABIO ARANGO" in vendedor_key:
        return "No gestión", 4

    # Criterio 6: aplica solo cuando realmente existe un estado.
    if estado_key and "ACTIVO" not in estado_key:
        return "No gestión", 6

    # Criterio 3: puede reemplazar Gestión A&S del criterio 2.
    if "ANGELINA" in lider_key:
        return "Autogestión cliente", 3

    # Criterio 2.
    if "ACTIVO" in estado_key:
        return "Gestión A&S", 2

    # Criterio 7.
    return None, 7


def construir_formato(
    fuente: pd.DataFrame,
    soat: pd.DataFrame,
    movilidad: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    placas = sorted(set(fuente["Placa"].dropna().tolist()))

    fuente_por_placa = {
        placa: grupo.sort_values("_orden_origen", kind="stable")
        for placa, grupo in fuente.groupby("Placa", sort=False)
    }

    soat_idx = (
        soat.set_index("Placa", drop=False)
        if not soat.empty
        else pd.DataFrame()
    )

    movilidad_idx = (
        movilidad.set_index("Placa", drop=False)
        if not movilidad.empty
        else pd.DataFrame()
    )

    filas: list[dict[str, object]] = []
    auditoria: list[dict[str, object]] = []

    for placa in placas:
        fila_soat = (
            soat_idx.loc[placa]
            if not soat.empty and placa in soat_idx.index
            else None
        )

        fila_mov = (
            movilidad_idx.loc[placa]
            if not movilidad.empty and placa in movilidad_idx.index
            else None
        )

        if isinstance(fila_soat, pd.DataFrame):
            fila_soat = fila_soat.iloc[0]

        if isinstance(fila_mov, pd.DataFrame):
            fila_mov = fila_mov.iloc[0]

        grupo = fuente_por_placa[placa]

        # Mapa del Formato informe:
        # - Vendedor, Empresa y fecha de renovación: SOAT y Movilidad.
        # - Campos SOAT: fila SOAT seleccionada.
        # - Campos Movilidad: fila Movilidad seleccionada.
        vendedor = primer_no_vacio(
            obtener(fila_soat, "Vendedor"),
            obtener(fila_mov, "Vendedor"),
        )

        empresa = primer_no_vacio(
            obtener(fila_soat, "Empresa"),
            obtener(fila_mov, "Empresa"),
        )

        fecha_renovacion = primer_no_vacio(
            obtener(fila_soat, "Fecha renovación SOAT (Riesgo)"),
            obtener(fila_mov, "Fecha renovación SOAT (Riesgo)"),
        )

        lider_movilidad = obtener(fila_mov, "Líder Comercial")
        estado_movilidad = obtener(fila_mov, "Estado asegurado")
        motivo_cancelacion = obtener(fila_mov, "Motivo cancelación")
        estado_poliza_movilidad = obtener(fila_mov, "Estado de la póliza")

        gestion, criterio = determinar_gestion(
            vendedor=vendedor,
            lider_movilidad=lider_movilidad,
            estado_movilidad=estado_movilidad,
            motivo_cancelacion=motivo_cancelacion,
            estado_poliza_movilidad=estado_poliza_movilidad,
        )

        id_soat = obtener(
            fila_soat,
            "ID de registro (Asegurado)",
        )
        id_movilidad = obtener(
            fila_mov,
            "ID de registro (Asegurado)",
        )
        id_carga = primer_no_vacio(id_soat, id_movilidad)

        fila_salida = {
            "Placa": placa,
            "Líder Comercial Movilidad": lider_movilidad,
            "Líder Comercial SOAT": obtener(
                fila_soat,
                "Líder Comercial",
            ),
            "Vendedor": vendedor,
            "Empresa": empresa,
            "Gestión SOAT A&S": gestion,
            "Fecha renovación SOAT (Riesgo)": fecha_renovacion,
            "Póliza - Fecha fin vigencia SOAT": obtener(
                fila_soat,
                "Póliza - Fecha fin vigencia",
            ),
            "Póliza SOAT": obtener(
                fila_soat,
                "Póliza (Póliza)",
            ),
            "Estado de la póliza SOAT": obtener(
                fila_soat,
                "Estado de la póliza",
            ),
            "Tomador Póliza SOAT": obtener(
                fila_soat,
                "Tomador (Póliza)",
            ),
            "Asegurado Póliza SOAT": obtener(
                fila_soat,
                "Nombre completo",
            ),
            "Ramo Movilidad": obtener(
                fila_mov,
                "Ramo (Póliza)",
            ),
            "Póliza Movilidad": obtener(
                fila_mov,
                "Póliza (Póliza)",
            ),
            "Estado de la póliza Movilidad": obtener(
                fila_mov,
                "Estado de la póliza",
            ),
            "Motivo cancelación  póliza Movilidad": motivo_cancelacion,
            "Estado asegurado Movilidad": estado_movilidad,
            "Tomador Póliza Movilidad": obtener(
                fila_mov,
                "Tomador (Póliza)",
            ),
            "Asegurado Póliza Movilidad": obtener(
                fila_mov,
                "Nombre completo",
            ),
            "Afiliado Póliza Movilidad": obtener(
                fila_mov,
                "Afiliado",
            ),
            "Aseguradora Movilidad": obtener(
                fila_mov,
                "Aseguradora (Póliza)",
            ),
            "Correo electrónico": primer_no_vacio(
                obtener(fila_soat, "Correo electrónico"),
                obtener(fila_mov, "Correo electrónico"),
            ),
            "Correo electrónico (Asegurado)": primer_no_vacio(
                obtener(fila_soat, "Correo electrónico (Asegurado)"),
                obtener(fila_mov, "Correo electrónico (Asegurado)"),
            ),
            "Correo electrónico afiliado": primer_no_vacio(
                obtener(fila_soat, "Correo electrónico afiliado"),
                obtener(fila_mov, "Correo electrónico afiliado"),
            ),
            "Correo comercial": primer_no_vacio(
                obtener(fila_soat, "Correo comercial"),
                obtener(fila_mov, "Correo comercial"),
            ),
            "ID de registro (Asegurado) SOAT": id_soat,
            "ID de registro (Asegurado) Movilidad": id_movilidad,
            "ID de registro (Asegurado) CARGA": id_carga,
            COLUMNA_CRITERIO_INTERNO: criterio,
        }

        filas.append(fila_salida)

        candidatos_soat = grupo[
            grupo["Ramo (Póliza)"].map(clave_texto) == "SOAT"
        ]
        candidatos_movilidad = grupo[
            grupo["Ramo (Póliza)"].map(clave_texto) != "SOAT"
        ]

        auditoria.append(
            {
                "Placa": placa,
                "Cantidad registros fuente": len(grupo),
                "Candidatos SOAT": len(candidatos_soat),
                "Candidatos Movilidad": len(candidatos_movilidad),
                "Criterio de gestión aplicado": criterio,
                "Gestión asignada": gestion,
                "Vendedor evaluado": vendedor,
                "Líder Movilidad evaluado": lider_movilidad,
                "Estado asegurado Movilidad evaluado": estado_movilidad,
                "Motivo cancelación evaluado": motivo_cancelacion,
                "Póliza SOAT seleccionada": obtener(
                    fila_soat,
                    "Póliza (Póliza)",
                ),
                "Póliza Movilidad seleccionada": obtener(
                    fila_mov,
                    "Póliza (Póliza)",
                ),
                "ID SOAT seleccionado": id_soat,
                "ID Movilidad seleccionado": id_movilidad,
                "ID CARGA": id_carga,
                "Requiere revisión por múltiples registros": (
                    "Sí"
                    if len(candidatos_soat) > 1
                    or len(candidatos_movilidad) > 1
                    else "No"
                ),
            }
        )

    formato = pd.DataFrame(filas)
    trazabilidad = pd.DataFrame(auditoria)
    return formato, trazabilidad


def aplicar_formato_excel(
    path: Path,
    criterios_por_placa: dict[str, int],
) -> None:
    wb = load_workbook(path)

    azul = "17365D"
    azul_claro = "D9EAF7"
    blanco = "FFFFFF"
    amarillo = "FFF2CC"
    borde = Side(style="thin", color="B7B7B7")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=azul)
            cell.font = Font(color=blanco, bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = Border(bottom=borde)

        ws.row_dimensions[1].height = 42

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=False,
                )

                if isinstance(cell.value, datetime):
                    cell.number_format = "dd/mm/yyyy"

        for column_cells in ws.columns:
            letra = get_column_letter(column_cells[0].column)
            encabezado = str(column_cells[0].value or "")

            if "Correo" in encabezado:
                ancho = 30
            elif any(
                texto in encabezado
                for texto in (
                    "Asegurado",
                    "Tomador",
                    "Líder",
                    "Motivo",
                )
            ):
                ancho = 28
            elif "Fecha" in encabezado:
                ancho = 18
            elif "ID" in encabezado or "Póliza" in encabezado:
                ancho = 24
            elif encabezado == "Placa":
                ancho = 12
            else:
                ancho = 20

            ws.column_dimensions[letra].width = ancho

    if "Formato informe" in wb.sheetnames:
        ws = wb["Formato informe"]
        encabezados = {
            cell.value: cell.column
            for cell in ws[1]
        }

        col_placa = encabezados["Placa"]
        col_gestion = encabezados["Gestión SOAT A&S"]
        col_id_carga = encabezados.get(
            "ID de registro (Asegurado) CARGA"
        )

        for fila_excel in range(2, ws.max_row + 1):
            placa = normalizar_placa(
                ws.cell(fila_excel, col_placa).value
            )
            criterio = criterios_por_placa.get(placa or "")
            celda_gestion = ws.cell(
                fila_excel,
                col_gestion,
            )

            if criterio in {6, 7, 8}:
                celda_gestion.fill = PatternFill(
                    "solid",
                    fgColor=COLOR_CRITERIOS_6_7_8,
                )
            else:
                celda_gestion.fill = PatternFill(
                    "solid",
                    fgColor=blanco,
                )

        if col_id_carga:
            for fila_excel in range(2, ws.max_row + 1):
                ws.cell(
                    fila_excel,
                    col_id_carga,
                ).fill = PatternFill(
                    "solid",
                    fgColor=azul_claro,
                )

    if "Trazabilidad" in wb.sheetnames:
        ws = wb["Trazabilidad"]
        encabezados = {
            cell.value: cell.column
            for cell in ws[1]
        }

        col_revision = encabezados.get(
            "Requiere revisión por múltiples registros"
        )

        if col_revision:
            for fila_excel in range(2, ws.max_row + 1):
                celda = ws.cell(
                    fila_excel,
                    col_revision,
                )

                if celda.value == "Sí":
                    celda.fill = PatternFill(
                        "solid",
                        fgColor=amarillo,
                    )
                    celda.font = Font(bold=True)

    wb.save(path)


def exportar(
    salida: Path,
    formato: pd.DataFrame,
    fuente: pd.DataFrame,
    soat: pd.DataFrame,
    movilidad: pd.DataFrame,
    trazabilidad: pd.DataFrame,
    hojas: tuple[str, ...] | None = None,
) -> None:
    salida.parent.mkdir(parents=True, exist_ok=True)
    hojas = hojas or (
        "Formato informe", "Trazabilidad", "SOAT seleccionados",
        "Movilidad seleccionada", "Fuente Zoho",
    )

    criterios_por_placa = {
        str(fila["Placa"]): int(fila[COLUMNA_CRITERIO_INTERNO])
        for _, fila in formato.iterrows()
    }

    formato_exportable = formato[COLUMNAS_SALIDA].copy()
    fuente_exportable = fuente.drop(
        columns=["_orden_origen"],
        errors="ignore",
    )
    soat_exportable = soat.drop(
        columns=["_orden_origen"],
        errors="ignore",
    )
    movilidad_exportable = movilidad.drop(
        columns=["_orden_origen"],
        errors="ignore",
    )

    with pd.ExcelWriter(
        salida,
        engine="openpyxl",
    ) as writer:
        exportables = {
            "Formato informe": formato_exportable,
            "Trazabilidad": trazabilidad,
            "SOAT seleccionados": soat_exportable,
            "Movilidad seleccionada": movilidad_exportable,
            "Fuente Zoho": fuente_exportable,
        }
        for hoja in hojas:
            exportables[hoja].to_excel(writer, sheet_name=hoja, index=False)

    aplicar_formato_excel(
        salida,
        criterios_por_placa,
    )


def validar_politicas(formato: pd.DataFrame) -> None:
    errores: list[str] = []

    for _, fila in formato.iterrows():
        esperado, criterio = determinar_gestion(
            vendedor=fila["Vendedor"],
            lider_movilidad=fila["Líder Comercial Movilidad"],
            estado_movilidad=fila["Estado asegurado Movilidad"],
            motivo_cancelacion=fila[
                "Motivo cancelación  póliza Movilidad"
            ],
            estado_poliza_movilidad=fila["Estado de la póliza Movilidad"],
        )

        actual = texto_limpio(
            fila["Gestión SOAT A&S"]
        )
        esperado_limpio = texto_limpio(esperado)

        if actual != esperado_limpio:
            errores.append(
                f"{fila['Placa']}: esperado={esperado!r}, "
                f"actual={fila['Gestión SOAT A&S']!r}, "
                f"criterio={criterio}"
            )

    if errores:
        raise ValueError(
            "Se encontraron incumplimientos de políticas:\n"
            + "\n".join(errores[:20])
        )


def validar_resultado(
    formato: pd.DataFrame,
    fuente: pd.DataFrame,
    soat: pd.DataFrame,
    movilidad: pd.DataFrame,
) -> None:
    if formato.empty:
        raise ValueError("El informe resultó vacío.")

    if formato["Placa"].duplicated().any():
        duplicadas = formato.loc[
            formato["Placa"].duplicated(),
            "Placa",
        ].tolist()

        raise ValueError(
            "El resultado contiene placas duplicadas: "
            f"{duplicadas[:10]}"
        )

    universo = set(fuente["Placa"].dropna())
    resultado = set(formato["Placa"].dropna())

    if universo != resultado:
        faltantes = sorted(universo - resultado)
        adicionales = sorted(resultado - universo)

        raise ValueError(
            "El universo de placas no coincide. "
            f"Faltantes={faltantes[:10]} | "
            f"adicionales={adicionales[:10]}"
        )

    if soat["Placa"].duplicated().any():
        raise ValueError(
            "La tabla SOAT seleccionada aún contiene placas duplicadas."
        )

    if movilidad["Placa"].duplicated().any():
        raise ValueError(
            "La tabla Movilidad seleccionada aún contiene placas duplicadas."
        )

    esperado_id_carga = formato[
        "ID de registro (Asegurado) SOAT"
    ].combine_first(
        formato["ID de registro (Asegurado) Movilidad"]
    )

    actual_id_carga = formato[
        "ID de registro (Asegurado) CARGA"
    ]

    if not esperado_id_carga.fillna("").equals(
        actual_id_carga.fillna("")
    ):
        raise ValueError(
            "La regla de ID CARGA no se aplicó correctamente."
        )

    validar_politicas(formato)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Genera el formato consolidado SOAT con selección priorizada, "
            "políticas de asignación e ID de carga."
        )
    )

    parser.add_argument(
        "--base-dir",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Raíz del proyecto. Por defecto, la carpeta del script.",
    )

    parser.add_argument(
        "--entrada",
        type=Path,
        default=None,
        help=(
            "Excel exportado desde Zoho. Si se omite, usa el más reciente "
            "de 'Informe Zoho'."
        ),
    )

    parser.add_argument(
        "--referencia",
        type=Path,
        default=None,
        help=(
            "Libro que contiene la hoja Movilidad con Motivo cancelación. "
            "Si se omite, usa el más reciente de 'Formato referencia'. "
            "Es opcional si la fuente ya contiene el motivo."
        ),
    )

    parser.add_argument(
        "--salida",
        type=Path,
        default=None,
        help=(
            "Excel de salida. Si se omite, se genera en 'Salidas'."
        ),
    )

    parser.add_argument(
        "--hoja",
        default="Sheet0",
        help=(
            "Hoja preferida del reporte de Zoho. "
            "Default: %(default)s"
        ),
    )

    parser.add_argument(
        "--hoja-movilidad-referencia",
        default="Movilidad",
        help=(
            "Hoja del libro de referencia usada para complementar "
            "Motivo cancelación. Default: %(default)s"
        ),
    )

    parser.add_argument(
        "-v",
        "--verbose",
        action="store_true",
    )

    return parser.parse_args()


def resolver_ruta(
    ruta: Path | None,
    base_dir: Path,
) -> Path | None:
    if ruta is None:
        return None

    if ruta.is_absolute():
        return ruta

    return base_dir / ruta


def main() -> None:
    args = parse_args()

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(levelname)s: %(message)s",
    )

    base_dir = args.base_dir.resolve()

    cfg = Config(
        base_dir=base_dir,
        hoja_preferida=args.hoja,
        hoja_movilidad_referencia=args.hoja_movilidad_referencia,
    )

    entrada = resolver_ruta(
        args.entrada,
        base_dir,
    )

    if entrada is None:
        entrada = obtener_excel_mas_reciente(
            cfg.dir_entrada,
            "la fuente Zoho",
        )

    if not entrada.is_file():
        raise FileNotFoundError(
            f"No existe el archivo de entrada: {entrada}"
        )

    referencia = resolver_ruta(
        args.referencia,
        base_dir,
    )

    if referencia is None:
        referencias = listar_excels(
            cfg.dir_referencia
        )
        referencia = referencias[0] if referencias else None

    if referencia is not None and not referencia.is_file():
        raise FileNotFoundError(
            f"No existe el archivo de referencia: {referencia}"
        )

    salida = resolver_ruta(
        args.salida,
        base_dir,
    )

    if salida is None:
        sello = datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )
        salida = (
            cfg.dir_salida
            / f"Informe_SOAT_Analistas_{sello}.xlsx"
        )

    fuente = leer_fuente(
        entrada,
        cfg.hoja_preferida,
    )

    referencia_movilidad = None

    if referencia is not None:
        referencia_movilidad = leer_movilidad_referencia(
            referencia,
            cfg.hoja_movilidad_referencia,
        )

    fuente = enriquecer_motivo_cancelacion(
        fuente,
        referencia_movilidad,
    )

    es_soat = (
        fuente["Ramo (Póliza)"].map(clave_texto)
        == cfg.ramo_soat
    )

    base_soat = fuente[es_soat].copy()
    base_movilidad = fuente[~es_soat].copy()

    soat = seleccionar_por_placa(
        base_soat
    )
    movilidad = seleccionar_por_placa(
        base_movilidad
    )

    formato, trazabilidad = construir_formato(
        fuente,
        soat,
        movilidad,
    )

    validar_resultado(
        formato,
        fuente,
        soat,
        movilidad,
    )

    exportar(
        salida,
        formato,
        fuente,
        soat,
        movilidad,
        trazabilidad,
    )

    conteo_gestion = (
        formato["Gestión SOAT A&S"]
        .fillna("En blanco")
        .value_counts()
        .to_dict()
    )

    conteo_criterios = (
        formato[COLUMNA_CRITERIO_INTERNO]
        .value_counts()
        .sort_index()
        .to_dict()
    )

    multiples = int(
        (
            trazabilidad[
                "Requiere revisión por múltiples registros"
            ]
            == "Sí"
        ).sum()
    )

    motivos_movilidad = int(
        movilidad["Motivo cancelación"]
        .map(texto_limpio)
        .notna()
        .sum()
    )

    print("Proceso finalizado correctamente")
    print(f"Entrada: {entrada}")
    print(
        "Referencia: "
        f"{referencia if referencia is not None else 'No utilizada'}"
    )
    print(f"Registros fuente: {len(fuente)}")
    print(f"Placas únicas: {len(formato)}")
    print(
        "Placas con SOAT: "
        f"{formato['ID de registro (Asegurado) SOAT'].notna().sum()}"
    )
    print(
        "Placas con Movilidad: "
        f"{formato['ID de registro (Asegurado) Movilidad'].notna().sum()}"
    )
    print(
        "Movilidades seleccionadas con motivo de cancelación: "
        f"{motivos_movilidad}"
    )
    print(
        "Placas con múltiples candidatos: "
        f"{multiples}"
    )
    print(f"Gestiones: {conteo_gestion}")
    print(f"Criterios aplicados: {conteo_criterios}")
    print(f"Salida: {salida}")


if __name__ == "__main__":
    main()
