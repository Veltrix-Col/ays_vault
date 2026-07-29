from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from time import monotonic

import pandas as pd
from django.conf import settings
from openpyxl import load_workbook

from . import legacy_processor as legacy


class SoatProcessingError(ValueError):
    pass


@dataclass(frozen=True)
class SoatResult:
    content: bytes
    filename: str
    summary: dict[str, object]


def _neutralize(value):
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _sanitize_frame(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.map(_neutralize)


def _validate_dimensions(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        for sheet in workbook.worksheets:
            if sheet.max_row > settings.SOAT_MAX_ROWS:
                raise SoatProcessingError("El archivo supera el límite de filas permitido.")
            if sheet.max_column > settings.SOAT_MAX_COLUMNS:
                raise SoatProcessingError("El archivo supera el límite de columnas permitido.")
    finally:
        workbook.close()


def _remove_hyperlinks(path: Path) -> None:
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.hyperlink = None
    workbook.save(path)


def process_soat(uploaded_file) -> SoatResult:
    started = monotonic()

    with TemporaryDirectory(prefix="ays-soat-") as directory:
        root = Path(directory)
        source = root / "source.xlsx"
        output = root / "result.xlsx"
        with source.open("wb") as stream:
            for chunk in uploaded_file.chunks():
                stream.write(chunk)

        _validate_dimensions(source)
        try:
            fuente = legacy.leer_fuente(source, "Sheet0")
            es_soat = fuente["Ramo (Póliza)"].map(legacy.clave_texto) == "SOAT"
            soat = legacy.seleccionar_por_placa(fuente[es_soat].copy())
            movilidad = legacy.seleccionar_por_placa(fuente[~es_soat].copy())
            formato, trazabilidad = legacy.construir_formato(fuente, soat, movilidad)
            legacy.validar_resultado(formato, fuente, soat, movilidad)
        except (KeyError, ValueError, OSError) as exc:
            raise SoatProcessingError(str(exc)) from exc

        formato = _sanitize_frame(formato)
        trazabilidad = _sanitize_frame(trazabilidad)
        fuente = _sanitize_frame(fuente)
        soat = _sanitize_frame(soat)
        movilidad = _sanitize_frame(movilidad)
        legacy.exportar(output, formato, fuente, soat, movilidad, trazabilidad)
        _remove_hyperlinks(output)

        criteria = formato[legacy.COLUMNA_CRITERIO_INTERNO].value_counts().sort_index().to_dict()
        management = formato["Gestión SOAT A&S"].fillna("En blanco").value_counts().to_dict()
        summary = {
            "source_records": int(len(fuente)),
            "unique_plates": int(len(formato)),
            "with_soat": int(formato["ID de registro (Asegurado) SOAT"].notna().sum()),
            "with_mobility": int(formato["ID de registro (Asegurado) Movilidad"].notna().sum()),
            "mobility_with_reason": int(movilidad["Motivo cancelación"].map(legacy.texto_limpio).notna().sum()),
            "multiple_candidates": int((trazabilidad["Requiere revisión por múltiples registros"] == "Sí").sum()),
            "management": {str(key): int(value) for key, value in management.items()},
            "criteria": {str(key): int(value) for key, value in criteria.items()},
            "duration_seconds": round(monotonic() - started, 2),
        }
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return SoatResult(
            content=output.read_bytes(),
            filename=f"Informe_SOAT_A&S_{stamp}.xlsx",
            summary=summary,
        )
