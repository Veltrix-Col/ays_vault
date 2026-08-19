from concurrent.futures import ThreadPoolExecutor
from email import policy
from email.parser import BytesParser
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook

from vault.models import AuditEvent

from .services import process_soat
from .services import legacy_processor as legacy


def workbook_bytes(rows, sheet_name="Sheet0"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def multipart_files(response):
    raw = (
        f"Content-Type: {response['Content-Type']}\r\n"
        "MIME-Version: 1.0\r\n\r\n"
    ).encode() + response.content
    message = BytesParser(policy=policy.default).parsebytes(raw)
    return {
        part.get_param("name", header="content-disposition"): (
            part.get_filename(), part.get_content_type(), part.get_payload(decode=True)
        )
        for part in message.iter_attachments()
    }


HEADERS = [
    "Placa", "Líder Comercial", "Analista", "Vendedor", "Empresa",
    "Gestión SOAT A&S", "Fecha renovación SOAT (Riesgo)",
    "Póliza - Fecha fin vigencia", "Estado asegurado", "Estado de la póliza",
    "Motivo cancelación", "Póliza (Póliza)", "Tomador (Póliza)",
    "Nombre completo", "Número ID", "Afiliado", "Ramo (Póliza)",
    "Aseguradora (Póliza)", "Correo electrónico",
    "Correo electrónico (Asegurado)", "Correo electrónico afiliado",
    "Correo comercial", "Responsable SOAT", "ID de registro (Asegurado)",
]


class SoatPublicTests(TestCase):
    def setUp(self):
        self.source = workbook_bytes([
            ["Reporte Zoho"],
            HEADERS,
            ["ABC123", "Laura", "", "Fonconstruimos", "A&S", "", "", "2030-12-31", "Activo", "Vigente", "", "SOAT-1", "", "", "", "", "SOAT", "", "", "", "", "", "", "100"],
            ["ABC123", "Angelina", "", "", "A&S", "", "", "2030-01-01", "Activo", "Vigente", "", "MOV-1", "", "", "", "", "Movilidad", "", "", "", "", "", "", "200"],
        ])

    def upload(self, name="SOAT_prueba_4.xlsx", content=None):
        return SimpleUploadedFile(
            name,
            content if content is not None else self.source,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def test_public_page_requires_no_authentication(self):
        response = self.client.get(reverse("soat:upload"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Gestión SOAT")
        self.assertNotIn("/login/", response.url if hasattr(response, "url") else "")

    def test_any_filename_with_valid_structure_is_accepted(self):
        response = self.client.post(reverse("soat:upload"), {"source_file": self.upload("cualquier_nombre.xlsx")})
        self.assertEqual(response.status_code, 200)

    def test_missing_structure_and_invalid_excel_are_rejected(self):
        sin_estructura = workbook_bytes([["Columna A", "Columna B"], ["x", "y"]])
        missing = self.client.post(
            reverse("soat:upload"),
            {"source_file": self.upload("otro.xlsx", content=sin_estructura)},
        )
        corrupt = self.client.post(reverse("soat:upload"), {"source_file": self.upload(content=b"not-xlsx")})
        self.assertEqual(missing.status_code, 422)
        self.assertContains(missing, "estructura", status_code=422)
        self.assertEqual(corrupt.status_code, 422)
        self.assertContains(corrupt, "Excel válido", status_code=422)

    def test_valid_processing_returns_two_workbooks_without_vault_audit(self):
        before = AuditEvent.objects.count()
        response = self.client.post(reverse("soat:upload"), {"source_file": self.upload()})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["X-Content-Type-Options"], "nosniff")
        self.assertTrue(response["Content-Type"].startswith("multipart/form-data;"))
        files = multipart_files(response)
        self.assertEqual(set(files), {"report", "support"})
        report_name, report_type, report_content = files["report"]
        support_name, support_type, support_content = files["support"]
        self.assertEqual(report_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(support_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertRegex(report_name, r"^Informe_SOAT_A&S_(\d{8}_\d{6})\.xlsx$")
        self.assertRegex(support_name, r"^Soporte_SOAT_A&S_(\d{8}_\d{6})\.xlsx$")
        self.assertEqual(report_name.split("_")[-2:], support_name.split("_")[-2:])
        self.assertEqual(load_workbook(BytesIO(report_content)).sheetnames, ["Formato informe"])
        self.assertEqual(load_workbook(BytesIO(support_content)).sheetnames, [
            "Trazabilidad", "SOAT seleccionados", "Movilidad seleccionada", "Fuente Zoho",
        ])
        self.assertNotIn("zip", response["Content-Type"].lower())
        self.assertEqual(AuditEvent.objects.count(), before)

    def test_formula_values_are_neutralized(self):
        rows = [list(row) for row in [
            HEADERS,
            ["ABC123", "Laura", "", "Fonconstruimos", "A&S", "", "", "2030-12-31", "Activo", "Vigente", "", "SOAT-1", "", "", "", "", "SOAT", "", "+CMD()", "", "", "", "", "", "100"],
        ]]
        result = process_soat(self.upload(content=workbook_bytes(rows)))
        workbook = load_workbook(BytesIO(result.support_content), data_only=False)
        values = [cell.value for row in workbook["Fuente Zoho"].iter_rows() for cell in row]
        self.assertIn("'+CMD()", values)
        self.assertFalse(any(getattr(cell, "data_type", "") == "f" for row in workbook["Fuente Zoho"].iter_rows() for cell in row))

    def test_concurrent_processing_is_isolated(self):
        def run():
            return process_soat(self.upload()).report_content
        with ThreadPoolExecutor(max_workers=2) as pool:
            outputs = list(pool.map(lambda _: run(), range(2)))
        self.assertEqual(len(outputs), 2)
        self.assertTrue(all(load_workbook(BytesIO(content)).sheetnames[0] == "Formato informe" for content in outputs))

    def test_nine_management_criteria_match_delivered_script(self):
        cases = [
            (("Fonconstruimos", "", "", "", ""), ("Gestión A&S", 1)),
            (("", "", "Activo", "", ""), ("Gestión A&S", 2)),
            (("", "Angelina", "Activo", "", ""), ("Autogestión cliente", 3)),
            (("Fabio Arango", "", "Activo", "", ""), ("No gestión", 4)),
            (("", "Angelina", "Cancelado", "Por venta", ""), ("No gestión", 5)),
            (("", "", "Cancelado", "Otro", ""), ("No gestión", 6)),
            (("", "", "", "", ""), (None, 7)),
            (("Fonconstruimos", "", "Excluido", "", ""), ("No gestión", 8)),
            (("", "", "Excluido", "", "Vigente"), ("No gestión", 9)),
        ]
        for arguments, expected in cases:
            with self.subTest(criterion=expected[1]):
                self.assertEqual(legacy.determinar_gestion(*arguments), expected)
