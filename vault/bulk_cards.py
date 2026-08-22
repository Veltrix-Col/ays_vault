"""Importación efímera y atómica de tarjetas desde la plantilla oficial."""

from dataclasses import dataclass
from io import BytesIO
import re
import unicodedata

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .crypto import fingerprint
from .forms import CardForm
from .models import PaymentCard


@dataclass(frozen=True)
class BulkColumn:
    header: str
    field_name: str
    instruction: str
    safe_example: str = ""


COLUMNS = (
    BulkColumn("Empresa", "company_name", "Razón social o empresa asociada.", "Empresa Ejemplo S.A.S."),
    BulkColumn("Alias", "client_name", "Nombre corto para identificar la tarjeta.", "Operativa principal"),
    BulkColumn("Titular", "cardholder_name", "Nombre completo del titular.", "Persona de ejemplo"),
    BulkColumn("Cédula / Documento de identidad", "identity_document", "Documento del titular.", "DOC-0001"),
    BulkColumn("Correo electrónico", "email", "Correo con formato válido.", "persona@example.invalid"),
    BulkColumn("Teléfono", "phone", "Número de contacto; admite espacios, paréntesis, guiones y prefijo internacional.", "+57 300 000 0000"),
    BulkColumn("Franquicia", "brand", "Seleccione exclusivamente una franquicia soportada por CardManager."),
    BulkColumn("Referencia", "purpose", "Referencia administrativa de la tarjeta.", "Operación interna"),
    BulkColumn("Número de tarjeta", "pan", "Número que cumpla las validaciones vigentes de CardManager."),
    BulkColumn("Vencimiento", "expiry", "Formato obligatorio MM/AA.", "05/28"),
    BulkColumn("Código", "code", "Código de seguridad requerido según las reglas vigentes de CardManager."),
)
HEADERS = tuple(column.header for column in COLUMNS)
FIELD_MAP = {column.header: column.field_name for column in COLUMNS}
SUPPORTED_BRANDS = tuple(PaymentCard.BRAND)
MAX_ROWS = 5000
TEMPLATE_DATA_ROWS = 5000
_INVISIBLE_CHARACTERS = "\ufeff\u200b\u200c\u200d\u2060"


@dataclass(frozen=True)
class BulkValidationResult:
    forms: tuple
    errors: tuple


def build_template():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tarjetas"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="152036")
        sheet.column_dimensions[get_column_letter(column)].width = min(max(len(header) + 4, 16), 36)
    for column in (4, 6, 9, 10, 11):
        sheet.column_dimensions[get_column_letter(column)].number_format = "@"

    brand_labels = tuple(label for _value, label in SUPPORTED_BRANDS)
    brand_validation = DataValidation(
        type="list",
        formula1=f'"{",".join(brand_labels)}"',
        allow_blank=False,
    )
    brand_validation.error = "Seleccione una franquicia de la lista."
    brand_validation.errorTitle = "Franquicia no válida"
    brand_validation.prompt = "Opciones soportadas: " + ", ".join(brand_labels)
    brand_validation.promptTitle = "Franquicia"
    brand_validation.showErrorMessage = True
    brand_validation.showInputMessage = True
    sheet.add_data_validation(brand_validation)
    brand_column = HEADERS.index("Franquicia") + 1
    brand_validation.add(f"{get_column_letter(brand_column)}2:{get_column_letter(brand_column)}{TEMPLATE_DATA_ROWS + 1}")

    instructions = workbook.create_sheet("Instrucciones")
    instructions.freeze_panes = "A2"
    instruction_headers = ("Campo", "Obligatorio", "Instrucción", "Ejemplo seguro")
    for column_number, header in enumerate(instruction_headers, start=1):
        cell = instructions.cell(row=1, column=column_number, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="152036")
    supported_text = ", ".join(brand_labels)
    for row_number, column in enumerate(COLUMNS, start=2):
        instruction = column.instruction
        if column.header == "Franquicia":
            instruction = f"{instruction} Opciones: {supported_text}."
        instructions.append((column.header, "Sí", instruction, column.safe_example))
        instructions.cell(row=row_number, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        instructions.cell(row=row_number, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    instructions.auto_filter.ref = f"A1:D{len(COLUMNS) + 1}"
    instructions.column_dimensions["A"].width = 34
    instructions.column_dimensions["B"].width = 14
    instructions.column_dimensions["C"].width = 72
    instructions.column_dimensions["D"].width = 30

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _safe_cell_value(cell):
    if cell.data_type == "f":
        return None, "no se permiten fórmulas"
    value = cell.value
    if value is None:
        return "", None
    if isinstance(value, str):
        return value.strip(), None
    return str(value).strip(), None


def _normalize_text(value):
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.translate({ord(character): None for character in _INVISIBLE_CHARACTERS})
    text = text.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def _normalized_key(value):
    return _normalize_text(value).casefold()


def _brand_aliases():
    aliases = {}
    for value, label in SUPPORTED_BRANDS:
        for candidate in (value, label):
            normalized = _normalized_key(candidate)
            aliases[normalized] = value
            aliases[normalized.replace(" ", "")] = value
    return aliases


BRAND_ALIASES = _brand_aliases()


def normalize_brand(value):
    normalized = _normalized_key(value)
    return BRAND_ALIASES.get(normalized, BRAND_ALIASES.get(normalized.replace(" ", ""), _normalize_text(value)))


def _header_contract(cells):
    raw_headers = [cell.value for cell in cells]
    while raw_headers and _normalize_text(raw_headers[-1]) == "":
        raw_headers.pop()
    expected = {_normalized_key(column.header): column for column in COLUMNS}
    found = {}
    errors = []
    for index, raw_header in enumerate(raw_headers):
        visible = _normalize_text(raw_header)
        normalized = _normalized_key(raw_header)
        if not normalized:
            errors.append("Se encontró una columna sin encabezado.")
            continue
        if normalized in found:
            official = expected.get(normalized)
            errors.append(f"Se encontró una columna duplicada: {official.header if official else visible}.")
            continue
        found[normalized] = index
        if normalized not in expected:
            errors.append(f"Se encontró una columna no esperada: {visible[:80]}.")
    for normalized, column in expected.items():
        if normalized not in found:
            errors.append(f"Falta la columna obligatoria: {column.header}.")
    if errors:
        return {}, tuple(dict.fromkeys(errors))
    return {column.header: found[_normalized_key(column.header)] for column in COLUMNS}, ()


def validate_workbook(uploaded_file):
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=False)
    except Exception:
        return BulkValidationResult((), ("El archivo no es un libro .xlsx válido.",))
    try:
        if "Tarjetas" not in workbook.sheetnames:
            return BulkValidationResult((), ("Falta la hoja obligatoria: Tarjetas.",))
        sheet = workbook["Tarjetas"]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, max_col=sheet.max_column))
        header_positions, header_errors = _header_contract(header_cells)
        if header_errors:
            return BulkValidationResult((), header_errors)
        forms = []
        errors = []
        fingerprints = {}
        nonempty_rows = 0
        for row_number, cells in enumerate(sheet.iter_rows(min_row=2, max_col=sheet.max_column), start=2):
            raw_values = [cell.value for cell in cells]
            if all(value is None or str(value).strip() == "" for value in raw_values):
                continue
            nonempty_rows += 1
            if nonempty_rows > MAX_ROWS:
                errors.append(f"El archivo supera el máximo permitido de {MAX_ROWS} filas con datos.")
                break
            data = {"active": "on"}
            row_has_formula = False
            for header in HEADERS:
                cell = cells[header_positions[header]]
                value, error = _safe_cell_value(cell)
                if error:
                    errors.append(f"Fila {row_number} — {header}: {error}.")
                    row_has_formula = True
                if header == "Franquicia":
                    value = normalize_brand(value)
                data[FIELD_MAP[header]] = value
            if row_has_formula:
                continue
            form = CardForm(data)
            if not form.is_valid():
                for field_name, field_errors in form.errors.items():
                    label = form.fields[field_name].label if field_name in form.fields else "Fila"
                    for error in field_errors:
                        errors.append(f"Fila {row_number} — {label}: {error}")
                continue
            pan_fingerprint = fingerprint(form.cleaned_data["pan"])
            previous_row = fingerprints.get(pan_fingerprint)
            if previous_row:
                errors.append(f"Fila {row_number} — Número de tarjeta: duplicada dentro del archivo (fila {previous_row}).")
                continue
            fingerprints[pan_fingerprint] = row_number
            forms.append(form)
        if nonempty_rows == 0:
            errors.append("El archivo no contiene filas para cargar.")
        return BulkValidationResult(tuple(forms), tuple(errors))
    finally:
        workbook.close()


def create_cards_atomically(validated_forms, user):
    with transaction.atomic():
        return tuple(form.save(user=user) for form in validated_forms)
