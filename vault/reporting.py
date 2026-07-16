"""Consultas y renderizado seguro de informes de A&S Vault.

Este módulo nunca descifra tarjetas ni serializa metadatos de auditoría sin filtrar.
"""

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re

from django.conf import settings
from django.contrib.auth import get_user_model
from django.contrib.staticfiles import finders
from django.db.models import Q
from django.template.loader import render_to_string
from django.utils import timezone

from .models import (
    AccessException,
    AuditEvent,
    NotificationRecord,
    PaymentCard,
    PolicyEvaluationRun,
    SecurityAlert,
    SecureSession,
    UserDevice,
    UserProfile,
)
from .security import cached_chain_status


OPERATIONAL_ACTIONS = ["VIEW", "REVEAL", "COPY", "COPY_ATTEMPT", "CREATE", "UPDATE", "DEACTIVATE", "DENIED"]
SENSITIVE_ACTIONS = ["REVEAL", "COPY", "COPY_ATTEMPT", "CREATE", "UPDATE", "DEACTIVATE", "PASSWORD_CHANGED", "MFA_RESET", "POLICY_CHANGED"]
QUICK_ACTIONS = {
    "accesses": ["LOGIN", "LOGIN_FAILED", "ACCESS", "MFA_SUCCESS", "MFA_FAILED"],
    "reveals": ["REVEAL"],
    "copies": ["COPY", "COPY_ATTEMPT"],
    "alerts": ["ALERT_CREATED", "ALERT_REVIEWED", "ALERT_CLOSED", "ALERT_ESCALATED", "ALERT_REOPENED"],
}
REPORT_TYPES_BY_ROLE = {
    UserProfile.ADMIN: ["TIMELINE", "ALERTS", "ACCESS", "ADOPTION", "HEALTH"],
    UserProfile.LEADER: ["TIMELINE", "ALERTS", "ACCESS", "ADOPTION", "CARDS"],
    UserProfile.ANALYST: ["TIMELINE", "ALERTS", "ACCESS"],
}
ROLE_LABELS = dict(UserProfile.ROLES)
EVENT_LABELS = dict(AuditEvent.ACTIONS)
EVENT_LABELS.update({"LOGIN": "Inicio de sesión", "SESSION_REPLACED": "Sesión reemplazada", "REPORT_EXPORT": "Exportación de informe"})
ALERT_LABELS = {
    "LOGIN": "Inicio de sesión", "LOGIN_FAILED": "Inicio de sesión fallido",
    "SESSION_REPLACED": "Sesión reemplazada", "REPORT_EXPORT": "Exportación de informe",
    "OUTSIDE_HOURS": "Acceso fuera de horario", "NEW_DEVICE": "Dispositivo nuevo",
    "MFA_BLOCKED": "MFA bloqueado", "USER_BLOCKED": "Usuario bloqueado",
    "POLICY_CHANGED": "Política modificada", "EXCEPTION_CREATED": "Excepción creada",
    "EXCEPTION_REVOKED": "Excepción revocada", "SYSTEM_INACTIVITY": "Sistema sin uso",
    "POSSIBLE_PARALLEL_TOOL_USE": "Posible uso paralelo de herramienta no autorizada",
    "AUDIT_INTEGRITY_REVIEW": "Revisión de integridad de auditoría", "CRITICAL_ALERT": "Alerta crítica",
    "EMAIL_FAILURE": "Fallo de correo", "INACTIVE_USER": "Usuario inactivo",
}
RESULT_LABELS = {"SUCCESS": "Exitoso", "FAILED": "Fallido", "DENIED": "Denegado", "BLOCKED": "Bloqueado"}
SCHEDULE_LABELS = {"inside": "Dentro del horario", "outside": "Fuera del horario"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
CONFIDENTIALITY = "Uso interno y confidencial. Este informe no contiene números completos de tarjeta ni vencimientos."


class ReportValidationError(Exception):
    pass


@dataclass
class ReportData:
    title: str
    columns: list[str]
    rows: list[list]
    filters: dict
    orientation: str = "landscape"
    observations: str = "Información limitada al alcance autorizado del usuario que genera el informe."


def allowed_report_types(user):
    profile = getattr(user, "vault_profile", None)
    if not user.is_active or not profile or not profile.active:
        return []
    return REPORT_TYPES_BY_ROLE.get(profile.role, [])


def timeline_queryset(user):
    queryset = AuditEvent.objects.select_related("user", "card", "securityalert")
    role = user.vault_profile.role
    if role == UserProfile.ADMIN:
        return queryset
    if role == UserProfile.LEADER:
        return queryset.filter(action__in=OPERATIONAL_ACTIONS)
    return queryset.filter(user=user)


def apply_timeline_filters(queryset, filters):
    """Aplica únicamente filtros previamente validados por TimelineFilterForm."""
    if filters.get("date_from"):
        queryset = queryset.filter(created_at__date__gte=filters["date_from"])
    if filters.get("date_to"):
        queryset = queryset.filter(created_at__date__lte=filters["date_to"])
    if filters.get("user"):
        queryset = queryset.filter(user=filters["user"])
    if filters.get("role"):
        queryset = queryset.filter(actor_role=filters["role"])
    if filters.get("event_type"):
        queryset = queryset.filter(action=filters["event_type"])
    if filters.get("severity"):
        queryset = queryset.filter(risk_level=filters["severity"])
    if filters.get("result"):
        queryset = queryset.filter(result=filters["result"])
    if filters.get("ip"):
        queryset = queryset.filter(ip_address__startswith=filters["ip"])
    if filters.get("device"):
        queryset = queryset.filter(user_agent__icontains=filters["device"])
    if filters.get("schedule") in SCHEDULE_LABELS:
        queryset = queryset.filter(outside_office_hours=filters["schedule"] == "outside")
    card = filters.get("card", "")
    if card:
        queryset = queryset.filter(card__last4=card) if len(card) == 4 else queryset.filter(card_id=int(card))
    if filters.get("alert"):
        queryset = queryset.filter(securityalert__pk=filters["alert"])
    if filters.get("method"):
        queryset = queryset.filter(method=filters["method"])
    if filters.get("path"):
        queryset = queryset.filter(path__icontains=filters["path"])
    for key in ("device_type", "browser", "operating_system"):
        if filters.get(key):
            queryset = queryset.filter(user_agent__icontains=filters[key])
    if filters.get("session"):
        queryset = queryset.filter(session_key__startswith=filters["session"])
    if filters.get("alert_status"):
        queryset = queryset.filter(securityalert__status=filters["alert_status"])
    if filters.get("policy"):
        queryset = queryset.filter(securityalert__policy_id=filters["policy"])
    if filters.get("exception"):
        queryset = queryset.filter(securityalert__access_exception_id=filters["exception"])
    if filters.get("sensitive_only"):
        queryset = queryset.filter(action__in=SENSITIVE_ACTIONS)
    if filters.get("with_alert"):
        queryset = queryset.filter(securityalert__isnull=False)
    if filters.get("failed_only"):
        queryset = queryset.exclude(result="SUCCESS")
    if filters.get("critical_only"):
        queryset = queryset.filter(risk_level="CRITICAL")
    quick = filters.get("quick_event", "")
    if quick in QUICK_ACTIONS:
        queryset = queryset.filter(action__in=QUICK_ACTIONS[quick])
    elif quick == "outside":
        queryset = queryset.filter(outside_office_hours=True)
    elif quick == "failed":
        queryset = queryset.exclude(result="SUCCESS")
    return queryset.distinct().order_by("sequence" if filters.get("order") == "asc" else "-sequence")


def safe_filter_summary(filters):
    action_labels = dict(AuditEvent.ACTIONS)
    severity_labels = dict(AuditEvent.RISK_LEVELS)
    safe = {}
    if filters.get("date_from"):
        safe["Fecha inicial"] = filters["date_from"].strftime("%d/%m/%Y")
    if filters.get("date_to"):
        safe["Fecha final"] = filters["date_to"].strftime("%d/%m/%Y")
    if filters.get("user"):
        safe["Usuario"] = filters["user"].get_full_name() or filters["user"].username
    mappings = {
        "role": ("Rol", ROLE_LABELS), "event_type": ("Evento", action_labels),
        "severity": ("Severidad", severity_labels), "result": ("Resultado", RESULT_LABELS),
        "schedule": ("Horario", SCHEDULE_LABELS),
    }
    for key, (label, choices) in mappings.items():
        if filters.get(key):
            safe[label] = choices.get(filters[key], filters[key])
    for key, label in (("device", "Dispositivo"), ("ip", "Dirección IP"), ("card", "Tarjeta segura"), ("alert", "Alerta"), ("method", "Método HTTP")):
        if filters.get(key):
            safe[label] = str(filters[key])[:80]
    boolean_labels = {"sensitive_only": "Solo operaciones sensibles", "with_alert": "Solo eventos con alerta", "failed_only": "Solo fallidos", "critical_only": "Solo críticos"}
    for key, label in boolean_labels.items():
        if filters.get(key):
            safe[label] = "Sí"
    return safe


def filter_chips(querydict, filters):
    chips = []
    for label, value in safe_filter_summary(filters).items():
        related = {
            "Fecha inicial": "date_from", "Fecha final": "date_to", "Usuario": "user", "Rol": "role",
            "Evento": "event_type", "Severidad": "severity", "Resultado": "result", "Horario": "schedule",
            "Dispositivo": "device", "Dirección IP": "ip", "Tarjeta segura": "card", "Alerta": "alert",
            "Método HTTP": "method", "Solo operaciones sensibles": "sensitive_only", "Solo eventos con alerta": "with_alert",
            "Solo fallidos": "failed_only", "Solo críticos": "critical_only",
        }[label]
        params = querydict.copy()
        params.pop(related, None)
        if related in {"date_from", "date_to"}:
            params.pop("period", None)
        params.pop("page", None)
        chips.append({"label": label, "value": value, "remove_url": "?" + params.urlencode() if params else "?"})
    return chips


def safe_reason(value):
    text = str(value or "—")[:240]
    text = re.sub(r"(?<!\d)\d{13,19}(?!\d)", "[dato protegido]", text)
    text = re.sub(r"(?<!\d)(0[1-9]|1[0-2])[/\-]\d{2,4}(?!\d)", "[fecha protegida]", text)
    return text


def safe_device(value):
    return safe_reason(value)[:90]


def _local_parts(value):
    local = timezone.localtime(value)
    return local.strftime("%d/%m/%Y"), local.strftime("%H:%M:%S")


def _timeline_data(user, filters):
    events = apply_timeline_filters(timeline_queryset(user), filters)
    rows = []
    for event in events[: filters.get("_row_limit", settings.REPORT_XLSX_MAX_ROWS) + 1]:
        day, hour = _local_parts(event.created_at)
        alert_id = getattr(getattr(event, "securityalert", None), "pk", None)
        safe_object = f"Tarjeta #{event.card_id} · **** {event.card.last4}" if event.card else "—"
        rows.append([day, hour, event.user.get_full_name() or event.user.username if event.user else "Sistema", ROLE_LABELS.get(event.actor_role, "Sistema"), EVENT_LABELS.get(event.action, "Evento del sistema"), RESULT_LABELS.get(event.result, event.result), event.get_risk_level_display(), str(event.ip_address or "—"), safe_device(event.user_agent), "Fuera" if event.outside_office_hours else "Dentro", safe_object, f"Alerta #{alert_id}" if alert_id else "—", safe_reason(event.reason)])
    return ReportData("Informe de Línea de Tiempo", ["Fecha", "Hora", "Usuario", "Rol", "Evento", "Resultado", "Severidad", "IP", "Dispositivo", "Horario", "Objeto seguro", "Alerta relacionada", "Motivo seguro"], rows, safe_filter_summary(filters))


def _scoped_alerts(user):
    queryset = SecurityAlert.objects.select_related("event", "actor", "affected_user", "device", "policy", "access_exception", "assigned_to")
    role = user.vault_profile.role
    if role == UserProfile.ADMIN:
        return queryset
    if role == UserProfile.LEADER:
        return queryset.filter(event__action__in=OPERATIONAL_ACTIONS)
    return queryset.filter(Q(actor=user) | Q(affected_user=user) | Q(event__user=user))


def _date_filter(queryset, filters, field="created_at"):
    if filters.get("date_from"):
        queryset = queryset.filter(**{f"{field}__date__gte": filters["date_from"]})
    if filters.get("date_to"):
        queryset = queryset.filter(**{f"{field}__date__lte": filters["date_to"]})
    return queryset


def _alerts_data(user, filters):
    rows = []
    items = _date_filter(_scoped_alerts(user), filters).order_by("-created_at")
    for item in items[: filters.get("_row_limit", settings.REPORT_XLSX_MAX_ROWS) + 1]:
        rows.append([item.pk, ALERT_LABELS.get(item.alert_type, EVENT_LABELS.get(item.alert_type, "Alerta de seguridad")), item.get_severity_display(), item.get_status_display(), timezone.localtime(item.created_at).strftime("%d/%m/%Y %H:%M"), str(item.actor or "Sistema"), str(item.affected_user or "—"), str(item.ip_address or "—"), safe_device(item.device.friendly_name if item.device else "—"), f"#{item.policy_id}" if item.policy_id else "—", f"#{item.access_exception_id}" if item.access_exception_id else "—", timezone.localtime(item.due_at).strftime("%d/%m/%Y %H:%M") if item.due_at else "—", str(item.assigned_to or "—"), safe_reason(item.review_note), timezone.localtime(item.closed_at).strftime("%d/%m/%Y %H:%M") if item.closed_at else "—"])
    return ReportData("Informe de Alertas", ["ID", "Tipo", "Severidad", "Estado", "Fecha", "Actor", "Usuario afectado", "IP", "Dispositivo", "Política", "Excepción", "Fecha límite", "Responsable", "Comentario de cierre", "Fecha de cierre"], rows, safe_filter_summary(filters))


def _access_data(user, filters):
    access_filters = dict(filters)
    queryset = timeline_queryset(user).filter(action__in=QUICK_ACTIONS["accesses"] + ["SESSION_REPLACED"])
    queryset = apply_timeline_filters(queryset, access_filters)
    rows = []
    for event in queryset[: filters.get("_row_limit", settings.REPORT_XLSX_MAX_ROWS) + 1]:
        day, hour = _local_parts(event.created_at)
        rows.append([str(event.user or "Sistema"), ROLE_LABELS.get(event.actor_role, "Sistema"), day, hour, RESULT_LABELS.get(event.result, event.result), "Exitoso" if event.action == "MFA_SUCCESS" else ("Fallido" if event.action == "MFA_FAILED" else "No aplica"), str(event.ip_address or "—"), safe_device(event.user_agent), "Sí" if event.outside_office_hours else "No", "Sí" if event.action == "SESSION_REPLACED" else "No", "Sí" if hasattr(event, "securityalert") else "No"])
    return ReportData("Informe de Accesos", ["Usuario", "Rol", "Fecha", "Hora", "Resultado", "MFA", "IP", "Dispositivo", "Fuera de horario", "Sesión reemplazada", "Alerta generada"], rows, safe_filter_summary(filters))


def _scoped_users(user):
    users = get_user_model().objects.filter(is_active=True, vault_profile__active=True).select_related("vault_profile")
    return users.filter(pk=user.pk) if user.vault_profile.role == UserProfile.ANALYST else users


def _adoption_data(user, filters):
    rows = []
    today = timezone.localdate()
    for item in _scoped_users(user):
        events = _date_filter(timeline_queryset(user).filter(user=item), filters)
        last_login = AuditEvent.objects.filter(user=item, action="LOGIN", result="SUCCESS").order_by("-created_at").first()
        last_reveal = AuditEvent.objects.filter(user=item, action="REVEAL").order_by("-created_at").first()
        last_copy = AuditEvent.objects.filter(user=item, action="COPY", result="SUCCESS").order_by("-created_at").first()
        days = (today - timezone.localdate(last_login.created_at)).days if last_login else "Nunca"
        rows.append([item.get_full_name() or item.username, item.vault_profile.get_role_display(), timezone.localtime(last_login.created_at).strftime("%d/%m/%Y %H:%M") if last_login else "Nunca", days, timezone.localtime(last_reveal.created_at).strftime("%d/%m/%Y %H:%M") if last_reveal else "Nunca", timezone.localtime(last_copy.created_at).strftime("%d/%m/%Y %H:%M") if last_copy else "Nunca", events.filter(action="LOGIN", result="SUCCESS").count(), events.filter(action="REVEAL").count(), events.filter(action="COPY", result="SUCCESS").count(), "Activo" if item.vault_profile.mfa_enabled else "Pendiente", "Sí" if item.vault_devices.filter(status=UserDevice.TRUSTED).exists() else "No", "Validar adopción" if days == "Nunca" or isinstance(days, int) and days >= 30 else "Sin señal"])
    return ReportData("Informe de Adopción", ["Usuario", "Rol", "Último ingreso", "Días sin ingresar", "Último revelado", "Última copia", "Accesos", "Revelados", "Copias", "Estado MFA", "Dispositivo reconocido", "Posible baja adopción"], rows, safe_filter_summary(filters))


def _cards_data(user, filters):
    if user.vault_profile.role != UserProfile.LEADER:
        raise ReportValidationError("Su rol no permite generar informes de tarjetas.")
    queryset = _date_filter(PaymentCard.objects.select_related("created_by"), filters)
    rows = []
    for card in queryset.order_by("-created_at")[: filters.get("_row_limit", settings.REPORT_XLSX_MAX_ROWS) + 1]:
        last_view = AuditEvent.objects.filter(card=card, action="VIEW").order_by("-created_at").first()
        last_copy = AuditEvent.objects.filter(card=card, action="COPY", result="SUCCESS").order_by("-created_at").first()
        rows.append([card.pk, safe_reason(card.client_name), safe_reason(card.cardholder_name), card.get_brand_display(), card.last4, "Activa" if card.active else "Inactiva", timezone.localtime(card.created_at).strftime("%d/%m/%Y %H:%M"), str(card.created_by), timezone.localtime(last_view.created_at).strftime("%d/%m/%Y %H:%M") if last_view else "Nunca", timezone.localtime(last_copy.created_at).strftime("%d/%m/%Y %H:%M") if last_copy else "Nunca", "No incluido por seguridad", safe_reason(card.purpose)])
    return ReportData("Informe Seguro de Tarjetas", ["ID interno", "Cliente", "Alias autorizado", "Franquicia", "Últimos cuatro", "Estado", "Fecha de creación", "Creada por", "Última consulta", "Última copia", "Próximo vencimiento", "Observación no sensible"], rows, safe_filter_summary(filters), observations="El PAN y el vencimiento se excluyen expresamente. El periodo de vencimiento no se exporta en esta fase.")


def _health_data(user, filters):
    if user.vault_profile.role != UserProfile.ADMIN:
        raise ReportValidationError("Su rol no permite generar el informe de salud operativa.")
    chain_ok, position = cached_chain_status()
    active_users = _scoped_users(user)
    last_operation = AuditEvent.objects.filter(action__in=OPERATIONAL_ACTIONS).order_by("-created_at").first()
    days_without_use = (timezone.now() - last_operation.created_at).days if last_operation else "Sin actividad registrada"
    last_run = PolicyEvaluationRun.objects.order_by("-started_at").first()
    reasons = []
    critical = SecurityAlert.objects.filter(severity="CRITICAL").exclude(status__in=["CLOSED", "JUSTIFIED"]).count()
    if not chain_ok: reasons.append("Integridad de auditoría requiere atención")
    if critical: reasons.append(f"{critical} alerta(s) crítica(s) abierta(s)")
    if isinstance(days_without_use, int) and days_without_use >= 30: reasons.append(f"{days_without_use} días sin uso operativo")
    state = "Saludable" if not reasons else ("Crítico" if not chain_ok or critical else "Atención")
    rows = [[state, "; ".join(reasons) or "Sin hallazgos activos", f"Íntegra hasta evento {position}" if chain_ok else "Fallo detectado", f"{active_users.filter(vault_profile__mfa_enabled=True).count()} de {active_users.count()} usuarios", SecureSession.objects.filter(status="ACTIVE").count(), UserDevice.objects.filter(status=UserDevice.TRUSTED).count(), critical, days_without_use, NotificationRecord.objects.filter(result__in=["FAILED", "RETRY"]).count(), AccessException.objects.filter(status="ACTIVE", ends_at__gte=timezone.now()).count(), timezone.localtime(last_run.finished_at).strftime("%d/%m/%Y %H:%M") if last_run and last_run.finished_at else "Sin evaluación", timezone.localtime(last_run.started_at).strftime("%d/%m/%Y %H:%M") if last_run else "Sin ejecución"]]
    return ReportData("Informe de Salud Operativa", ["Estado general", "Motivos", "Integridad de auditoría", "MFA", "Sesiones activas", "Dispositivos reconocidos", "Alertas críticas", "Días sin uso", "Fallos de correo", "Excepciones activas", "Última evaluación", "Última tarea programada"], rows, safe_filter_summary(filters), orientation="landscape")


BUILDERS = {"TIMELINE": _timeline_data, "ALERTS": _alerts_data, "ACCESS": _access_data, "ADOPTION": _adoption_data, "CARDS": _cards_data, "HEALTH": _health_data}


def build_report(report_type, user, filters):
    if report_type not in allowed_report_types(user) or report_type not in BUILDERS:
        raise ReportValidationError("Su rol no permite generar este informe.")
    return BUILDERS[report_type](user, filters)


def excel_safe(value):
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def build_excel(data, generated_by, actor_role):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.sheet_view.showGridLines = False
    summary["A1"] = "A&S Vault"
    summary["A1"].font = Font(size=20, bold=True, color="123D6A")
    summary["A3"] = data.title
    summary["A3"].font = Font(size=16, bold=True, color="123D6A")
    logo_path = finders.find("img/branding/logo-ays-azul.png")
    if logo_path:
        try:
            from openpyxl.drawing.image import Image
            logo = Image(logo_path)
            logo.width, logo.height = 150, 48
            summary.add_image(logo, "D1")
        except (ImportError, OSError, ValueError):
            pass
    summary_rows = [
        ("Fecha de generación", timezone.localtime().strftime("%d/%m/%Y %H:%M")),
        ("Generado por", generated_by.get_full_name() or generated_by.username),
        ("Rol", ROLE_LABELS.get(actor_role, actor_role)),
        ("Total de registros", len(data.rows)),
    ] + [(key, value) for key, value in data.filters.items()]
    for index, (label, value) in enumerate(summary_rows, start=5):
        summary.cell(index, 1, label).font = Font(bold=True, color="123D6A")
        summary.cell(index, 2, excel_safe(value))
    warning_row = 6 + len(summary_rows)
    summary.cell(warning_row, 1, CONFIDENTIALITY)
    summary.merge_cells(start_row=warning_row, start_column=1, end_row=warning_row, end_column=6)
    summary.cell(warning_row, 1).fill = PatternFill("solid", fgColor="E8F4FA")
    summary.cell(warning_row, 1).alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 48

    sheet = workbook.create_sheet("Datos")
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    for column, header in enumerate(data.columns, start=1):
        cell = sheet.cell(1, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="123D6A")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_index, values in enumerate(data.rows, start=2):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, excel_safe(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    last_row = max(2, len(data.rows) + 1)
    last_column = sheet.cell(1, len(data.columns)).column_letter
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    if data.rows:
        table = Table(displayName="DatosInforme", ref=f"A1:{last_column}{last_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
        sheet.add_table(table)
    for index, header in enumerate(data.columns, start=1):
        maximum = max([len(str(header))] + [len(str(row[index - 1])) for row in data.rows[:500]])
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(max(maximum + 2, 12), 38)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf(data, generated_by, actor_role, orientation="auto", base_url=None):
    from weasyprint import HTML

    selected_orientation = data.orientation if orientation == "auto" else orientation
    logo_path = finders.find("img/branding/logo-ays-azul.png")
    html = render_to_string("vault/reports/report_pdf.html", {
        "report": data,
        "generated_at": timezone.localtime(),
        "generated_by": generated_by,
        "actor_role": ROLE_LABELS.get(actor_role, actor_role),
        "confidentiality": CONFIDENTIALITY,
        "orientation": selected_orientation,
        "logo_uri": Path(logo_path).resolve().as_uri() if logo_path else "",
    })
    return HTML(string=html, base_url=base_url or str(settings.BASE_DIR)).write_pdf()


def report_filename(data, export_format):
    clean_title = re.sub(r"[^A-Za-zÁÉÍÓÚáéíóúÑñ0-9 _-]", "", data.title).strip()[:80]
    stamp = timezone.localdate().isoformat()
    prefix = "PDF - " if export_format == "PDF" else ""
    return f"{prefix}A&S Vault - {clean_title} - {stamp}.{export_format.lower() if export_format == 'PDF' else 'xlsx'}"
