from io import BytesIO
from datetime import timedelta
from unittest.mock import patch
from xml.etree import ElementTree
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.contrib.staticfiles import finders
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from django_otp.plugins.otp_totp.models import TOTPDevice
from openpyxl import load_workbook

from .forms import TimelineFilterForm
from .crypto import encrypt
from .models import (
    AuditEvent,
    PaymentCard,
    PolicyConfiguration,
    ProtectedOperationContext,
    ReportExport,
    SecureSession,
    SecurityAlert,
    SensitiveOperationWindow,
    UserProfile,
)
from .reporting import excel_safe
from .report_views import _technical_error_code
from .security import audit, session_hash, verify_audit_chain


PASSWORD = "ReportsSecure123!"
TEST_SETTINGS = dict(
    APP_ENV="development",
    FIELD_ENCRYPTION_KEY="MDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDA=",
    FIELD_FINGERPRINT_KEY="reports-test-key",
    PASSWORD_HASHERS=["django.contrib.auth.hashers.MD5PasswordHasher"],
    ALERT_EMAIL_BACKEND="console",
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
    AXES_ENABLED=False,
)


@override_settings(**TEST_SETTINGS)
class ReportingTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.admin = User.objects.create_user("admin.reports", password=PASSWORD, first_name="Ana")
        cls.leader = User.objects.create_user("leader.reports", password=PASSWORD, first_name="Laura")
        cls.analyst = User.objects.create_user("analyst.reports", password=PASSWORD, first_name="Andrés")
        for user, role in ((cls.admin, UserProfile.ADMIN), (cls.leader, UserProfile.LEADER), (cls.analyst, UserProfile.ANALYST)):
            profile = user.vault_profile
            profile.role, profile.active, profile.mfa_enabled, profile.mfa_status = role, True, True, UserProfile.MFA_ACTIVE
            profile.save()
            TOTPDevice.objects.create(user=user, confirmed=True)
        cls.card = PaymentCard(company_name="Empresa de reportes", client_name="=SUM(1,1)", cardholder_name="Alias seguro", brand="VISA", purpose="Uso interno", created_by=cls.leader)
        cls.card.set_pan("4111111111111111")
        cls.card.set_expiry("12/29")
        cls.card.set_code("CODIGO-NO-EXPORTABLE")
        cls.card.save()

    def setUp(self):
        PolicyConfiguration.objects.get_or_create(singleton=1)

    def login(self, user):
        client = Client(REMOTE_ADDR=f"10.20.0.{user.pk}", HTTP_USER_AGENT=f"Report Test Browser {user.pk}")
        client.force_login(user)
        session = client.session
        session["otp_device_id"] = TOTPDevice.objects.get(user=user).persistent_id
        session.save()
        now = timezone.now()
        SecureSession.objects.update_or_create(session_hash=session_hash(session.session_key), defaults={
            "user": user, "encrypted_session_key": encrypt(session.session_key), "last_activity_at": now,
            "expires_at": now + timedelta(minutes=10), "initial_ip": f"10.20.0.{user.pk}",
            "last_ip": f"10.20.0.{user.pk}", "user_agent": f"Report Test Browser {user.pk}",
            "status": SecureSession.ACTIVE, "mfa_completed": True, "mfa_completed_at": now,
        })
        return client

    def test_timeline_form_validates_dates_ip_and_rejects_pan(self):
        invalid_date = TimelineFilterForm({"date_from": "2026-07-15", "date_to": "2026-07-01"}, user=self.admin)
        self.assertFalse(invalid_date.is_valid())
        invalid_ip = TimelineFilterForm({"ip": "dirección-no-válida"}, user=self.admin)
        self.assertFalse(invalid_ip.is_valid())
        pan = TimelineFilterForm({"card": "4111111111111111"}, user=self.admin)
        self.assertFalse(pan.is_valid())

    def test_quick_filter_chips_and_responsive_controls_render(self):
        audit(None, "COPY", user=self.admin)
        response = self.login(self.admin).get(reverse("vault:timeline"), {"period": "7d", "quick_event": "copies", "page_size": "25"})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Últimos 7 días")
        self.assertContains(response, "Más filtros")
        self.assertContains(response, "active-filter-chip")
        self.assertContains(response, "Exportar Excel")
        self.assertEqual(response.context["page"].paginator.per_page, 25)

    def test_timeline_renders_actor_and_system_events_with_null_relations(self):
        audit(None, "VIEW", user=self.admin)
        audit(None, "POLICY_EVALUATION", user=None)
        client = self.login(self.admin)
        for params in ({"view": "compact"}, {"view": "detail"}, {"period": "7d"}):
            response = client.get(reverse("vault:timeline"), params)
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, "Ana")
            self.assertContains(response, "Sistema")
            self.assertContains(response, "No disponible")
            self.assertContains(response, "No aplica")

    def test_analyst_cannot_access_timeline_or_reports(self):
        audit(None, "VIEW", user=self.analyst, card=self.card)
        audit(None, "VIEW", user=self.leader, card=self.card)
        client = self.login(self.analyst)
        response = client.get(reverse("vault:timeline"), {"user": self.leader.pk})
        self.assertEqual(response.status_code, 403)
        export = client.post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]))
        self.assertEqual(export.status_code, 403)
        self.assertFalse(ReportExport.objects.filter(user=self.analyst).exists())

    def test_excel_is_valid_styled_filtered_and_audited(self):
        audit(
            None,
            "COPY",
            user=self.admin,
            reason="Motivo seguro",
            metadata={"context_id": "referencia-historica-no-uuid"},
        )
        response = self.login(self.admin).post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]), {"quick_event": "copies"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Datos"])
        self.assertEqual(workbook.active.title, "Datos")
        self.assertEqual(workbook["Datos"].freeze_panes, "A2")
        self.assertEqual(workbook["Datos"].auto_filter.ref, workbook["Datos"].calculate_dimension())
        self.assertFalse(workbook["Datos"].tables)
        headers = [cell.value for cell in workbook["Datos"][1]]
        self.assertTrue(all(headers))
        self.assertEqual(len(headers), len({header.casefold() for header in headers}))
        self.assertIn("Referencia", headers)
        self.assertIn("Número certificado recibo - Zoho", headers)
        self.assertNotIn("Motivo seguro", headers)
        resaved = BytesIO()
        workbook.save(resaved)
        resaved.seek(0)
        reopened = load_workbook(resaved, data_only=False)
        self.assertEqual(reopened.sheetnames, ["Datos"])
        self.assertEqual(reopened.active.title, "Datos")
        self.assertEqual(reopened["Datos"].auto_filter.ref, reopened["Datos"].calculate_dimension())
        self.assertFalse(reopened["Datos"].tables)
        with ZipFile(BytesIO(response.content)) as archive:
            self.assertIsNone(archive.testzip())
            names = archive.namelist()
            self.assertEqual(
                [name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")],
                ["xl/worksheets/sheet1.xml"],
            )
            self.assertFalse(any(name.startswith("xl/tables/") for name in names))
            self.assertNotIn(b"Resumen", archive.read("xl/workbook.xml"))
            for name in names:
                if name.endswith((".xml", ".rels")):
                    ElementTree.fromstring(archive.read(name))
        content = " ".join(str(cell.value) for sheet in workbook for row in sheet for cell in row)
        self.assertNotIn("4111111111111111", content)
        self.assertNotIn("12/29", content)
        self.assertNotIn("CODIGO-NO-EXPORTABLE", content)
        self.assertTrue(ReportExport.objects.filter(user=self.admin, result="SUCCESS", export_format="XLSX").exists())
        self.assertTrue(AuditEvent.objects.filter(user=self.admin, action="REPORT_EXPORT").exists())
        self.assertTrue(verify_audit_chain()[0])

    def test_timeline_excel_uses_exact_protected_context_zoho_reference(self):
        now = timezone.now()
        window = SensitiveOperationWindow.objects.create(
            user=self.admin,
            session_hash="a" * 64,
            expires_at=now + timedelta(minutes=15),
        )
        context = ProtectedOperationContext.objects.create(
            identity_window=window,
            user=self.admin,
            session_hash="a" * 64,
            card=self.card,
            reason="Consulta operativa",
            internal_reference="ZOHO-EVENTO-001",
            expires_at=now + timedelta(minutes=5),
        )
        audit(
            None,
            "COPY",
            user=self.admin,
            card=self.card,
            reason="Referencia administrativa",
            metadata={"context_id": str(context.public_id), "reference": "ZOHO-HISTORICO"},
        )
        response = self.login(self.admin).post(
            reverse("vault:export_report", args=["TIMELINE", "XLSX"]),
            {"quick_event": "copies"},
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        rows = list(workbook["Datos"].iter_rows(values_only=True))
        headers = list(rows[0])
        exported = dict(zip(headers, rows[1]))
        self.assertEqual(exported["Referencia"], "Referencia administrativa")
        self.assertEqual(exported["Número certificado recibo - Zoho"], "ZOHO-EVENTO-001")

    def test_card_excel_uses_latest_context_for_each_card(self):
        now = timezone.now()
        window = SensitiveOperationWindow.objects.create(
            user=self.leader,
            session_hash="b" * 64,
            expires_at=now + timedelta(minutes=15),
        )
        older = ProtectedOperationContext.objects.create(
            identity_window=window,
            user=self.leader,
            session_hash="b" * 64,
            card=self.card,
            reason="Operación anterior",
            internal_reference="ZOHO-ANTERIOR",
            expires_at=now + timedelta(minutes=5),
            closed_at=now,
        )
        ProtectedOperationContext.objects.filter(pk=older.pk).update(
            created_at=now - timedelta(minutes=2)
        )
        ProtectedOperationContext.objects.create(
            identity_window=window,
            user=self.leader,
            session_hash="b" * 64,
            card=self.card,
            reason="Operación reciente",
            internal_reference="ZOHO-RECIENTE",
            expires_at=now + timedelta(minutes=5),
        )
        response = self.login(self.admin).post(
            reverse("vault:export_report", args=["CARDS", "XLSX"])
        )
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        rows = list(workbook["Datos"].iter_rows(values_only=True))
        headers = list(rows[0])
        exported = dict(zip(headers, rows[1]))
        self.assertEqual(exported["Referencia"], "Uso interno")
        self.assertEqual(exported["Número certificado recibo - Zoho"], "ZOHO-RECIENTE")

    def test_formula_injection_is_neutralized_in_card_report(self):
        response = self.login(self.admin).post(reverse("vault:export_report", args=["CARDS", "XLSX"]))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        values = [cell.value for row in workbook["Datos"] for cell in row]
        self.assertIn("'=SUM(1,1)", values)
        self.assertFalse(any(cell.data_type == "f" for row in workbook["Datos"] for cell in row))

    def test_pdf_is_valid_private_and_contains_no_sensitive_values(self):
        audit(None, "REVEAL", user=self.admin, reason="Consulta autorizada")
        response = self.login(self.admin).post(reverse("vault:export_report", args=["TIMELINE", "PDF"]), {"quick_event": "reveals"})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content[:4], b"%PDF")
        self.assertIn("no-store", response["Cache-Control"])
        self.assertNotIn(b"4111111111111111", response.content)
        self.assertNotIn(b"12/29", response.content)
        self.assertTrue(ReportExport.objects.filter(user=self.admin, result="SUCCESS", export_format="PDF").exists())

    def test_pdf_brand_logo_is_available_from_staticfiles(self):
        logo_path = finders.find(
            "img/branding/cardmanager/Logo-CardManager-CO-COLOR.png"
        )
        self.assertIsNotNone(logo_path)
        self.assertTrue(logo_path.endswith("Logo-CardManager-CO-COLOR.png"))

    def test_excel_failure_is_logged_and_stored_without_security_alert(self):
        client = self.login(self.admin)
        alerts_before = SecurityAlert.objects.count()
        with patch("vault.report_views.build_excel", side_effect=RuntimeError("fallo simulado de Excel")):
            with self.assertLogs("vault.report_views", level="ERROR") as captured:
                response = client.post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]), follow=True)
        record = ReportExport.objects.latest("created_at")
        self.assertEqual(record.result, "FAILED")
        self.assertEqual(record.safe_error, "EXCEL_RENDER_ERROR")
        self.assertContains(response, "No fue posible generar el informe")
        log_text = " ".join(captured.output)
        self.assertNotIn("4111111111111111", log_text)
        self.assertNotIn("12/29", log_text)
        self.assertEqual(SecurityAlert.objects.count(), alerts_before)

    def test_pdf_failure_is_logged_and_classified(self):
        client = self.login(self.admin)
        with patch("vault.report_views.build_pdf", side_effect=RuntimeError("fallo simulado de PDF")):
            with self.assertLogs("vault.report_views", level="ERROR"):
                response = client.post(reverse("vault:export_report", args=["TIMELINE", "PDF"]), follow=True)
        self.assertContains(response, "El intento quedó registrado sin exponer detalles técnicos")
        self.assertEqual(ReportExport.objects.latest("created_at").safe_error, "PDF_RENDER_ERROR")

    def test_missing_dependency_has_distinct_safe_code(self):
        self.assertEqual(_technical_error_code("XLSX", ModuleNotFoundError("openpyxl")), "DEPENDENCY_MISSING")

    def test_successful_export_does_not_create_unnecessary_critical_alert(self):
        client = self.login(self.admin)
        critical_before = SecurityAlert.objects.filter(severity="CRITICAL").count()
        response = client.post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(SecurityAlert.objects.filter(severity="CRITICAL").count(), critical_before)

    def test_role_matrix_allows_admin_safe_cards_and_blocks_operators(self):
        admin = self.login(self.admin)
        leader = self.login(self.leader)
        analyst = self.login(self.analyst)
        self.assertEqual(admin.post(reverse("vault:export_report", args=["CARDS", "XLSX"])).status_code, 200)
        self.assertEqual(leader.post(reverse("vault:export_report", args=["CARDS", "XLSX"])).status_code, 403)
        self.assertEqual(analyst.post(reverse("vault:export_report", args=["TIMELINE", "XLSX"])).status_code, 403)
        self.assertContains(admin.get(reverse("vault:report_center")), "Inventario operativo")
        self.assertEqual(leader.get(reverse("vault:report_center")).status_code, 403)
        self.assertEqual(analyst.get(reverse("vault:report_center")).status_code, 403)

    def test_report_configuration_uses_reusable_modal(self):
        response = self.login(self.admin).get(reverse("vault:report_center"))
        self.assertContains(response, 'id="report-dialog"')
        self.assertContains(response, "Configurar informe")
        self.assertNotContains(response, "<details")

    def test_reports_cannot_be_generated_by_get(self):
        client = self.login(self.admin)
        self.assertEqual(client.get(reverse("vault:export_report", args=["TIMELINE", "XLSX"])).status_code, 405)

    def test_export_post_requires_csrf_token(self):
        authenticated = self.login(self.admin)
        protected = Client(enforce_csrf_checks=True, REMOTE_ADDR=f"10.20.0.{self.admin.pk}")
        protected.cookies = authenticated.cookies
        response = protected.post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]))
        self.assertEqual(response.status_code, 403)
        self.assertFalse(ReportExport.objects.exists())

    @override_settings(REPORT_XLSX_MAX_ROWS=1)
    def test_export_limit_requires_narrower_filters(self):
        audit(None, "ACCESS", user=self.admin)
        audit(None, "ACCESS", user=self.admin)
        response = self.login(self.admin).post(reverse("vault:export_report", args=["TIMELINE", "XLSX"]))
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "límite seguro", status_code=400)
        self.assertTrue(ReportExport.objects.filter(result="LIMITED", safe_error="ROW_LIMIT_EXCEEDED").exists())

    def test_excel_safe_helper_neutralizes_all_formula_prefixes(self):
        for value in ("=1+1", "+cmd", "-2+3", "@SUM(A1:A2)"):
            self.assertTrue(excel_safe(value).startswith("'"))
