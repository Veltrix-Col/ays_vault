from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from django_otp.plugins.otp_totp.models import TOTPDevice
from openpyxl import Workbook, load_workbook

from .bulk_cards import HEADERS, SUPPORTED_BRANDS, validate_workbook
from .crypto import encrypt
from .forms import CardEditForm, CardForm
from .models import AuditEvent, PaymentCard, PolicyConfiguration, SecureSession, SensitiveOperationWindow, UserProfile
from .security import audit, session_hash


PASSWORD = "DeliverySecure123!"
NEW_PASSWORD = "ChangedSecure456!"


@override_settings(
    FIELD_ENCRYPTION_KEY="MDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDAwMDA=",
    FIELD_FINGERPRINT_KEY="delivery-test-fingerprint",
    PASSWORD_HASHERS=["django.contrib.auth.hashers.MD5PasswordHasher"],
    AXES_ENABLED=False,
    ALERT_EMAIL_BACKEND="console",
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
)
class CardManagerDeliveryTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        User = get_user_model()
        cls.admin = User.objects.create_superuser("admin.delivery", "admin@example.invalid", PASSWORD)
        cls.leader = User.objects.create_user("leader.delivery", password=PASSWORD)
        cls.analyst = User.objects.create_user("analyst.delivery", password=PASSWORD)
        for user, role in ((cls.admin, UserProfile.ADMIN), (cls.leader, UserProfile.LEADER), (cls.analyst, UserProfile.ANALYST)):
            profile = user.vault_profile
            profile.role = role
            profile.active = True
            profile.mfa_enabled = True
            profile.mfa_status = UserProfile.MFA_ACTIVE
            profile.save()
            TOTPDevice.objects.create(user=user, name="Delivery", confirmed=True)

    def setUp(self):
        PolicyConfiguration.objects.update_or_create(singleton=1, defaults={"outside_hours_behavior": "ALLOW"})

    def authenticated_client(self, user, with_window=True):
        client = Client(REMOTE_ADDR="10.20.30.40", HTTP_USER_AGENT="Delivery Test Browser")
        client.force_login(user)
        session = client.session
        session["otp_device_id"] = TOTPDevice.objects.get(user=user).persistent_id
        session.save()
        now = timezone.now()
        SecureSession.objects.create(
            user=user,
            session_hash=session_hash(session.session_key),
            encrypted_session_key=encrypt(session.session_key),
            last_activity_at=now,
            expires_at=now + timedelta(hours=2),
            status=SecureSession.ACTIVE,
            mfa_completed=True,
            mfa_completed_at=now,
        )
        if with_window:
            SensitiveOperationWindow.objects.create(
                user=user,
                session_hash=session_hash(session.session_key),
                purpose="sensitive_operations",
                expires_at=now + timedelta(minutes=60),
            )
        return client

    @staticmethod
    def card_data(**overrides):
        data = {
            "company_name": "Empresa Ejemplo S.A.S.",
            "client_name": "Tarjeta operativa",
            "cardholder_name": "Persona autorizada",
            "identity_document": "1000000010",
            "email": "persona@example.invalid",
            "phone": "+57 300 000 0010",
            "brand": "VISA",
            "purpose": "Operación autorizada",
            "active": "on",
            "pan": "4111111111111111",
            "expiry": "12/30",
            "code": "CODIGO-SEGURO-001",
        }
        data.update(overrides)
        return data

    @staticmethod
    def workbook_upload(rows, headers=HEADERS, name="tarjetas.xlsx"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Tarjetas"
        sheet.append(list(headers))
        for row in rows:
            sheet.append(row)
        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        return SimpleUploadedFile(
            name,
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @staticmethod
    def bulk_row(**overrides):
        values = {
            "Empresa": "Empresa Uno S.A.S.",
            "Alias": "Operativa uno",
            "Titular": "Persona Uno",
            "Cédula / Documento de identidad": "1000000020",
            "Correo electrónico": "uno@example.invalid",
            "Teléfono": "+57 300 000 0020",
            "Franquicia": "VISA",
            "Referencia": "Operación uno",
            "Número de tarjeta": "4111111111111111",
            "Vencimiento": "12/30",
            "Código": "CODIGO-SEGURO-020",
        }
        values.update(overrides)
        return [values[header] for header in HEADERS]

    def test_new_contact_fields_are_required_on_create_and_legacy_edit_remains_compatible(self):
        missing = CardForm(self.card_data(identity_document="", email="", phone=""))
        self.assertFalse(missing.is_valid())
        self.assertEqual(set(missing.errors) & {"identity_document", "email", "phone"}, {"identity_document", "email", "phone"})

        legacy = PaymentCard(company_name="Histórica", client_name="Histórica", cardholder_name="Titular", brand="VISA", purpose="Histórica", created_by=self.leader)
        legacy.set_pan("4012888888881881")
        legacy.set_expiry("10/30")
        legacy.set_code("CODIGO-HISTORICO")
        legacy.save()
        edit = CardEditForm(
            {"company_name": "Histórica", "client_name": "Histórica", "cardholder_name": "Titular", "identity_document": "", "email": "", "phone": "", "brand": "VISA", "purpose": "Histórica", "active": "on", "pan": "", "expiry": "", "code": ""},
            instance=legacy,
        )
        self.assertTrue(edit.is_valid(), edit.errors)

    def test_contact_fields_are_saved_on_create_and_edit_without_encryption_or_reveal(self):
        creation = CardForm(self.card_data())
        self.assertTrue(creation.is_valid(), creation.errors)
        card = creation.save(user=self.leader)
        self.assertEqual(card.identity_document, "1000000010")
        self.assertEqual(card.email, "persona@example.invalid")
        self.assertEqual(card.phone, "+57 300 000 0010")
        edit_data = self.card_data(
            identity_document="1000000099",
            email="actualizado@example.invalid",
            phone="+57 300 000 0099",
            pan="",
            expiry="",
            code="",
        )
        edition = CardEditForm(edit_data, instance=card)
        self.assertTrue(edition.is_valid(), edition.errors)
        edition.save(user=self.leader)
        card.refresh_from_db()
        self.assertEqual(card.identity_document, "1000000099")
        self.assertEqual(card.email, "actualizado@example.invalid")
        self.assertEqual(card.phone, "+57 300 000 0099")

    def test_detail_displays_and_copies_each_administrative_value_without_protected_fetch(self):
        card = CardForm(self.card_data()).save(user=self.leader)
        response = self.authenticated_client(self.leader).get(reverse("vault:card_detail", args=[card.pk]))
        self.assertEqual(response.status_code, 200)
        for value in (card.cardholder_name, card.identity_document, card.email, card.phone):
            self.assertContains(response, value)
        for target in ("visible-cardholder", "visible-document", "visible-email", "visible-phone"):
            self.assertContains(response, f'data-copy-visible="{target}"')
        administrative = response.content.decode().split('<section class="panel protected-data-panel"', 1)[0]
        self.assertNotIn("protected-action", administrative)

    def test_bulk_routes_are_admin_only_and_hidden_from_operational_navigation(self):
        admin_response = self.authenticated_client(self.admin).get(reverse("vault:bulk_card_upload"))
        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, "Cargue masivo de tarjetas")
        for user in (self.leader, self.analyst):
            client = self.authenticated_client(user)
            self.assertEqual(client.get(reverse("vault:bulk_card_upload")).status_code, 403)
            self.assertNotContains(client.get(reverse("vault:card_list")), "Cargue masivo")

    def test_downloaded_template_is_reopenable_with_instructions_and_brand_dropdown(self):
        response = self.authenticated_client(self.admin).get(reverse("vault:bulk_card_template"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        workbook = load_workbook(BytesIO(response.content))
        self.assertEqual(workbook.sheetnames, ["Tarjetas", "Instrucciones"])
        sheet = workbook["Tarjetas"]
        self.assertEqual(tuple(cell.value for cell in next(sheet.iter_rows(max_row=1))), HEADERS)
        self.assertEqual(sheet.max_row, 1)
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet.auto_filter.ref, "A1:K1")
        validations = list(sheet.data_validations.dataValidation)
        self.assertEqual(len(validations), 1)
        self.assertEqual(validations[0].type, "list")
        for _value, label in SUPPORTED_BRANDS:
            self.assertIn(label, validations[0].formula1)
        instructions = workbook["Instrucciones"]
        instruction_text = " ".join(str(cell.value or "") for row in instructions.iter_rows() for cell in row)
        for header in HEADERS:
            self.assertIn(header, instruction_text)
        self.assertIn("Diners", instruction_text)
        resaved = BytesIO()
        workbook.save(resaved)
        workbook.close()
        reopened = load_workbook(BytesIO(resaved.getvalue()))
        self.assertEqual(reopened.sheetnames, ["Tarjetas", "Instrucciones"])
        reopened.close()

    def test_bulk_normalizes_all_supported_brand_names_and_codes(self):
        cases = (
            ("VISA", "4111111111111111", "VISA"),
            ("Visa", "4111111111111111", "VISA"),
            ("visa", "4111111111111111", "VISA"),
            ("MC", "5555555555554444", "MC"),
            ("Mastercard", "5555555555554444", "MC"),
            ("MASTER CARD", "5555555555554444", "MC"),
            ("AMEX", "378282246310005", "AMEX"),
            ("American Express", "378282246310005", "AMEX"),
            ("DINERS", "30569309025904", "DINERS"),
            ("Diners", "30569309025904", "DINERS"),
            ("diners", "30569309025904", "DINERS"),
        )
        for supplied, pan, expected in cases:
            with self.subTest(supplied=supplied):
                result = validate_workbook(self.workbook_upload([self.bulk_row(**{"Franquicia": supplied, "Número de tarjeta": pan})]))
                self.assertEqual(result.errors, (), result.errors)
                self.assertEqual(result.forms[0].cleaned_data["brand"], expected)

    def test_bulk_header_normalization_accepts_irrelevant_differences(self):
        headers = tuple(
            ("\ufeff  " + header.upper() + "  ") if index == 0 else ("  " + header.swapcase() + "  ")
            for index, header in enumerate(HEADERS)
        )
        response = self.authenticated_client(self.admin).post(
            reverse("vault:bulk_card_upload"),
            {"file": self.workbook_upload([self.bulk_row(**{"Franquicia": "Visa"})], headers=headers)},
        )
        self.assertRedirects(response, reverse("vault:bulk_card_upload"))
        self.assertEqual(PaymentCard.objects.count(), 1)

    def test_bulk_headers_report_missing_additional_and_duplicate_columns_without_saving(self):
        cases = (
            (HEADERS[:-1], self.bulk_row()[:-1], "Falta la columna obligatoria: Código"),
            (HEADERS + ("Observaciones",), self.bulk_row() + ["texto"], "columna no esperada: Observaciones"),
            (HEADERS + ("Empresa",), self.bulk_row() + ["Otra"], "columna duplicada: Empresa"),
        )
        for headers, row, expected in cases:
            with self.subTest(expected=expected):
                response = self.authenticated_client(self.admin).post(
                    reverse("vault:bulk_card_upload"),
                    {"file": self.workbook_upload([row], headers=headers)},
                )
                self.assertContains(response, expected)
                self.assertContains(response, "No se creó ninguna tarjeta")
                self.assertEqual(PaymentCard.objects.count(), 0)

    def test_bulk_upload_validates_all_rows_then_encrypts_atomically(self):
        second = self.bulk_row(**{
            "Alias": "Operativa dos",
            "Cédula / Documento de identidad": "1000000021",
            "Correo electrónico": "dos@example.invalid",
            "Teléfono": "+57 300 000 0021",
            "Número de tarjeta": "5555555555554444",
            "Franquicia": "MC",
            "Código": "CODIGO-SEGURO-021",
        })
        client = self.authenticated_client(self.admin)
        response = client.post(reverse("vault:bulk_card_upload"), {"file": self.workbook_upload([self.bulk_row(), second])})
        self.assertRedirects(response, reverse("vault:bulk_card_upload"))
        cards = PaymentCard.objects.order_by("pk")
        self.assertEqual(cards.count(), 2)
        self.assertEqual(cards[0].identity_document, "1000000020")
        self.assertEqual(cards[0].email, "uno@example.invalid")
        self.assertEqual(cards[0].phone, "+57 300 000 0020")
        self.assertNotIn("4111111111111111", cards[0].encrypted_pan)
        self.assertNotIn("12/30", cards[0].encrypted_expiry)
        self.assertNotIn("CODIGO-SEGURO-020", cards[0].encrypted_code)
        event = AuditEvent.objects.filter(action="CREATE", reason="Cargue masivo de tarjetas").get()
        self.assertEqual(event.metadata["created_count"], 2)
        self.assertNotIn("4111111111111111", str(event.metadata))
        self.assertNotIn("CODIGO-SEGURO-020", str(event.metadata))

    def test_bulk_upload_rolls_back_entire_file_and_returns_safe_row_errors(self):
        invalid = self.bulk_row(**{"Correo electrónico": "correo-invalido", "Número de tarjeta": "5555555555554444", "Franquicia": "MC"})
        third = self.bulk_row(**{
            "Alias": "Operativa tres",
            "Cédula / Documento de identidad": "1000000022",
            "Correo electrónico": "tres@example.invalid",
            "Teléfono": "+57 300 000 0022",
            "Franquicia": "AMEX",
            "Número de tarjeta": "378282246310005",
        })
        response = self.authenticated_client(self.admin).post(
            reverse("vault:bulk_card_upload"),
            {"file": self.workbook_upload([self.bulk_row(), invalid, third])},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(PaymentCard.objects.count(), 0)
        self.assertContains(response, "Fila 3")
        self.assertContains(response, "Correo electrónico")
        self.assertNotContains(response, "4111111111111111")
        self.assertNotContains(response, "5555555555554444")

    def test_bulk_rejects_wrong_headers_wrong_extension_and_duplicate_pan(self):
        client = self.authenticated_client(self.admin)
        wrong = client.post(reverse("vault:bulk_card_upload"), {"file": self.workbook_upload([self.bulk_row()], headers=("Otro",))})
        self.assertContains(wrong, "columna no esperada: Otro")
        self.assertContains(wrong, "Falta la columna obligatoria: Empresa")
        extension = client.post(reverse("vault:bulk_card_upload"), {"file": SimpleUploadedFile("tarjetas.csv", b"x")})
        self.assertContains(extension, "extensión .xlsx")
        duplicate = client.post(reverse("vault:bulk_card_upload"), {"file": self.workbook_upload([self.bulk_row(), self.bulk_row()])})
        self.assertContains(duplicate, "duplicada dentro del archivo")
        self.assertEqual(PaymentCard.objects.count(), 0)

    def test_bulk_duplicate_against_database_blocks_entire_import(self):
        existing = CardForm(self.card_data())
        self.assertTrue(existing.is_valid(), existing.errors)
        existing.save(user=self.leader)
        second = self.bulk_row(**{
            "Alias": "Duplicada existente",
            "Cédula / Documento de identidad": "1000000098",
            "Correo electrónico": "duplicada@example.invalid",
            "Teléfono": "+57 300 000 0098",
        })
        response = self.authenticated_client(self.admin).post(
            reverse("vault:bulk_card_upload"),
            {"file": self.workbook_upload([second])},
        )
        self.assertContains(response, "Fila 2")
        self.assertContains(response, "La tarjeta ya está registrada")
        self.assertEqual(PaymentCard.objects.count(), 1)
        self.assertFalse(AuditEvent.objects.filter(reason="Cargue masivo de tarjetas").exists())

    def test_real_equivalent_file_with_visa_persists_contact_fields_and_renders_detail(self):
        response = self.authenticated_client(self.admin).post(
            reverse("vault:bulk_card_upload"),
            {"file": self.workbook_upload([self.bulk_row(**{"Franquicia": "Visa"})])},
        )
        self.assertRedirects(response, reverse("vault:bulk_card_upload"))
        card = PaymentCard.objects.get()
        self.assertEqual(card.brand, "VISA")
        self.assertEqual(card.identity_document, "1000000020")
        self.assertEqual(card.email, "uno@example.invalid")
        self.assertEqual(card.phone, "+57 300 000 0020")
        self.assertNotIn("4111111111111111", card.encrypted_pan)
        detail = self.authenticated_client(self.analyst).get(reverse("vault:card_detail", args=[card.pk]))
        for value in (card.identity_document, card.email, card.phone):
            self.assertContains(detail, value)
        administrative = detail.content.decode().split('<section class="panel protected-data-panel"', 1)[0]
        self.assertNotIn("No registrado", administrative)

    def test_all_bulk_columns_are_required_and_whitespace_is_empty(self):
        for header in HEADERS:
            with self.subTest(header=header):
                response = self.authenticated_client(self.admin).post(
                    reverse("vault:bulk_card_upload"),
                    {"file": self.workbook_upload([self.bulk_row(**{header: "   "})])},
                )
                self.assertContains(response, "Fila 2")
                self.assertContains(response, "campo es obligatorio")
                self.assertContains(response, "No se creó ninguna tarjeta")
                self.assertEqual(PaymentCard.objects.count(), 0)

    def test_diners_is_supported_in_bulk_upload(self):
        response = self.authenticated_client(self.admin).post(
            reverse("vault:bulk_card_upload"),
            {"file": self.workbook_upload([self.bulk_row(**{"Franquicia": "Diners", "Número de tarjeta": "30569309025904"})])},
        )
        self.assertRedirects(response, reverse("vault:bulk_card_upload"))
        self.assertEqual(PaymentCard.objects.get().brand, "DINERS")

    def test_diners_is_supported_on_create_and_edit(self):
        creation = CardForm(self.card_data(brand="DINERS", pan="30569309025904"))
        self.assertTrue(creation.is_valid(), creation.errors)
        card = creation.save(user=self.leader)
        edition = CardEditForm(self.card_data(brand="DINERS", pan="", expiry="", code=""), instance=card)
        self.assertTrue(edition.is_valid(), edition.errors)

    def test_bulk_reuses_required_card_expiry_and_brand_validation(self):
        cases = (
            ({"Teléfono": ""}, "campo es obligatorio"),
            ({"Vencimiento": "13/30"}, "Use formato MM/AA"),
            ({"Número de tarjeta": "1234567890123"}, "validación requerida"),
            ({"Franquicia": "MC"}, "franquicia no coincide"),
            ({"Código": ""}, "campo es obligatorio"),
        )
        for overrides, expected in cases:
            with self.subTest(overrides=overrides):
                response = self.authenticated_client(self.admin).post(
                    reverse("vault:bulk_card_upload"),
                    {"file": self.workbook_upload([self.bulk_row(**overrides)])},
                )
                self.assertEqual(response.status_code, 200)
                self.assertContains(response, expected)
                self.assertEqual(PaymentCard.objects.count(), 0)

    def test_admin_standard_password_change_hashes_password_and_no_longer_returns_403(self):
        target = get_user_model().objects.create_user("target.delivery", password=PASSWORD)
        client = self.authenticated_client(self.admin)
        url = reverse("admin:auth_user_password_change", args=[target.pk])
        self.assertEqual(client.get(url).status_code, 200)
        response = client.post(url, {"password1": NEW_PASSWORD, "password2": NEW_PASSWORD})
        self.assertEqual(response.status_code, 302)
        target.refresh_from_db()
        self.assertTrue(target.check_password(NEW_PASSWORD))
        self.assertNotEqual(target.password, NEW_PASSWORD)
        self.assertTrue(AuditEvent.objects.filter(action="PASSWORD_CHANGED", metadata__target_user_id=target.pk).exists())
        self.assertEqual(client.get(reverse("admin:auth_user_change", args=[target.pk])).status_code, 200)
        self.assertTrue(
            SensitiveOperationWindow.objects.filter(
                user=self.admin,
                session_hash=session_hash(client.session.session_key),
                revoked_at__isnull=True,
            ).exists()
        )

    def test_admin_deletes_only_user_without_historical_audit(self):
        removable = get_user_model().objects.create_user("removable.delivery", password=PASSWORD)
        historical = get_user_model().objects.create_user("historical.delivery", password=PASSWORD)
        audit(None, "LOGIN", user=historical)
        client = self.authenticated_client(self.admin)
        removable_url = reverse("admin:auth_user_delete", args=[removable.pk])
        delete_page = client.get(removable_url)
        self.assertEqual(delete_page.status_code, 200)
        self.assertFalse(delete_page.context["perms_lacking"], delete_page.context["perms_lacking"])
        deletion = client.post(removable_url, {"post": "yes"})
        self.assertEqual(deletion.status_code, 302, deletion.content.decode())
        self.assertFalse(get_user_model().objects.filter(pk=removable.pk).exists())
        historical_change = client.get(reverse("admin:auth_user_change", args=[historical.pk]))
        self.assertNotContains(historical_change, "Eliminar")
        self.assertTrue(get_user_model().objects.filter(pk=historical.pk).exists())

    def test_recovery_screen_uses_existing_cardmanager_brand(self):
        response = self.authenticated_client(self.admin).get(reverse("recovery_codes_confirm"))
        self.assertContains(response, "/static/img/branding/cardmanager/Logo-CardManager-CO-COLOR.png")
        self.assertNotContains(response, "/static/img/branding/logo-ays-azul.png")

    def test_policy_is_the_only_visible_operation_reference_label(self):
        card = CardForm(self.card_data()).save(user=self.leader)
        client = self.authenticated_client(self.analyst, with_window=False)
        start = client.post(reverse("vault:reveal", args=[card.pk]), {"field": "pan", "action": "reveal"})
        self.assertEqual(start.status_code, 428)
        identity = client.post(reverse("vault:protected_reauthenticate"), {"intent": start.json()["intent"], "password": PASSWORD})
        self.assertContains(identity, "Póliza")
        self.assertNotContains(identity, "Número certificado")
        self.assertNotContains(identity, "recibo Zoho")
