"""Puente entre los archivos subidos (Django) y el motor `conciliador`.

Mismo patrón que `soat/services/processor.py`: se recibe el contenido ya validado
por el formulario, se materializa en un directorio temporal (preservando el
nombre original del archivo de cobro, del que algunos ramos infieren el periodo),
se ejecuta la conciliación y se devuelve el Excel + un resumen serializable. El
Excel se endurece (neutralización anti fórmula y sin hipervínculos) antes de salir.
"""

from __future__ import annotations

import io
import logging
import os
import re
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from time import monotonic

from ays_zoho_sdk.exceptions import ZohoError
from django.conf import settings
from openpyxl import load_workbook

from conciliador.ramos import obtener_ramo
from conciliador.rules.valor import ComparacionExactaRule
from conciliador.service import (
    ConciliacionArchivos,
    ConciliacionService,
    ConciliacionServiceError,
)
from conciliador.sources.foundry_recibo import ReciboExtraido
from conciliador.sources.zoho_api import resolver_cobros_poliza, resolver_id_poliza
from integrations.zoho import get_zoho

logger = logging.getLogger("conciliacion")

# El boton "Facturar" enlaza directo a los Cobros (modulo Opeeraciones,
# CustomModule6 en la interfaz web) de la poliza, con la poliza misma
# (CustomModule4) como respaldo si no se encuentra ningun cobro. Facturar es
# una accion real de negocio: los enlaces siempre apuntan a Produccion, sin
# importar en que perfil se corrio la conciliacion (conciliar en Sandbox no
# debe llevar a facturar contra datos de prueba).
_ZOHO_CRM_WEB_ORG = "753703967"
_ZOHO_CRM_POLIZAS_TAB = "CustomModule4"
_ZOHO_CRM_COBROS_TAB = "CustomModule6"

# Modulo API de Cobros (mismo modulo cuya tab web es CustomModule6) y los tres
# campos de la interfaz que "Facturar cobro" prellena antes de redirigir, con
# sus nombres API confirmados contra el snapshot de metadatos (docs/zoho/*/
# latest/fields.json): "Certificado (recibo, documento, anexo)",
# "Fecha expedición" y "Pago total cuota" respectivamente.
_MODULO_COBROS = "Opeeraciones"
_CAMPO_CERTIFICADO = "N_mero_de_certificado"
_CAMPO_FECHA_EXPEDICION = "Fecha_de_expedici_n_de_p_liza"
_CAMPO_PAGO_TOTAL_CUOTA = "Valor_de_cuota"


def _url_poliza(poliza_id: str) -> str:
    return f"https://crm.zoho.com/crm/org{_ZOHO_CRM_WEB_ORG}/tab/{_ZOHO_CRM_POLIZAS_TAB}/{poliza_id}"


def _url_cobro(cobro_id: str) -> str:
    return f"https://crm.zoho.com/crm/org{_ZOHO_CRM_WEB_ORG}/tab/{_ZOHO_CRM_COBROS_TAB}/{cobro_id}"


def _resolver_poliza_url(poliza: str) -> str | None:
    try:
        zoho_produccion = get_zoho(profile="production")
        poliza_id = resolver_id_poliza(zoho_produccion, poliza=poliza)
    except ZohoError:
        logger.warning(
            "No fue posible resolver el enlace de la póliza %s", poliza, exc_info=True
        )
        return None
    if not poliza_id:
        return None
    return _url_poliza(poliza_id)


def _resolver_cobros(poliza: str) -> list[dict[str, object]] | None:
    """None => no fue posible consultar Zoho (perfil Producción no disponible,
    error de red, etc.). Lista vacía => se consultó correctamente pero la
    póliza no tiene ninguna Operación/Cobro registrada."""
    try:
        zoho_produccion = get_zoho(profile="production")
        candidatos = resolver_cobros_poliza(zoho_produccion, poliza=poliza)
    except ZohoError:
        logger.warning(
            "No fue posible resolver los cobros para la póliza %s", poliza, exc_info=True
        )
        return None
    return [{**candidato, "url": _url_cobro(candidato["id"])} for candidato in candidatos]


class ConciliacionProcessingError(ValueError):
    """Error de negocio al conciliar; una vista puede mostrar el mensaje tal cual."""


class CobroPrefillError(ValueError):
    """Error de negocio al prellenar el Cobro; una vista puede mostrar el mensaje tal cual."""


class CobroPrefillDisabled(CobroPrefillError):
    """El interruptor `CONCILIACION_COBRO_PREFILL_ENABLED` está apagado."""


class CobroNotFound(CobroPrefillError):
    """El `cobro_id` recibido no está entre los Cobros resueltos para la póliza."""


class CobroPrefillNoData(CobroPrefillError):
    """Ninguno de los 3 campos trae valor (recibo no extraído o campos vacíos)."""


def recibo_prefill_fields(recibo: object) -> dict[str, object] | None:
    """Los 3 valores del recibo (PDF) que "Facturar cobro" puede prellenar en
    Zoho, listos para el summary que ve el frontend. None si el recibo no se
    extrajo (PDF ausente o extracción fallida): en ese caso `ReciboConciliacionRule`
    ya deja la advertencia 'N/D' explicando que el cobro no se prellenará."""
    if not isinstance(recibo, ReciboExtraido):
        return None
    return {
        "certificado": recibo.numero_recibo,
        "fecha_expedicion": recibo.fecha_expedicion,
        "pago_total_cuota": recibo.valor_total_a_pagar,
    }


def prellenar_cobro(
    *, poliza: str, cobro_id: str,
    certificado: str | None, fecha_expedicion: str | None, pago_total_cuota: float | None,
) -> dict[str, object]:
    """Escribe Certificado / Fecha expedición / Pago total cuota en el Cobro
    de Zoho Producción antes de que "Facturar cobro" redirija ahí -- solo los
    campos con valor (prellenado parcial si el recibo no trajo alguno).

    Siempre contra Producción, igual que `_resolver_cobros`/`_url_cobro`: no
    tiene sentido prellenar un cobro de Sandbox si el enlace de facturación
    real apunta a Producción. Antes de escribir, vuelve a resolver los cobros
    de la póliza y exige que `cobro_id` esté entre ellos -- el mismo alcance
    ya público vía `_resolver_cobros`, así el llamador nunca puede dirigir la
    escritura a un id que no haya sido primero resuelto por la póliza dada."""
    if not getattr(settings, "CONCILIACION_COBRO_PREFILL_ENABLED", False):
        raise CobroPrefillDisabled("El prellenado del cobro está deshabilitado.")

    campos: dict[str, object] = {}
    if certificado:
        campos[_CAMPO_CERTIFICADO] = certificado
    if fecha_expedicion:
        campos[_CAMPO_FECHA_EXPEDICION] = fecha_expedicion
    if pago_total_cuota is not None:
        campos[_CAMPO_PAGO_TOTAL_CUOTA] = pago_total_cuota
    if not campos:
        raise CobroPrefillNoData("No hay datos del recibo para prellenar.")

    try:
        zoho_produccion = get_zoho(profile="production")
        candidatos = resolver_cobros_poliza(zoho_produccion, poliza=poliza)
    except ZohoError as exc:
        raise CobroPrefillError(f"No fue posible validar el cobro en Zoho ({exc.category}).") from exc
    if not any(str(candidato["id"]) == str(cobro_id) for candidato in candidatos):
        raise CobroNotFound("El cobro indicado no corresponde a la póliza consultada.")

    try:
        resultado = zoho_produccion.records.update(
            module=_MODULO_COBROS,
            records=({"id": cobro_id, **campos},),
        )
    except ZohoError as exc:
        logger.warning(
            "Fallo al prellenar el cobro %s de la póliza %s", cobro_id, poliza, exc_info=True
        )
        raise CobroPrefillError(f"Falló la escritura en Zoho ({exc.category}).") from exc

    registros = tuple(getattr(resultado, "records", ()))
    if len(registros) != 1 or not getattr(registros[0], "succeeded", False):
        code = str(getattr(registros[0], "code", "") or "WRITE_REJECTED") if registros else "WRITE_REJECTED"
        raise CobroPrefillError(f"Zoho rechazó la actualización del cobro ({code}).")

    return {"cobro_id": str(cobro_id), "campos": sorted(campos)}


@dataclass(frozen=True)
class ConciliacionOutput:
    content: bytes
    filename: str
    summary: dict[str, object]


_NOMBRE_SEGURO = re.compile(r"[^A-Za-z0-9 ._()\-À-ſ]")


def _nombre_seguro(nombre: str, respaldo: str) -> str:
    """Basename saneado que conserva tokens de mes/año para inferir el periodo."""
    base = os.path.basename((nombre or "").replace("\\", "/"))
    base = _NOMBRE_SEGURO.sub("_", base).strip().strip(".")
    return base or respaldo


def _volcar(archivo, destino: Path) -> Path:
    archivo.seek(0)
    with destino.open("wb") as stream:
        for chunk in archivo.chunks():
            stream.write(chunk)
    archivo.seek(0)
    return destino


def _neutralizar(valor):
    if isinstance(valor, str) and valor[:1] in {"=", "+", "-", "@"}:
        return "'" + valor
    return valor


def _endurecer_xlsx(contenido: bytes) -> bytes:
    """Neutraliza celdas de texto con prefijo de fórmula y quita hipervínculos."""
    workbook = load_workbook(io.BytesIO(contenido))
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.hyperlink = None
                    if isinstance(cell.value, str):
                        cell.value = _neutralizar(cell.value)
        salida = io.BytesIO()
        workbook.save(salida)
        return salida.getvalue()
    finally:
        workbook.close()


def _aplicar_umbral_valor_exacto(ramo_codigo: str) -> None:
    """Sobreescribe `ComparacionExactaRule.umbral_pesos` (diferencia en pesos
    por debajo de la cual una diferencia de valor no es incidente) con lo que
    haya en `settings.CONCILIACION_UMBRAL_VALOR_EXACTO`, para poder ajustar el
    umbral por configuracion sin tocar codigo. Mismo patron que ya usa
    `cli.py` con `ComparacionEstadisticaRule.zscore_umbral`."""
    umbral = settings.CONCILIACION_UMBRAL_VALOR_EXACTO
    for regla in obtener_ramo(ramo_codigo).reglas:
        if isinstance(regla, ComparacionExactaRule):
            regla.umbral_pesos = umbral


def procesar_conciliacion(*, ramo: str, poliza: str, archivos: dict) -> ConciliacionOutput:
    """`archivos` es {slot: UploadedFile|None} con las claves cobro, recibo,
    novedades (recibo/novedades pueden faltar). Relación de asegurados y
    Personas siempre se consultan directo a Zoho (Full API), filtradas por
    `poliza`. El perfil (sandbox o producción) lo decide `settings.ZOHO_ACTIVE_PROFILE`
    de forma global para toda la aplicación: esta función no lo recibe como
    parámetro ni lo deja elegir por request."""
    started = monotonic()
    _aplicar_umbral_valor_exacto(ramo)

    with TemporaryDirectory(prefix="ays-conciliacion-") as directorio:
        raiz = Path(directorio)
        rutas: dict[str, Path] = {}
        respaldos = {"novedades": "novedades.xlsx", "cobro": "cobro", "recibo": "recibo.pdf"}
        for slot, archivo in archivos.items():
            if archivo in (None, False):
                continue
            nombre = _nombre_seguro(getattr(archivo, "name", ""), respaldos.get(slot, slot))
            rutas[slot] = _volcar(archivo, raiz / nombre)

        try:
            zoho = get_zoho()
        except ZohoError as exc:
            raise ConciliacionProcessingError(
                f"No fue posible conectar con Zoho ({exc.category}). Intente nuevamente más tarde."
            ) from exc

        entrada = ConciliacionArchivos(
            cobro=rutas["cobro"],
            novedades=rutas.get("novedades"),
            recibo=rutas.get("recibo"),
            zoho=zoho,
            poliza=poliza,
        )
        try:
            resultado = ConciliacionService().ejecutar(ramo, entrada)
        except ConciliacionServiceError as exc:
            raise ConciliacionProcessingError(str(exc)) from exc
        except ZohoError as exc:
            raise ConciliacionProcessingError(
                f"Falló la consulta a Zoho ({exc.category}). Intente nuevamente más tarde."
            ) from exc

        contenido = _endurecer_xlsx(resultado.contenido_excel)
        reporte = resultado.reporte
        # Solo tiene sentido resolver los cobros cuando la conciliación queda sin
        # incidentes bloqueantes: es el único caso en que el frontend muestra la
        # sección. Las advertencias (p. ej. recibo/PDF sin validar) no bloquean.
        poliza_url = _resolver_poliza_url(poliza) if reporte.esta_vacio else None
        cobros = _resolver_cobros(poliza) if reporte.esta_vacio else None
        summary = {
            "ramo": reporte.ramo,
            "periodo": reporte.periodo,
            "poliza": poliza,
            "total_incidentes": int(reporte.total_incidentes),
            "total_advertencias": int(reporte.total_advertencias),
            "sin_incidentes": bool(reporte.esta_vacio),
            "por_tipo": {str(k): int(v) for k, v in resultado.resumen.items()},
            "filename": resultado.nombre_archivo,
            "duration_seconds": round(monotonic() - started, 2),
            "poliza_url": poliza_url,
            "cobros": cobros,
            "recibo_cobro": recibo_prefill_fields(resultado.recibo),
            "cobro_prefill_enabled": bool(getattr(settings, "CONCILIACION_COBRO_PREFILL_ENABLED", False)),
        }
        return ConciliacionOutput(
            content=contenido,
            filename=resultado.nombre_archivo,
            summary=summary,
        )
