from __future__ import annotations

import io
import json
import zipfile
from datetime import timedelta
from unittest.mock import Mock, patch

from django.conf import settings
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import resolve, reverse
from django.utils import timezone
from openpyxl import load_workbook

from vault.crypto import decrypt, encrypt

from cotizacion_colectivos.models import (
    AccesoExternoSolicitudColectivo,
    AdjuntoSolicitudColectivo,
    CambioSolicitudColectivo,
    ColectivosTaskOutbox,
    RespuestaSolicitudColectivo,
    RevisionSolicitudColectivo,
    SolicitudColectivo,
    SolicitudColectivoRegistro,
    NotificacionColectivos,
)
from cotizacion_colectivos.services.attachments import store_attachment
from cotizacion_colectivos.services.excel_roundtrip import (
    build_approved_consolidated,
    build_novelties_template,
    build_response_workbook,
    parse_novelties,
)
from cotizacion_colectivos.services.external import (
    ExternalAccessError,
    generate_access,
    issue_otp,
    resolve_external_session,
    resolve_token,
    save_response,
    submit_response,
    update_access_recipient,
    verify_otp,
)
from cotizacion_colectivos.services.review import finalize_review, record_reviews


@override_settings(
    COLECTIVOS_EXTERNAL_LINK_TTL_SECONDS=3600,
    COLECTIVOS_EXTERNAL_LINK_MAX_TTL_SECONDS=7200,
    COLECTIVOS_EXTERNAL_OTP_MAX_ATTEMPTS=3,
    COLECTIVOS_EXTERNAL_OTP_TTL_SECONDS=600,
)
class ExternalWorkflowTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.admin = User.objects.create_superuser(
            "external-admin", "external-admin@example.test", "Password123!"
        )
        self.request = SolicitudColectivo.objects.create(
            source_kind="company",
            source_reference_hash="a" * 64,
            policy_reference_hash="b" * 64,
            encrypted_policy_token=encrypt("opaque-internal-token"),
            masked_policy_reference="Póliza terminada en 1234",
            client_label="Cliente de prueba",
            branch_code="91",
            branch_name="Salud colectivo",
            request_type=SolicitudColectivo.RequestType.UPDATE,
            status=SolicitudColectivo.Status.READY,
            assigned_to=self.admin,
            deadline=timezone.localdate() + timedelta(days=10),
            zoho_profile="sandbox",
            encrypted_snapshot=encrypt('{"version": 1, "policy": {}, "group": [], "warnings": []}'),
            created_by=self.admin,
        )
        self.record = SolicitudColectivoRegistro.objects.create(
            request=self.request,
            element_type=SolicitudColectivoRegistro.ElementType.PERSON,
            role="Asegurado",
            external_reference_hash="c" * 64,
            initial_status="Activo",
            plan="Plan vigente",
            original_position=1,
            checksum="d" * 64,
        )

    def access(self):
        return generate_access(
            request=self.request,
            actor=self.admin,
            recipient="cliente@example.test",
        )

    def enter_with_otp(self, generated):
        with patch("cotizacion_colectivos.services.external.secrets.randbelow", return_value=123456):
            entry = self.client.get(
                reverse("colectivos_external:entry", args=[generated.token])
            )
        self.assertEqual(entry.status_code, 200)
        self.assertContains(entry, "Verifique su acceso")
        verified = self.client.post(
            reverse("colectivos_external:verify", args=[generated.token]),
            {"code": "123456"},
        )
        self.assertEqual(verified.status_code, 302)
        self.assertIn("colectivos_external_session", verified.cookies)
        return verified

    def test_static_portal_route_is_not_interpreted_as_a_token(self):
        match = resolve(reverse("colectivos_external:portal"))
        self.assertEqual(match.url_name, "portal")

    def test_reused_access_persists_edited_recipient_for_the_next_otp(self):
        original = "original@example.test"
        edited = "edited@example.test"
        generated = generate_access(
            request=self.request, actor=self.admin, recipient=original,
        )
        generated.access.otp_hash = "pending-for-original-recipient"
        generated.access.otp_expires_at = timezone.now() + timedelta(minutes=5)
        generated.access.save(update_fields=("otp_hash", "otp_expires_at"))

        self.assertTrue(update_access_recipient(
            access=generated.access, actor=self.admin, recipient=edited,
        ))
        generated.access.refresh_from_db()
        self.assertEqual(decrypt(generated.access.encrypted_recipient), edited)
        self.assertEqual(generated.access.otp_hash, "")

        backend = Mock(name="edited_recipient_backend")
        backend.name = "smtp"
        backend.send.return_value = "accepted"
        with patch(
            "cotizacion_colectivos.services.external.secrets.randbelow", return_value=112233,
        ), patch("vault.notifications.get_backend", return_value=backend):
            self.assertTrue(issue_otp(generated.access))
        self.assertEqual(backend.send.call_args.args[3], edited)
        self.assertNotEqual(backend.send.call_args.args[3], original)

    def test_external_portal_requires_otp_and_uses_isolated_cookie_without_django_login(self):
        generated = self.access()
        self.request.status = self.request.Status.SENT
        self.request.save(update_fields=("status",))
        self.enter_with_otp(generated)
        portal_response = self.client.get(reverse("colectivos_external:portal"))
        self.assertEqual(portal_response.status_code, 200)
        self.assertNotIn("_auth_user_id", self.client.session)
        self.assertEqual(portal_response["Cache-Control"], "max-age=0, no-cache, no-store, must-revalidate, private")
        self.assertContains(portal_response, "data-row-filter", html=False)
        self.assertContains(portal_response, "data-progress-count", html=False)
        self.assertContains(portal_response, "js/colectivos-external.js", html=False)
        self.assertContains(portal_response, "Mi póliza")
        self.assertContains(portal_response, "Mis pólizas y mi grupo")
        self.assertContains(portal_response, "Buscar dentro de mi información")
        self.assertNotContains(portal_response, "Riesgos1")
        self.assertNotContains(portal_response, "lookups")

    @patch(
        "cotizacion_colectivos.external_views._policy_sections",
        return_value=({"grouping_warnings": ("relación inconsistente",)},),
    )
    def test_portal_warning_log_uses_request_correlation_without_model_attribute(self, _sections):
        generated = self.access()
        self.enter_with_otp(generated)
        with self.assertLogs("cotizacion_colectivos", level="WARNING") as captured:
            portal = self.client.get(
                reverse("colectivos_external:portal"),
                HTTP_X_CORRELATION_ID="qa-safe-correlation",
            )
        self.assertEqual(portal.status_code, 200)
        self.assertIn("correlation=qa-safe-correlation", " ".join(captured.output))
        self.assertNotIn("Cliente de prueba", " ".join(captured.output))

    def test_client_can_confirm_without_creating_a_draft_first(self):
        generated = self.access()
        self.enter_with_otp(generated)
        response = self.client.post(
            reverse("colectivos_external:submit"),
            {"declaration": "on", "client_observations": ""},
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("No hay cambios preparados para enviar.", response.content.decode())
        self.assertFalse(self.request.responses.exists())
        self.assertFalse(ColectivosTaskOutbox.objects.exists())
        self.request.refresh_from_db()
        self.assertNotEqual(self.request.status, self.request.Status.ANSWERED)
        self.assertFalse(NotificacionColectivos.objects.filter(notification_type="CLIENT_RESPONSE").exists())

    @patch("cotizacion_colectivos.services.external._send_submission_receipt")
    def test_valid_novelty_submit_never_sends_automatic_response_receipt(self, receipt):
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{"record": str(self.record.public_key), "action": "RETIRAR", "fecha_retiro": "2026-09-01"}],
            observations="",
        )
        submit_response(access=access, response=response, declaration=True)
        receipt.assert_not_called()

    def test_otp_endpoint_rejects_an_unissued_code(self):
        generated = self.access()
        response = self.client.post(
            reverse("colectivos_external:verify", args=[generated.token]),
            {"code": "123456"},
        )
        self.assertEqual(response.status_code, 400)
        generated.access.refresh_from_db()
        self.assertFalse(generated.access.otp_hash)

    def test_all_confirmed_branches_have_editable_portal(self):
        expected_field = {code: "include_fecha_nacimiento" for code in ("91", "86", "28", "83", "40")}
        for index, branch_code in enumerate(expected_field):
            with self.subTest(branch=branch_code):
                self.request.branch_code = branch_code
                self.request.status = self.request.Status.READY
                self.request.save(update_fields=("branch_code", "status"))
                generated = generate_access(
                    request=self.request,
                    actor=self.admin,
                    recipient="cliente@example.test",
                    regenerate=index > 0,
                )
                self.enter_with_otp(generated)
                portal = self.client.get(reverse("colectivos_external:portal"))
                self.assertContains(portal, "Mis pólizas y mi grupo")
                self.assertContains(portal, expected_field[branch_code])
                saved = self.client.post(reverse("colectivos_external:save_draft"), {})
                self.assertEqual(saved.status_code, 302)

    def test_portal_uses_metadata_backed_identification_and_minimal_ingress(self):
        generated = self.access()
        self.request.status = self.request.Status.SENT
        self.request.save(update_fields=("status",))
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        self.assertContains(portal, 'name="include_rol" value="Asegurado"', html=False)
        self.assertContains(portal, '<select name="include_tipo_id">', html=False)
        self.assertContains(portal, 'name="include_fecha_nacimiento"', html=False)
        self.assertContains(portal, 'name="include_fecha_ingreso"', html=False)
        self.assertNotContains(portal, 'name="include_plan"', html=False)

    def test_external_entities_render_structured_rows_and_drawers(self):
        cases = {
            "91": ("Nombre", "Rol", "Solicitar retiro"),
            "86": ("Nombre", "Parentesco", "Solicitar retiro"),
            "28": ("Inmueble 1", "Calle 10 # 20-30", "Solicitar retiro"),
            "83": ("Asegurado", "Valor asegurado", "Solicitar retiro"),
            "40": ("Vehículo 1", "ABC123", "Solicitar retiro"),
        }
        member = {
            "display_name": "Persona de prueba",
            "insured_name": "Persona de prueba",
            "insured_key": "a" * 64,
            "state": "Activo",
            "plan": "Plan vigente",
            "relationship": "Hija",
            "risk_key": "b" * 64,
            "risk_summary": "Bien registrado",
            "risk_attributes": {
                "direccion": "Calle 10 # 20-30", "ciudad": "Medellín",
                "tipo_uso": "Residencial", "anio_construccion": "2018",
                "vehiculo": "Mazda CX-5", "placa": "ABC123",
                "marca": "Mazda CX-5", "modelo": "2024",
            },
        }
        self.record.economic_values = {"Valor asegurado": "$450.000.000"}
        self.record.save(update_fields=("economic_values",))
        for index, (branch_code, expected) in enumerate(cases.items()):
            with self.subTest(branch=branch_code):
                self.request.branch_code = branch_code
                self.request.status = self.request.Status.READY
                self.request.encrypted_snapshot = encrypt(json.dumps({
                    "version": 1, "policy": {}, "group": [member], "warnings": [],
                }, ensure_ascii=False))
                self.request.save(update_fields=("branch_code", "status", "encrypted_snapshot"))
                generated = generate_access(
                    request=self.request, actor=self.admin,
                    recipient="cliente@example.test", regenerate=index > 0,
                )
                self.enter_with_otp(generated)
                portal = self.client.get(reverse("colectivos_external:portal"))
                self.assertEqual(portal.status_code, 200)
                for text in expected:
                    self.assertContains(portal, text)
                self.assertContains(portal, "data-functional-table", html=False)
                self.assertContains(portal, "data-record-summary", html=False)
                self.assertContains(portal, "data-edit-panel", html=False)
                self.assertContains(portal, "data-edit-panel hidden", html=False)
                self.assertContains(portal, 'role="dialog"', html=False)
                self.assertNotContains(portal, ">Modificar<")
                self.assertNotContains(portal, "data-edit-disclosure", html=False)
                self.assertNotContains(portal, "Titular o principal")
                self.assertNotContains(portal, "Ver familia, beneficiarios y coberturas")

    def test_progressive_editor_keeps_the_existing_backend_field_contract(self):
        member = {
            "display_name": "Persona de prueba", "insured_name": "Persona de prueba",
            "insured_key": "a" * 64, "state": "Activo", "plan": "Plan vigente",
        }
        self.request.encrypted_snapshot = encrypt(json.dumps({
            "version": 1, "policy": {}, "group": [member], "warnings": [],
        }))
        self.request.save(update_fields=("encrypted_snapshot",))
        generated = self.access()
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        functional_key = "a" * 64
        self.assertContains(portal, f'name="action_entity_{functional_key}"', html=False)
        self.assertContains(portal, f'name="fecha_retiro_entity_{functional_key}"', html=False)
        self.assertNotContains(portal, f'name="plan_entity_{functional_key}"', html=False)
        self.assertContains(portal, f'name="source_records_{functional_key}"', html=False)

    def test_drawer_posts_the_unchanged_payload_contract(self):
        functional_key = "a" * 64
        member = {
            "display_name": "Persona de prueba", "insured_name": "Persona de prueba",
            "insured_key": functional_key, "state": "Activo", "plan": "Plan vigente",
        }
        self.request.encrypted_snapshot = encrypt(json.dumps({
            "version": 1, "policy": {}, "group": [member], "warnings": [],
        }))
        self.request.save(update_fields=("encrypted_snapshot",))
        generated = self.access()
        self.enter_with_otp(generated)
        saved = self.client.post(reverse("colectivos_external:save_draft"), {
            f"source_records_{functional_key}": str(self.record.public_key),
            f"action_entity_{functional_key}": "RETIRAR",
            f"fecha_retiro_entity_{functional_key}": "2026-09-01",
        })
        self.assertEqual(saved.status_code, 302)
        response = self.request.responses.get(status=RespuestaSolicitudColectivo.Status.DRAFT)
        change = response.changes.get(functional_field="fecha_retiro")
        self.assertEqual(decrypt(change.encrypted_new_value), "2026-09-01")

    def test_external_table_keeps_final_save_and_local_pagination_controls(self):
        generated = self.access()
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        self.assertContains(portal, "Guardar mis cambios")
        self.assertContains(portal, "data-page-size", html=False)
        self.assertContains(portal, '<option value="25">25</option>', html=False)
        self.assertContains(portal, '<option value="50">50</option>', html=False)
        self.assertContains(portal, '<option value="100">100</option>', html=False)
        self.assertContains(portal, "data-page-previous", html=False)
        self.assertContains(portal, "data-page-next", html=False)

    def test_external_drawer_accessibility_and_javascript_contract(self):
        generated = self.access()
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        self.assertContains(portal, 'aria-modal="true"', html=False)
        self.assertContains(portal, "data-drawer-title", html=False)
        self.assertContains(portal, "data-drawer-close", html=False)
        self.assertContains(portal, "data-drawer-backdrop", html=False)
        script = (settings.BASE_DIR / "static" / "js" / "colectivos-external.js").read_text(encoding="utf-8")
        self.assertIn('event.key === "Escape"', script)
        self.assertIn('event.key !== "Tab"', script)
        self.assertIn("activeTrigger", script)
        self.assertIn("focusableElements", script)

    def test_people_with_same_confirmed_key_render_once_with_consolidated_roles(self):
        shared_key = "a" * 64
        second = SolicitudColectivoRegistro.objects.create(
            request=self.request,
            element_type=SolicitudColectivoRegistro.ElementType.BENEFICIARY,
            role="Beneficiario",
            external_reference_hash="e" * 64,
            initial_status="Activo",
            plan="Plan vigente",
            original_position=2,
            checksum="f" * 64,
        )
        members = [
            {"insured_name": "Persona de prueba", "insured_key": shared_key},
            {"beneficiary_name": "Persona de prueba", "beneficiary_key": shared_key},
        ]
        self.request.encrypted_snapshot = encrypt(json.dumps({
            "version": 1, "policy": {}, "group": members, "warnings": [],
        }))
        self.request.save(update_fields=("encrypted_snapshot",))
        generated = self.access()
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        self.assertEqual(portal.content.count(b"data-functional-entity"), 1)
        self.assertContains(portal, "Asegurado · Beneficiario")
        source_records = ",".join(sorted((str(self.record.public_key), str(second.public_key))))
        self.assertContains(
            portal,
            f'value="{source_records}"',
            html=False,
        )

    @patch("cotizacion_colectivos.zoho.get_zoho")
    def test_external_table_and_drawer_use_only_persisted_snapshot(self, get_zoho):
        generated = self.access()
        self.enter_with_otp(generated)
        portal = self.client.get(reverse("colectivos_external:portal"))
        self.assertEqual(portal.status_code, 200)
        get_zoho.assert_not_called()

    def test_external_table_and_drawer_have_mobile_layout_without_page_overflow(self):
        css = (settings.BASE_DIR / "static" / "css" / "colectivos_external.css").read_text(encoding="utf-8")
        self.assertIn("@media(max-width:1024px)", css)
        self.assertIn("@media(max-width:768px)", css)
        self.assertIn("@media(max-width:480px)", css)
        self.assertIn(".functional-drawer{width:100%}", css)
        self.assertIn("body main{width:min(100% - 1rem,1480px)}", css)

    def test_external_token_is_only_persisted_as_hash_and_is_tamper_evident(self):
        generated = self.access()
        self.assertNotIn(generated.token, generated.access.token_hash)
        self.assertNotIn("cliente@example.test", generated.access.encrypted_recipient)
        self.assertEqual(resolve_token(generated.token).pk, generated.access.pk)
        with self.assertRaises(ExternalAccessError):
            resolve_token(generated.token + "alterado")

    def test_regeneration_revokes_previous_access(self):
        first = self.access()
        second = generate_access(
            request=self.request,
            actor=self.admin,
            recipient="cliente@example.test",
            regenerate=True,
        )
        first.access.refresh_from_db()
        self.assertEqual(first.access.status, first.access.Status.REVOKED)
        self.assertEqual(resolve_token(second.token).pk, second.access.pk)
        with self.assertRaises(ExternalAccessError):
            resolve_token(first.token)

    def test_expired_token_fails_closed(self):
        generated = self.access()
        generated.access.expires_at = timezone.now() - timedelta(seconds=1)
        generated.access.save(update_fields=("expires_at",))
        with self.assertRaises(ExternalAccessError):
            resolve_token(generated.token)

    def test_access_expiry_uses_exact_seconds_not_end_of_day(self):
        before = timezone.now()
        generated = self.access()
        after = timezone.now()
        self.assertGreaterEqual(
            generated.access.expires_at,
            before + timedelta(seconds=settings.COLECTIVOS_EXTERNAL_LINK_TTL_SECONDS),
        )
        self.assertLessEqual(
            generated.access.expires_at,
            after + timedelta(seconds=settings.COLECTIVOS_EXTERNAL_LINK_TTL_SECONDS),
        )

    @override_settings(COLECTIVOS_EXTERNAL_LINK_TTL_SECONDS=172800, COLECTIVOS_EXTERNAL_LINK_MAX_TTL_SECONDS=604800)
    def test_link_ttl_is_elapsed_48_hours_even_when_deadline_is_today(self):
        self.request.deadline = timezone.localdate() + timedelta(days=1)
        self.request.save(update_fields=("deadline",))
        before = timezone.now()
        generated = self.access()
        self.assertGreaterEqual(generated.access.expires_at, before + timedelta(hours=47, minutes=59))
        self.assertLessEqual(generated.access.expires_at, before + timedelta(hours=48, seconds=2))
        self.assertEqual(resolve_token(generated.token).pk, generated.access.pk)

    def test_otp_expires_exactly_with_access_and_email_uses_local_deadline(self):
        generated = self.access()
        expected_expiry = generated.access.expires_at
        with patch(
            "cotizacion_colectivos.services.external.send_notification"
        ) as sender, patch(
            "cotizacion_colectivos.services.external.secrets.randbelow",
            return_value=123456,
        ):
            self.assertTrue(issue_otp(generated.access))
        generated.access.refresh_from_db()
        self.assertEqual(generated.access.otp_expires_at, expected_expiry)
        message = sender.call_args.kwargs
        self.assertIn("mientras el enlace permanezca vigente", message["text_body"])
        self.assertNotIn("minutos", message["text_body"])

    def test_entry_does_not_resend_an_unexpired_otp_after_link_bound_issue(self):
        generated = self.access()
        with patch("cotizacion_colectivos.services.external.send_notification") as sender, patch(
            "cotizacion_colectivos.services.external.secrets.randbelow", return_value=123456,
        ):
            first = self.client.get(reverse("colectivos_external:entry", args=[generated.token]))
            second = self.client.get(reverse("colectivos_external:entry", args=[generated.token]))
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(sender.call_count, 1)

    def test_email_backend_receives_real_otp_while_database_and_logs_do_not(self):
        generated = self.access()
        backend = Mock(name="otp_backend")
        backend.name = "smtp"
        backend.send.return_value = "accepted"
        with patch("cotizacion_colectivos.services.external.secrets.randbelow", return_value=123456), patch(
            "vault.notifications.get_backend", return_value=backend,
        ), patch("vault.notifications.logger") as notification_logger:
            self.assertTrue(issue_otp(generated.access))

        subject, text_body, html_body, recipient = backend.send.call_args.args
        self.assertEqual(recipient, "cliente@example.test")
        self.assertIn("123456", text_body)
        self.assertIn("123456", html_body)
        self.assertNotIn("[CÓDIGO OMITIDO]", text_body + html_body)
        self.assertIn("Código de verificación", subject + html_body)
        generated.access.refresh_from_db()
        self.assertNotEqual(generated.access.otp_hash, "123456")
        self.assertNotIn("123456", repr(generated.access.__dict__))
        self.assertNotIn("123456", repr(notification_logger.mock_calls))

    def test_otp_is_one_use_and_external_session_is_request_scoped(self):
        generated = self.access()
        generated.access.otp_hash = ""
        from django.contrib.auth.hashers import make_password

        generated.access.otp_hash = make_password("123456")
        generated.access.otp_expires_at = timezone.now() + timedelta(minutes=5)
        generated.access.save(update_fields=("otp_hash", "otp_expires_at"))
        cookie = verify_otp(generated.access, "123456")
        session_access = resolve_external_session(cookie)
        self.assertEqual(session_access.request_id, self.request.pk)
        session_access.refresh_from_db()
        self.assertEqual(session_access.otp_hash, "")
        with self.assertRaises(ExternalAccessError):
            verify_otp(session_access, "123456")

    def test_otp_blocks_after_bounded_failures(self):
        generated = self.access()
        from django.contrib.auth.hashers import make_password

        generated.access.otp_hash = make_password("123456")
        generated.access.otp_expires_at = timezone.now() + timedelta(minutes=5)
        generated.access.save(update_fields=("otp_hash", "otp_expires_at"))
        for _ in range(3):
            with self.assertRaises(ExternalAccessError):
                verify_otp(generated.access, "000000")
        generated.access.refresh_from_db()
        self.assertEqual(generated.access.status, generated.access.Status.BLOCKED)

    def verified_access(self):
        generated = self.access()
        generated.access.status = generated.access.Status.VERIFIED
        generated.access.save(update_fields=("status",))
        self.request.status = self.request.Status.OPENED
        self.request.save(update_fields=("status",))
        return generated.access

    def test_web_draft_versions_changes_and_submission_is_idempotent(self):
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{
                "record": str(self.record.public_key),
                "action": "RETIRAR",
                "fecha_retiro": "2026-09-01",
            }],
            observations="Observación del cliente",
        )
        self.assertEqual(response.origin, RespuestaSolicitudColectivo.Origin.WEB)
        self.assertNotIn("Observación del cliente", response.encrypted_client_observations)
        self.assertTrue(response.changes.filter(functional_field="fecha_retiro").exists())
        submitted = submit_response(access=access, response=response, declaration=True)
        again = submit_response(access=access, response=submitted, declaration=True)
        self.assertEqual(again.pk, submitted.pk)
        self.request.refresh_from_db()
        self.assertEqual(self.request.status, self.request.Status.ANSWERED)
        access.refresh_from_db()
        self.assertEqual(access.status, access.Status.USED)

    def test_personal_identifier_validation_for_inclusion(self):
        access = self.verified_access()
        with self.assertRaises(ExternalAccessError):
            save_response(
                access=access,
                rows=[{
                    "record": "",
                    "action": "INCLUIR",
                    "tipo_id": "CC",
                    "documento": "12-34",
                    "nombre": "Persona",
                    "fecha_efectiva": "2026-09-01",
                }],
                observations="",
            )

    def test_novelties_use_requested_dates_without_effective_date(self):
        access = self.verified_access()
        ingress = save_response(
            access=access,
            rows=[{
                "record": "",
                "action": "INCLUIR",
                "tipo_id": "CC",
                "documento": "12345678",
                "nombres": "Persona",
                "apellidos": "de prueba",
                "rol": "Asegurado",
                "fecha_nacimiento": "1990-01-01",
                "fecha_ingreso": "2026-09-01",
            }],
            observations="",
        )
        self.assertTrue(ingress.changes.filter(functional_field="fecha_ingreso").exists())
        self.assertTrue(ingress.changes.filter(functional_field="fecha_nacimiento").exists())
        self.assertFalse(ingress.changes.filter(functional_field="fecha_efectiva").exists())

        retirement = save_response(
            access=access,
            rows=[{
                "record": str(self.record.public_key),
                "action": "RETIRAR",
                "fecha_retiro": "2026-10-01",
            }],
            observations="",
        )
        self.assertTrue(retirement.changes.filter(functional_field="fecha_retiro").exists())
        self.assertFalse(retirement.changes.filter(functional_field="fecha_efectiva").exists())

    def test_template_round_trip_and_formula_rejection(self):
        content = build_novelties_template(self.request)
        upload = SimpleUploadedFile(
            "novedades.xlsx",
            content,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        preview = parse_novelties(upload, self.request)
        self.assertEqual(preview.counts["SIN_CAMBIOS"], 1)

        workbook = load_workbook(io.BytesIO(content))
        workbook["Novedades"]["A2"] = "=1+1"
        modified = io.BytesIO()
        workbook.save(modified)
        bad = SimpleUploadedFile("novedades.xlsx", modified.getvalue())
        with self.assertRaises(ValidationError):
            parse_novelties(bad, self.request)

    def test_functional_excel_consolidates_roles_and_keeps_signed_source_mapping(self):
        functional_key = "a" * 64
        payload = {
            "display_name": "Información protegida",
            "insured_name": "Información protegida",
            "insured_key": functional_key,
        }
        self.record.encrypted_branch_payload = encrypt(json.dumps(payload))
        self.record.save(update_fields=("encrypted_branch_payload",))
        second = SolicitudColectivoRegistro.objects.create(
            request=self.request,
            element_type=SolicitudColectivoRegistro.ElementType.PERSON,
            role="Asegurado",
            external_reference_hash="e" * 64,
            initial_status="Activo",
            plan="Plan vigente",
            encrypted_branch_payload=encrypt(json.dumps(payload)),
            original_position=2,
            checksum="f" * 64,
        )
        content = build_novelties_template(self.request)
        workbook = load_workbook(io.BytesIO(content), read_only=True)
        self.assertEqual(workbook["Novedades"].max_row, 2)
        upload = SimpleUploadedFile("novedades.xlsx", content)
        preview = parse_novelties(upload, self.request)
        self.assertEqual(preview.rows[0]["functional_key"], functional_key)
        self.assertEqual(
            set(preview.rows[0]["records"]),
            {str(self.record.public_key), str(second.public_key)},
        )

    def test_one_functional_novelty_references_multiple_technical_records(self):
        second = SolicitudColectivoRegistro.objects.create(
            request=self.request,
            element_type=SolicitudColectivoRegistro.ElementType.PERSON,
            role="Beneficiario",
            external_reference_hash="e" * 64,
            initial_status="Activo",
            plan="Plan vigente",
            original_position=2,
            checksum="f" * 64,
        )
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{
                "record": str(self.record.public_key),
                "records": (str(self.record.public_key), str(second.public_key)),
                "functional_key": "a" * 64,
                "action": "MODIFICAR",
                "plan": "Plan solicitado",
                "fecha_efectiva": "2026-09-01",
            }],
            observations="",
        )
        self.assertEqual(response.changes.filter(functional_field="accion").count(), 1)
        self.assertEqual(response.changes.filter(functional_field="plan").count(), 1)
        mapping = json.loads(decrypt(response.changes.get(functional_field="accion").encrypted_branch_payload))
        self.assertEqual(
            set(mapping["source_record_keys"]),
            {str(self.record.public_key), str(second.public_key)},
        )

    def test_macro_or_external_link_payload_is_rejected(self):
        content = build_novelties_template(self.request)
        source = zipfile.ZipFile(io.BytesIO(content))
        altered = io.BytesIO()
        with zipfile.ZipFile(altered, "w") as output:
            for info in source.infolist():
                output.writestr(info, source.read(info.filename))
            output.writestr("xl/vbaProject.bin", b"not-a-macro")
        uploaded = SimpleUploadedFile("novedades.xlsx", altered.getvalue())
        with self.assertRaises(ValidationError):
            parse_novelties(uploaded, self.request)

    @override_settings(COLECTIVOS_PRIVATE_ROOT=None)
    def test_attachment_setting_is_required(self):
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{"record": str(self.record.public_key), "action": "SIN_CAMBIOS"}],
            observations="",
        )
        uploaded = SimpleUploadedFile("soporte.pdf", b"%PDF-1.7\n%%EOF", content_type="application/pdf")
        with self.assertRaises((ValidationError, TypeError)):
            store_attachment(response=response, uploaded=uploaded)

    @override_settings(COLECTIVOS_PRIVATE_ROOT="private_assets/colectivos-test")
    def test_attachment_rejects_extension_content_mismatch(self):
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{"record": str(self.record.public_key), "action": "SIN_CAMBIOS"}],
            observations="",
        )
        uploaded = SimpleUploadedFile("soporte.pdf", b"not a pdf", content_type="application/pdf")
        with self.assertRaises(ValidationError):
            store_attachment(response=response, uploaded=uploaded)
        self.assertFalse(AdjuntoSolicitudColectivo.objects.exists())

    def test_no_write_api_is_exposed_by_external_services(self):
        import cotizacion_colectivos.services.external as external

        forbidden = {"create", "update", "delete", "upsert", "upload", "attach"}
        self.assertFalse(forbidden.intersection(set(dir(external))))

    def submitted_response(self):
        access = self.verified_access()
        response = save_response(
            access=access,
            rows=[{
                "record": str(self.record.public_key),
                "action": "MODIFICAR",
                "plan": "Plan aprobado",
                "fecha_efectiva": "2026-09-01",
            }],
            observations="",
        )
        return submit_response(access=access, response=response, declaration=True)

    def test_response_export_contains_no_review_and_approved_export_requires_decisions(self):
        response = self.submitted_response()
        response_content = build_response_workbook(response)
        response_book = load_workbook(io.BytesIO(response_content), data_only=False)
        self.assertEqual(response_book.sheetnames, ["Resumen", "Respuesta", "Metadatos"])
        with self.assertRaises(ValidationError):
            build_approved_consolidated(response)

        decisions = {
            change.pk: {"decision": RevisionSolicitudColectivo.Decision.APPROVE}
            for change in response.changes.all()
        }
        record_reviews(response=response, reviewer=self.admin, decisions=decisions)
        finalize_review(response=response, reviewer=self.admin, action="approve")
        response.refresh_from_db()
        content = build_approved_consolidated(response)
        approved_book = load_workbook(io.BytesIO(content), data_only=False)
        self.assertIn("Consolidado aprobado", approved_book.sheetnames)

    @patch("cotizacion_colectivos.services.external.send_invitation")
    def test_correction_reuses_same_request_and_rotates_access(self, send_invitation_mock):
        response = self.submitted_response()
        decisions = {
            change.pk: {
                "decision": RevisionSolicitudColectivo.Decision.CORRECTION,
                "client_observation": "Ajuste requerido",
            }
            for change in response.changes.all()
        }
        record_reviews(response=response, reviewer=self.admin, decisions=decisions)
        with self.captureOnCommitCallbacks(execute=True):
            result = finalize_review(response=response, reviewer=self.admin, action="correction")
        self.assertEqual(result.pk, self.request.pk)
        result.refresh_from_db()
        self.assertEqual(result.status, result.Status.CORRECTION)
        self.assertEqual(result.external_accesses.filter(status="ACTIVO").count(), 1)
        send_invitation_mock.assert_called_once()
