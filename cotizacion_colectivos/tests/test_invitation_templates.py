from __future__ import annotations

import hashlib
import io
import zipfile
import inspect
import tempfile
import posixpath
from dataclasses import replace
from pathlib import Path
from unittest.mock import patch
from xml.etree import ElementTree as ET

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from cotizacion_colectivos.dto import ContactSummary, GroupMember, PolicyDetail
from cotizacion_colectivos.invitation_templates.catalog import (
    INVITATION_TEMPLATE_CATALOG,
    templates_for_branch,
)
from cotizacion_colectivos.services.common import sign_record_id
from cotizacion_colectivos.services.common import ColectivosServiceError
from cotizacion_colectivos.services.invitation_templates import (
    _context,
    generate_invitation_templates,
    preview_invitation_templates,
    sign_branch_invitation_context,
)


POLICY_ID = "4234567890123456789"
SOURCE_ID = "5234567890123456789"
TOKEN = sign_record_id(
    POLICY_ID, "policy", context={"source_id": SOURCE_ID, "source_kind": "company"},
)
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def policy(branch="40"):
    return PolicyDetail(
        detail_token=TOKEN, masked_reference="Referencia terminada en 8971",
        full_reference="POLIZA-PRUEBA", branch_code=branch,
        branch_name="Movilidad colectivo" if branch == "40" else "Vida grupo deudores",
        classification="confirmed", insurer="Aseguradora actual", state="Vigente",
        holder="Empresa de prueba", start_date="2026-08-01", end_date="2027-08-01",
        renewable="Sí", payment_mode="Contado", frequency="Anual", installments="1",
        first_installment_date="", payment_calendar=(), insured=(), risks=(),
        active_count=1, excluded_count=0, source_kind="company",
        source_name="Empresa de prueba",
        source_summary=ContactSummary(
            person_type="Persona jurídica", id_type="NIT", masked_document="•••123",
            state="Cliente", city="Bogotá", document="900000001",
        ),
    )


def member(index=1):
    return GroupMember(
        role="Asegurado", display_name=f"Persona {index}", id_type="CC",
        masked_document="•••456", document=f"10000000{index}", state="Activo",
        entry_date="", exit_date="", plan="", relationship="Titular",
        risk_summary="Vehículo", insured_name=f"Persona {index}",
        insured_id_type="CC", insured_document=f"10000000{index}",
        associate_document="900000001", risk_key=f"risk-{index}",
        risk_attributes=(("placa", f"ABC{index:03d}"), ("modelo", "2025"),
                         ("marca", "Marca prueba"), ("ciudad", "Bogotá"),
                         ("tipo_uso", "Familiar")),
    )


def local_workspace(branch="40", members=None):
    return policy(branch), tuple(members or (member(),)), {
        "status": "hit", "storage": "database",
    }, {"source_kind": "company"}, "sandbox", "sdk"


def inline_value(archive: bytes, part: str, coordinate: str):
    with zipfile.ZipFile(io.BytesIO(archive)) as package:
        root = ET.fromstring(package.read(part))
    cell = root.find(f".//{{{NS}}}c[@r='{coordinate}']")
    value = cell.find(f"{{{NS}}}is/{{{NS}}}t") if cell is not None else None
    return value.text if value is not None else ""


class InvitationTemplateCatalogTests(TestCase):
    def test_catalog_is_closed_and_does_not_infer_insurer_from_filename(self):
        self.assertEqual(len(INVITATION_TEMPLATE_CATALOG), 4)
        general = next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "allianz_autos_collective")
        self.assertEqual(general.insurer_code, "ALLIANZ")
        self.assertNotEqual(general.insurer_name, "SURA")
        self.assertEqual({item.code for item in templates_for_branch("40", active_only=True)}, {
            "sura_autos_quote", "allianz_autos_collective",
        })
        sura = next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "sura_autos_quote")
        allianz = next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "allianz_autos_collective")
        self.assertTrue(sura.expandable_rows)
        self.assertFalse(sura.supports_chunking)
        self.assertTrue(allianz.supports_chunking)

    def test_legacy_xls_is_not_silently_converted_or_enabled(self):
        legacy = next(item for item in INVITATION_TEMPLATE_CATALOG if item.extension == "xls")
        self.assertFalse(legacy.active)
        self.assertEqual(legacy.generator, "unsupported_biff8")
        self.assertTrue(legacy.path.read_bytes().startswith(bytes.fromhex("D0CF11E0A1B11AE1")))

    def test_reorganized_sources_resolve_and_allianz_life_is_active(self):
        self.assertTrue(all(item.path.is_file() for item in INVITATION_TEMPLATE_CATALOG))
        allianz = next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "allianz_vg_collective")
        self.assertTrue(allianz.active)
        self.assertEqual(allianz.path.parent.name, "allianz")
        self.assertEqual({item.insurer_code for item in templates_for_branch("83")}, {"SURA", "ALLIANZ"})

    def test_local_loader_has_no_facade_or_zoho_query_dependency(self):
        from cotizacion_colectivos.services import invitation_templates as service
        source = inspect.getsource(service._local_workspace)
        self.assertIn("load_policy_preparation", source)
        self.assertNotIn("PolicyService", source)
        self.assertNotIn("get_zoho", source)

    def test_openpyxl_load_save_pure_is_not_byte_or_package_preserving(self):
        # This is a diagnostic guard: the generator deliberately does not use
        # openpyxl because a pure round-trip drops unsupported OOXML parts.
        with tempfile.TemporaryDirectory() as directory:
            for item in (
                next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "sura_autos_quote"),
                next(item for item in INVITATION_TEMPLATE_CATALOG if item.code == "allianz_autos_collective"),
            ):
                copy = Path(directory) / f"{item.code}.xlsx"
                load_workbook(item.path).save(copy)
                with zipfile.ZipFile(item.path) as original, zipfile.ZipFile(copy) as roundtrip:
                    self.assertNotEqual(set(original.namelist()), set(roundtrip.namelist()))

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_generated_worksheets_keep_ignorable_namespace_prefixes_and_relationship_targets(self, workspace):
        workspace.return_value = local_workspace()
        for insurer in ("SURA", "ALLIANZ"):
            content, _filename, content_type, errors = generate_invitation_templates(
                TOKEN, insurer_code=insurer,
            )
            self.assertEqual(errors, ())
            self.assertIn("spreadsheetml", content_type)
            with zipfile.ZipFile(io.BytesIO(content)) as package:
                names = set(package.namelist())
                for name in names:
                    if name.endswith(".xml") or name.endswith(".rels"):
                        ET.fromstring(package.read(name))
                for rels_name in (name for name in names if name.endswith(".rels")):
                    rels = ET.fromstring(package.read(rels_name))
                    if rels_name == "_rels/.rels":
                        base = Path(".")
                    else:
                        source_part = rels_name.replace("/_rels/", "/").removesuffix(".rels")
                        base = Path(source_part).parent
                    for relationship in rels:
                        target = relationship.get("Target", "")
                        if target.startswith("/"):
                            target_name = target.lstrip("/")
                        else:
                            target_name = posixpath.normpath(str((base / target).as_posix()))
                        self.assertIn(target_name, names, (rels_name, target_name))
                sheet = next(name for name in names if name.startswith("xl/worksheets/sheet1.xml"))
                raw = package.read(sheet)
                opening_start = raw.find(b"?>") + 2
                opening = raw[opening_start:raw.find(b">", opening_start)]
                self.assertIn(b"Ignorable=\"x14ac xr xr2 xr3\"", raw)
                for prefix in (b"x14ac", b"xr", b"xr2", b"xr3"):
                    self.assertIn(b"xmlns:" + prefix + b"=", opening)
                self.assertTrue(load_workbook(io.BytesIO(content)).sheetnames)


class InvitationTemplateGenerationTests(TestCase):
    def setUp(self):
        self.source_hashes = {
            item.path: hashlib.sha256(item.path.read_bytes()).hexdigest()
            for item in INVITATION_TEMPLATE_CATALOG
        }

    def tearDown(self):
        for path, digest in self.source_hashes.items():
            self.assertEqual(hashlib.sha256(path.read_bytes()).hexdigest(), digest)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_preview_uses_only_workspace_and_reports_all_branch_templates(self, workspace):
        workspace.return_value = local_workspace()
        detail, previews, metadata = preview_invitation_templates(TOKEN)
        self.assertEqual(detail.branch_code, "40")
        self.assertEqual({item.template.insurer_code for item in previews}, {"SURA", "ALLIANZ"})
        self.assertTrue(all(item.rows == 1 for item in previews))
        self.assertEqual(metadata["status"], "hit")
        workspace.assert_called_once_with(TOKEN)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_generation_returns_zip_for_all_active_templates_and_precaches_person_company_data(self, workspace):
        workspace.return_value = local_workspace()
        content, filename, content_type, errors = generate_invitation_templates(TOKEN)
        self.assertEqual(content_type, "application/zip")
        self.assertTrue(filename.endswith(".zip"))
        self.assertEqual(errors, ())
        with zipfile.ZipFile(io.BytesIO(content)) as bundle:
            names = set(bundle.namelist())
            self.assertEqual(names, {"sura_movilidad.xlsx", "allianz_movilidad.xlsx"})
            sura = bundle.read("sura_movilidad.xlsx")
            general = bundle.read("allianz_movilidad.xlsx")
        self.assertEqual(inline_value(sura, "xl/worksheets/sheet1.xml", "A2"), "ABC001")
        self.assertEqual(inline_value(sura, "xl/worksheets/sheet1.xml", "Q2"), "100000001")
        self.assertEqual(inline_value(general, "xl/worksheets/sheet1.xml", "B5"), "Empresa de prueba")
        self.assertEqual(inline_value(general, "xl/worksheets/sheet2.xml", "D2"), "ABC001")

    def test_holder_document_comes_from_exact_source_for_person_or_company(self):
        detail = policy()
        fixed_company, _rows = _context(detail, (member(),), {"source_kind": "company"})
        fixed_person, _rows = _context(detail, (member(),), {"source_kind": "person"})
        self.assertEqual(fixed_company["holder.document"], "900000001")
        self.assertEqual(fixed_person["holder.document"], "900000001")
        self.assertNotEqual(fixed_company["holder.document"], member().document)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_branch_preview_consolidates_same_client_branch_from_local_workspaces(self, workspace):
        second_token = sign_record_id(
            "4234567890123456790", "policy",
            context={"source_id": SOURCE_ID, "source_kind": "company"},
        )
        first_workspace = local_workspace(members=(member(1),))
        first_workspace = (
            first_workspace[0], first_workspace[1], first_workspace[2],
            {"source_kind": "company", "source_id": SOURCE_ID}, "sandbox", "sdk",
        )
        second_policy = replace(policy(), detail_token=second_token, full_reference="POLIZA-2")
        workspace.side_effect = (
            first_workspace,
            (second_policy, (member(2),), {"status": "hit"},
             {"source_kind": "company", "source_id": SOURCE_ID}, "sandbox", "sdk"),
        )
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN, second_token), branch_code="40",
            holder="Empresa de prueba", policy_references=("POLIZA-PRUEBA", "POLIZA-2"),
        )
        detail, previews, metadata = preview_invitation_templates(
            context_token, consolidated=True,
        )
        self.assertEqual(detail.full_reference, "2 pólizas vigentes del ramo")
        self.assertEqual({item.rows for item in previews}, {2})
        self.assertEqual(metadata["remote_queries"], 0)
        self.assertTrue(metadata["complete"])
        self.assertEqual(
            [item["policy_reference"] for item in metadata["operational_groups"]],
            ["POLIZA-PRUEBA", "POLIZA-2"],
        )
        self.assertEqual(workspace.call_count, 2)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_branch_missing_workspace_keeps_available_policy_operational(self, workspace):
        second_token = sign_record_id(
            "4234567890123456790", "policy",
            context={"source_id": SOURCE_ID, "source_kind": "company"},
        )
        first = local_workspace()
        workspace.side_effect = (
            (first[0], first[1], first[2],
             {"source_kind": "company", "source_id": SOURCE_ID}, "sandbox", "sdk"),
            ColectivosServiceError("workspace_unavailable", "No disponible"),
        )
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN, second_token), branch_code="40",
            holder="Empresa de prueba", policy_references=("POLIZA-1", "POLIZA-2"),
        )
        _detail, previews, metadata = preview_invitation_templates(
            context_token, consolidated=True,
        )
        self.assertTrue(previews)
        self.assertFalse(metadata["complete"])
        self.assertEqual(metadata["missing_workspaces"], ("POLIZA-2",))
        workspace.side_effect = (
            (first[0], first[1], first[2],
             {"source_kind": "company", "source_id": SOURCE_ID}, "sandbox", "sdk"),
            ColectivosServiceError("workspace_unavailable", "No disponible"),
        )
        content, filename, content_type, errors = generate_invitation_templates(
            context_token, consolidated=True, insurer_code="SURA",
        )
        self.assertEqual(content_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(filename, "sura_movilidad.xlsx")
        self.assertEqual(errors, ())
        self.assertEqual(inline_value(content, "xl/worksheets/sheet1.xml", "A2"), "ABC001")

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_branch_rejects_policy_from_another_branch(self, workspace):
        workspace.return_value = local_workspace(branch="83")
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN,), branch_code="40", holder="Empresa de prueba",
        )
        with self.assertRaises(ColectivosServiceError):
            preview_invitation_templates(context_token, consolidated=True)

    def test_branch_rejects_tampered_signed_context(self):
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN,), branch_code="40", holder="Empresa de prueba",
        )
        altered = context_token[:-1] + ("a" if context_token[-1] != "a" else "b")
        with self.assertRaises(ColectivosServiceError):
            preview_invitation_templates(altered, consolidated=True)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_ooxml_preserves_every_non_target_part_and_formula_validation_structure(self, workspace):
        workspace.return_value = local_workspace()
        content, _filename, _content_type, _errors = generate_invitation_templates(TOKEN)
        with zipfile.ZipFile(io.BytesIO(content)) as bundle:
            generated = bundle.read("allianz_movilidad.xlsx")
        source_path = next(item.path for item in INVITATION_TEMPLATE_CATALOG if item.code == "allianz_autos_collective")
        with zipfile.ZipFile(source_path) as before, zipfile.ZipFile(io.BytesIO(generated)) as after:
            self.assertEqual(set(before.namelist()), set(after.namelist()))
            changed = {"xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml", "docProps/custom.xml"}
            for name in before.namelist():
                if name not in changed:
                    self.assertEqual(after.read(name), before.read(name), name)
            sheet1 = ET.fromstring(after.read("xl/worksheets/sheet1.xml"))
            self.assertEqual(sheet1.find(f".//{{{NS}}}c[@r='B3']/{{{NS}}}f").text, "TODAY()")
            self.assertEqual(len(sheet1.findall(f".//{{{NS}}}dataValidation")), 3)
            sheet2 = ET.fromstring(after.read("xl/worksheets/sheet2.xml"))
            self.assertEqual(len(sheet2.findall(f".//{{{NS}}}dataValidation")), 2)
            custom = ET.fromstring(after.read("docProps/custom.xml"))
            self.assertEqual(list(custom), [])

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_large_group_is_chunked_without_truncation(self, workspace):
        workspace.return_value = local_workspace(members=tuple(member(i) for i in range(1, 302)))
        content, _filename, content_type, errors = generate_invitation_templates(TOKEN)
        self.assertEqual(content_type, "application/zip")
        self.assertEqual(errors, ())
        with zipfile.ZipFile(io.BytesIO(content)) as bundle:
            self.assertEqual(len([name for name in bundle.namelist() if name.startswith("sura_")]), 1)
            self.assertEqual(len([name for name in bundle.namelist() if name.startswith("allianz_")]), 2)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_sura_31_records_stays_one_file_in_general_download(self, workspace):
        workspace.return_value = local_workspace(members=tuple(member(i) for i in range(1, 31)))
        content, filename, content_type, errors = generate_invitation_templates(TOKEN)
        self.assertEqual(content_type, "application/zip")
        self.assertTrue(filename.endswith(".zip"))
        self.assertEqual(errors, ())
        with zipfile.ZipFile(io.BytesIO(content)) as bundle:
            self.assertEqual(set(bundle.namelist()), {
                "sura_movilidad.xlsx",
                "allianz_movilidad.xlsx",
            })

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_per_insurer_download_is_xlsx_or_exclusive_zip_without_truncation(self, workspace):
        workspace.return_value = local_workspace(members=tuple(member(i) for i in range(1, 32)))
        allianz, allianz_name, allianz_type, _errors = generate_invitation_templates(
            TOKEN, insurer_code="ALLIANZ",
        )
        self.assertTrue(allianz_name.startswith("allianz_"))
        self.assertIn("spreadsheetml", allianz_type)
        self.assertTrue(allianz.startswith(b"PK"))
        sura, sura_name, sura_type, _errors = generate_invitation_templates(
            TOKEN, insurer_code="SURA",
        )
        self.assertIn("spreadsheetml", sura_type)
        self.assertEqual(sura_name, "sura_movilidad.xlsx")
        with zipfile.ZipFile(io.BytesIO(sura)) as workbook:
            root = ET.fromstring(workbook.read("xl/worksheets/sheet1.xml"))
        record_count = sum(
            1 for cell in root.findall(f".//{{{NS}}}c")
            if cell.get("r", "").startswith("A")
            and cell.find(f"{{{NS}}}is/{{{NS}}}t") is not None
            and cell.find(f"{{{NS}}}is/{{{NS}}}t").text
        )
        self.assertEqual(record_count, 31)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_sura_136_records_expand_one_file_without_truncation(self, workspace):
        workspace.return_value = local_workspace(members=tuple(member(i) for i in range(1, 137)))
        _detail, previews, _metadata = preview_invitation_templates(TOKEN)
        sura_preview = next(item for item in previews if item.template.insurer_code == "SURA")
        self.assertEqual(sura_preview.output_files, 1)
        self.assertEqual(sura_preview.status, "ready_manual")
        content, filename, content_type, errors = generate_invitation_templates(
            TOKEN, insurer_code="SURA",
        )
        self.assertEqual(errors, ())
        self.assertEqual(filename, "sura_movilidad.xlsx")
        self.assertIn("spreadsheetml", content_type)
        with zipfile.ZipFile(io.BytesIO(content)) as workbook:
            root = ET.fromstring(workbook.read("xl/worksheets/sheet1.xml"))
        cells = {
            cell.get("r"): cell for cell in root.findall(f".//{{{NS}}}c")
            if cell.get("r", "").startswith("A")
        }
        populated = [cell for cell in cells.values()
                     if cell.find(f"{{{NS}}}is/{{{NS}}}t") is not None
                     and cell.find(f"{{{NS}}}is/{{{NS}}}t").text]
        self.assertEqual(len(populated), 136)
        self.assertEqual(cells["A137"].get("s"), cells["A22"].get("s"))
        self.assertEqual(root.find(f"{{{NS}}}dimension").get("ref"), "A1:T137")

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_sensitive_values_are_never_logged(self, workspace):
        workspace.return_value = local_workspace()
        with self.assertLogs("cotizacion_colectivos", level="INFO") as captured:
            generate_invitation_templates(TOKEN)
        combined = " ".join(captured.output)
        self.assertNotIn("900000001", combined)
        self.assertNotIn("100000001", combined)
        self.assertNotIn("Empresa de prueba", combined)

    @patch("cotizacion_colectivos.services.invitation_templates._local_workspace")
    def test_inactive_xls_is_reported_but_not_generated(self, workspace):
        workspace.return_value = local_workspace(branch="83")
        _detail, previews, _metadata = preview_invitation_templates(TOKEN)
        self.assertEqual(len(previews), 2)
        self.assertEqual({item.template.insurer_code: item.status for item in previews}["SURA"], "unavailable")
        content, filename, content_type, errors = generate_invitation_templates(TOKEN)
        self.assertEqual(filename, "allianz_83.xlsx")
        self.assertEqual(content_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(errors, ())
        self.assertEqual(inline_value(content, "xl/worksheets/sheet1.xml", "B10"), "Empresa de prueba")
        self.assertEqual(inline_value(content, "xl/worksheets/sheet1.xml", "B26"), "2027-08-01")
        self.assertIn(inline_value(content, "xl/worksheets/sheet1.xml", "B12"), ("", None))
        with zipfile.ZipFile(io.BytesIO(content)) as package:
            sheet = ET.fromstring(package.read("xl/worksheets/sheet1.xml"))
            self.assertEqual(sheet.find(f".//{{{NS}}}c[@r='B4']/{{{NS}}}f").text, "TODAY()")


class InvitationTemplateViewTests(TestCase):
    def setUp(self):
        self.admin = get_user_model().objects.create_superuser(
            username="invitation-admin", password="safe-test-password",
            email="admin@example.test",
        )
        self.client.force_login(self.admin)

    @patch("cotizacion_colectivos.views.preview_invitation_templates")
    def test_preview_is_get_anti_idor_and_contains_no_raw_identifier(self, preview):
        preview.return_value = policy(), (), {
            "status": "hit", "complete": True, "operational_groups": (),
        }
        url = reverse("cotizacion_colectivos:policy_invitation_preview", args=[TOKEN])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, POLICY_ID)
        self.assertNotContains(response, SOURCE_ID)
        altered = TOKEN[:-1] + ("a" if TOKEN[-1] != "a" else "b")
        with patch(
            "cotizacion_colectivos.views.preview_invitation_templates",
            side_effect=__import__("cotizacion_colectivos.services.common", fromlist=["ColectivosServiceError"]).ColectivosServiceError("invalid_record", "No válida"),
        ):
            self.assertEqual(self.client.get(reverse("cotizacion_colectivos:policy_invitation_preview", args=[altered])).status_code, 404)

    @patch("cotizacion_colectivos.views.preview_invitation_templates")
    def test_each_insurer_has_a_short_independent_mailto(self, preview):
        from cotizacion_colectivos.services.invitation_templates import TemplatePreview
        templates = templates_for_branch("40", active_only=True)
        preview.return_value = policy(), tuple(
            TemplatePreview(item, "ready", 1, 0, 1, 100, ())
            for item in templates
        ), {"status": "hit", "complete": True, "operational_groups": ({
            "policy_reference": "POLIZA-PRUEBA", "insurer": "Actual",
            "rows": ({"document": "100000001", "plate": "ABC001",
                      "model": "2025", "brand": "Marca", "city": "Bogotá",
                      "relationship": "Titular", "insured_name": "Persona 1"},),
        },)}
        response = self.client.get(reverse(
            "cotizacion_colectivos:policy_invitation_preview", args=[TOKEN],
        ))
        self.assertContains(response, "Preparar correo SURA")
        self.assertContains(response, "Preparar correo Allianz")
        self.assertEqual(response.content.count(b"invitation-preview-table"), 1)
        self.assertContains(response, "ABC001")
        self.assertContains(response, "100000001")
        for item in response.context["actions"]:
            self.assertLess(len(item["mailto_url"]), 700)
            self.assertIn(item["insurer_name"], item["mailto_url"])

    @patch("cotizacion_colectivos.views.preview_invitation_templates")
    def test_branch_view_groups_records_by_policy_and_actions_by_catalog_insurer(self, preview):
        from cotizacion_colectivos.services.invitation_templates import TemplatePreview
        templates = templates_for_branch("40", active_only=True)
        preview.return_value = replace(
            policy(), full_reference="2 pólizas vigentes del ramo",
        ), tuple(
            TemplatePreview(item, "ready", 1, 0, 2, 100, ())
            for item in templates
        ), {
            "complete": True, "remote_queries": 0,
            "operational_groups": (
                {"policy_reference": "POLIZA-1", "insurer": "SURA", "rows": ({
                    "document": "1001", "insured_name": "Persona uno", "plate": "AAA001",
                    "model": "2025", "brand": "Marca", "city": "Bogotá", "relationship": "Titular",
                },)},
                {"policy_reference": "POLIZA-2", "insurer": "Allianz", "rows": ({
                    "document": "1002", "insured_name": "Persona dos", "plate": "BBB002",
                    "model": "2024", "brand": "Marca", "city": "Medellín", "relationship": "Titular",
                },)},
            ),
        }
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN,), branch_code="40", holder="Empresa de prueba",
        )
        response = self.client.get(reverse(
            "cotizacion_colectivos:branch_invitation_preview", args=[context_token],
        ))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content.count(b"invitation-preview-table"), 1)
        self.assertContains(response, "POLIZA-1")
        self.assertContains(response, "POLIZA-2")
        self.assertEqual(
            {item["insurer_code"] for item in response.context["actions"]},
            {"SURA", "ALLIANZ"},
        )

    @patch("cotizacion_colectivos.views.generate_invitation_templates")
    def test_branch_download_requests_only_selected_catalog_insurer(self, generate):
        generate.return_value = (
            b"zip", "Invitaciones_SURA_40.zip", "application/zip", (),
        )
        context_token = sign_branch_invitation_context(
            policy_tokens=(TOKEN,), branch_code="40", holder="Empresa de prueba",
        )
        response = self.client.post(reverse(
            "cotizacion_colectivos:branch_invitation_download", args=[context_token],
        ), {"insurer_code": "SURA"})
        self.assertEqual(response.status_code, 200)
        generate.assert_called_once_with(
            context_token, template_code="", insurer_code="SURA", consolidated=True,
        )

    def test_local_filter_supports_policy_plate_document_name_and_clear(self):
        source = (Path(__file__).resolve().parents[2] / "static" / "js" / "colectivos-invitations.js").read_text(encoding="utf-8")
        self.assertIn("data-invitation-policy-filter", source)
        self.assertIn("row.dataset.policyKey", source)
        self.assertIn("row.dataset.invitationSearchText", source)
        self.assertIn("option.checked = false", source)
        self.assertNotIn("fetch(", source)

    @patch("cotizacion_colectivos.views.generate_invitation_templates")
    def test_download_requires_post_and_csrf_and_sets_private_headers(self, generate):
        generate.return_value = (b"xlsx", "Invitacion_SURA_40.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", ())
        url = reverse("cotizacion_colectivos:policy_invitation_download", args=[TOKEN])
        self.assertEqual(self.client.get(url).status_code, 405)
        response = self.client.post(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("no-store", response["Cache-Control"])
        self.assertIn("private", response["Cache-Control"])
        generate.assert_called_once_with(TOKEN, template_code="", insurer_code="")
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.admin)
        self.assertEqual(csrf_client.post(url).status_code, 403)
