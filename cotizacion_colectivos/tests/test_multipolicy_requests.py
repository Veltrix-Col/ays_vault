from __future__ import annotations

import io
from datetime import timedelta
from types import SimpleNamespace

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.utils import timezone
from openpyxl import load_workbook

from vault.crypto import decrypt, encrypt

from cotizacion_colectivos.filenames import download_filename
from cotizacion_colectivos.dto import BranchSummary, GroupMember, PolicyDetail, RelatedPolicy, RequestPolicyOption
from cotizacion_colectivos.forms import MultiPolicyRequestForm
from cotizacion_colectivos.models import (
    CambioSolicitudColectivo,
    SolicitudColectivo,
    SolicitudColectivoPoliza,
    SolicitudColectivoRegistro,
)
from cotizacion_colectivos.services.excel_roundtrip import build_novelties_template, parse_novelties
from cotizacion_colectivos.services.external import ExternalAccessError, generate_access, save_response
from cotizacion_colectivos.services.common import sign_record_id
from cotizacion_colectivos.services.requests import create_request_from_policies, regenerate_request_snapshot, request_snapshot
from cotizacion_colectivos.views import _builder_policies


class FakeMultiPolicyService:
    profile = "sandbox"

    def __init__(self):
        self.source_kinds = []

    def group(self, token, *, source_kind=None):
        self.source_kinds.append(source_kind)
        from cotizacion_colectivos.services.common import unsign_record_id
        policy_id = unsign_record_id(token, "policy")
        branch_code, branch_name, suffix = ("91", "Salud colectivo", "1814") if policy_id.endswith("1") else ("40", "Movilidad colectivo", "8971")
        detail = PolicyDetail(
            detail_token=token, masked_reference=f"Póliza terminada en {suffix}",
            branch_code=branch_code, branch_name=branch_name, classification="confirmed",
            insurer="Aseguradora", state="Vigente", holder="Empresa",
            start_date="2026-01-01", end_date="2026-12-31", renewable="Sí",
            payment_mode="", frequency="", installments="", first_installment_date="",
            payment_calendar=(), insured=(), risks=(), active_count=1, excluded_count=0,
        )
        member = GroupMember(
            role="Asegurado", display_name="Persona", id_type="CC",
            masked_document="••123", document="123", state="Activo",
            entry_date="", exit_date="", plan="", relationship="", risk_summary="",
        )
        return detail, (member,)


@override_settings(
    COLECTIVOS_EXTERNAL_LINK_TTL_SECONDS=3600,
    COLECTIVOS_EXTERNAL_LINK_MAX_TTL_SECONDS=7200,
)
class MultiPolicyRequestTests(TestCase):
    def setUp(self):
        self.actor = get_user_model().objects.create_superuser(
            "multi-admin", "multi-admin@example.test", "Password123!"
        )
        self.request = SolicitudColectivo.objects.create(
            source_kind="company",
            source_reference_hash="a" * 64,
            policy_reference_hash="b" * 64,
            encrypted_policy_token=encrypt("legacy-token"),
            masked_policy_reference="Póliza terminada en 1814",
            client_label="Empresa de prueba",
            branch_code="91",
            branch_name="Salud colectivo",
            request_type=SolicitudColectivo.RequestType.UPDATE,
            status=SolicitudColectivo.Status.OPENED,
            assigned_to=self.actor,
            deadline=timezone.localdate() + timedelta(days=10),
            zoho_profile="sandbox",
            encrypted_snapshot=encrypt('{"version": 1, "policy": {}, "group": [], "warnings": []}'),
            created_by=self.actor,
        )
        self.health = self._policy(1, "b", "1814", "91", "Salud colectivo", ["SIN_CAMBIOS", "MODIFICACION"])
        self.mobility = self._policy(2, "c", "8971", "40", "Movilidad colectivo", ["SIN_CAMBIOS", "INCLUSION"])
        self.record = SolicitudColectivoRegistro.objects.create(
            request=self.request,
            policy=self.health,
            element_type=SolicitudColectivoRegistro.ElementType.PERSON,
            role="Asegurado",
            external_reference_hash="d" * 64,
            initial_status="Activo",
            plan="Plan vigente",
            encrypted_branch_payload=encrypt('{"display_name":"Persona","id_type":"CC","document":"123"}'),
            original_position=1,
            checksum="e" * 64,
        )
        self.access = generate_access(request=self.request, actor=self.actor).access
        self.request.status = SolicitudColectivo.Status.OPENED
        self.request.save(update_fields=("status",))

    def _policy(self, position, hash_prefix, suffix, branch_code, branch_name, adjustments):
        payload = encrypt('{"version": 1, "policy": {}, "group": [], "warnings": []}')
        return SolicitudColectivoPoliza.objects.create(
            request=self.request,
            policy_reference_hash=hash_prefix * 64,
            encrypted_policy_token=encrypt(f"token-{position}"),
            masked_policy_reference=f"Póliza terminada en {suffix}",
            branch_code=branch_code,
            branch_name=branch_name,
            enabled_adjustments=adjustments,
            encrypted_snapshot=payload,
            snapshot_checksum="f" * 64,
            position=position,
        )

    def test_change_is_bound_to_policy_and_its_adjustment_allowlist(self):
        response = save_response(
            access=self.access,
            rows=[{
                "record": str(self.record.public_key),
                "action": "MODIFICAR",
                "fecha_efectiva": "2026-09-01",
                "plan": "Plan solicitado",
            }],
            observations="",
        )
        self.assertFalse(response.changes.exclude(policy=self.health).exists())
        with self.assertRaises(ExternalAccessError):
            save_response(
                access=self.access,
                rows=[{"record": str(self.record.public_key), "action": "RETIRAR", "fecha_efectiva": "2026-09-01"}],
                observations="",
            )

    def test_inclusion_requires_and_uses_a_policy_from_same_request(self):
        response = save_response(
            access=self.access,
            rows=[{
                "record": "", "policy": str(self.mobility.pk), "action": "INCLUIR",
                "tipo_id": "CC", "documento": "123456", "nombre": "Persona",
                "rol": "Asegurado", "fecha_efectiva": "2026-09-01",
            }],
            observations="",
        )
        self.assertFalse(response.changes.exclude(policy=self.mobility).exists())
        with self.assertRaises(ExternalAccessError):
            save_response(
                access=self.access,
                rows=[{
                    "record": "", "policy": "999999", "action": "INCLUIR",
                    "tipo_id": "CC", "documento": "123456", "nombre": "Persona",
                    "rol": "Asegurado", "fecha_efectiva": "2026-09-01",
                }],
                observations="",
            )

    def test_workbook_has_one_signed_sheet_per_policy_and_converges_on_response_rows(self):
        content = build_novelties_template(self.request)
        workbook = load_workbook(io.BytesIO(content))
        self.assertIn("Salud_colectivo_1814", workbook.sheetnames)
        self.assertIn("Movilidad_colectivo_8971", workbook.sheetnames)
        upload = SimpleUploadedFile("novedades.xlsx", content)
        preview = parse_novelties(upload, self.request)
        self.assertEqual(preview.counts["SIN_CAMBIOS"], 1)
        self.assertEqual(preview.rows[0]["policy"], str(self.health.pk))

    def test_adjustments_are_preselected_but_can_be_unchecked(self):
        option = RequestPolicyOption(
            detail_token="opaque", masked_reference="Póliza protegida",
            branch_code="91", branch_name="Salud colectivo", insurer="Aseguradora",
            state="Vigente",
        )
        initial = MultiPolicyRequestForm(policies=(option,))
        self.assertEqual(
            tuple(initial.fields["adjustments_0"].initial),
            ("SIN_CAMBIOS", "INCLUSION", "RETIRO", "MODIFICACION"),
        )
        bound = MultiPolicyRequestForm(
            {
                "request_type": SolicitudColectivo.RequestType.UPDATE,
                "deadline": (timezone.localdate() + timedelta(days=2)).isoformat(),
                "confirm_snapshot": "on", "policy_0": "on",
                "adjustments_0": ("RETIRO",),
            },
            policies=(option,),
        )
        self.assertTrue(bound.is_valid(), bound.errors)
        self.assertEqual(bound.cleaned_data["selections"][0]["adjustments"], ["RETIRO"])

    def test_builder_only_selects_confirmed_collective_operable_policies(self):
        source_id = "1000000000000000001"
        def related(identifier, *, state="Vigente", layout="collective"):
            return RelatedPolicy(
                detail_token=sign_record_id(
                    identifier, "policy", {"source_id": source_id, "source_kind": "company"}
                ),
                masked_reference="Póliza protegida", state=state,
                branch="Salud colectivo", insurer="Aseguradora",
                layout_category=layout,
            )
        branch = BranchSummary(
            code="91", slug="salud", name="Salud colectivo",
            classification="confirmed",
            policies=(
                related("2000000000000000001"),
                related("2000000000000000002", state="Cancelada"),
                related("2000000000000000003", layout="individual"),
            ),
            insured_count=3, risk_count=0, active_count=1, excluded_count=2,
        )
        available, unavailable = _builder_policies(SimpleNamespace(branches=(branch,)))
        self.assertEqual(len(available), 1)
        self.assertEqual(len(unavailable), 2)
        self.assertEqual({reason for _policy, reason in unavailable}, {
            "Cancelada.", "No corresponde a un diseño colectivo confirmado.",
        })

    def test_beneficiary_row_keeps_associate_and_insured_in_novelties_excel(self):
        payload = {
            "display_name": "Beneficiario", "id_type": "TI", "document": "300",
            "associate_name": "Asociado", "associate_id_type": "CC", "associate_document": "100",
            "insured_name": "Asegurado", "insured_id_type": "CC", "insured_document": "200",
            "beneficiary_name": "Beneficiario", "beneficiary_id_type": "TI", "beneficiary_document": "300",
            "relationship": "Hijo", "risk_attributes": {},
        }
        self.record.role = "Beneficiario"
        self.record.encrypted_branch_payload = encrypt(__import__("json").dumps(payload))
        self.record.save(update_fields=("role", "encrypted_branch_payload"))
        workbook = load_workbook(io.BytesIO(build_novelties_template(self.request)))
        sheet = workbook["Salud_colectivo_1814"]
        values = tuple(cell.value for cell in sheet[2])
        self.assertEqual(values[1:10], ("CC", "100", "Asociado", "CC", "200", "Asegurado", "TI", "300", "Beneficiario"))
        self.assertNotIn("No determinada", tuple(cell.value for row in workbook["Póliza"] for cell in row))

    def test_policy_sheet_mapping_is_tamper_evident(self):
        workbook = load_workbook(io.BytesIO(build_novelties_template(self.request)))
        metadata = workbook["Metadatos"]
        for row in metadata.iter_rows():
            if row[0].value == "sheet:Salud_colectivo_1814":
                row[1].value = self.mobility.position
        stream = io.BytesIO()
        workbook.save(stream)
        with self.assertRaises(ValidationError):
            parse_novelties(SimpleUploadedFile("novedades.xlsx", stream.getvalue()), self.request)

    def test_filename_is_safe_and_contains_no_document(self):
        name = download_filename(
            "Novedades", origin="Compañía / Cliente", request_id="COL-2026-ABC12345"
        )
        self.assertRegex(name, r"^Novedades_Compania_Cliente_COL-2026-ABC12345_\d{8}_\d{6}\.xlsx$")
        self.assertNotIn("/", name)

    def test_models_expose_no_zoho_write_actions(self):
        forbidden = {"create", "update", "delete", "upsert", "write", "upload", "attach"}
        from cotizacion_colectivos.services import external, requests
        public = {
            name for module in (external, requests)
            for name, value in vars(module).items()
            if callable(value) and not name.startswith("_")
        }
        self.assertFalse(forbidden & public)

    def test_constructor_creates_one_request_two_policy_snapshots_and_scopes_person(self):
        service = FakeMultiPolicyService()
        source_id = "1000000000000000001"
        tokens = (
            sign_record_id("2000000000000000001", "policy", {"source_id": source_id, "source_kind": "person"}),
            sign_record_id("2000000000000000002", "policy", {"source_id": source_id, "source_kind": "person"}),
        )
        item = create_request_from_policies(
            selections=(
                {"token": tokens[0], "adjustments": ("MODIFICACION",)},
                {"token": tokens[1], "adjustments": ("INCLUSION",)},
            ),
            source_kind="person",
            actor=self.actor,
            assigned_to=self.actor,
            request_type=SolicitudColectivo.RequestType.UPDATE,
            deadline=timezone.localdate() + timedelta(days=5),
            client_label="Persona",
            service=service,
        )
        self.assertEqual(item.policies.count(), 2)
        self.assertEqual(item.records.count(), 2)
        self.assertEqual(service.source_kinds, ["person", "person"])
        self.assertEqual(len(request_snapshot(item)["policies"]), 2)
        self.assertEqual(
            set(item.policies.values_list("enabled_adjustments", flat=True)[0]),
            {"SIN_CAMBIOS", "MODIFICACION"},
        )
        self.assertNotEqual(decrypt(item.policies.first().encrypted_policy_token), tokens[0])
        regenerated = regenerate_request_snapshot(request=item, actor=self.actor, service=service)
        self.assertEqual(regenerated.snapshot_revision, 2)
