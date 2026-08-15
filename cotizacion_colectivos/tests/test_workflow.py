from __future__ import annotations

import json
from datetime import timedelta
from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import OperationalError
from django.test import Client, TestCase, SimpleTestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from cotizacion_colectivos.branches import (
    COLLECTIVE_BRANCH_CONFIG,
    BranchConfigurationError,
    classify_branch,
    validate_branch_config,
)
from cotizacion_colectivos.dto import GroupMember, PolicyDetail
from cotizacion_colectivos.excel import build_current_policy_workbook
from cotizacion_colectivos.models import AccesoCotizacionIndividual, CambioSolicitudColectivo, CotizacionIndividual, EventoSolicitudColectivo, NotificacionCotizacionIndividual, NotificacionColectivos, RespuestaSolicitudColectivo, SolicitudColectivo
from cotizacion_colectivos.services.common import ColectivosServiceError, sign_record_id, unsign_record_id
from cotizacion_colectivos.services.requests import create_or_reuse_request_from_policy, create_request_from_policy, request_snapshot, transition_request
from cotizacion_colectivos.quotation_forms.security import sign_receipt
from cotizacion_colectivos.services.individual_access import generate_individual_access
from vault.crypto import encrypt


POLICY_ID = "4234567890123456789"
TOKEN = sign_record_id(POLICY_ID, "policy")


def policy_detail(**overrides):
    values = {
        "detail_token": TOKEN,
        "masked_reference": "Referencia terminada en 3456",
        "full_reference": "083002914855",
        "branch_code": "91",
        "branch_name": "Salud colectivo",
        "classification": "confirmed",
        "insurer": "Aseguradora",
        "state": "Vigente",
        "holder": "Empresa autorizada",
        "start_date": "2026-01-01",
        "end_date": "2026-12-31",
        "renewable": "Sí",
        "payment_mode": "Fraccionado",
        "frequency": "Mensual",
        "installments": "12",
        "first_installment_date": "2026-01-01",
        "payment_calendar": (),
        "insured": (),
        "risks": (),
        "active_count": 1,
        "excluded_count": 0,
    }
    values.update(overrides)
    return PolicyDetail(**values)


MEMBER = GroupMember(
    role="Asegurado", display_name="Persona interna", id_type="CC", document="=2+2",
    masked_document="••••890", state="Activo", entry_date="2026-01-01",
    exit_date="", plan="Plan A", relationship="Titular", risk_summary="",
    economic_values=(("Pago total", "=2+2"),),
)


class FakePolicyService:
    profile = "sandbox"

    def detail(self, token):
        assert token == TOKEN
        return policy_detail()

    def group(self, token, *, source_kind=None):
        return self.detail(token), (MEMBER,)

    def _relations(self, policy_id):
        assert policy_id == POLICY_ID
        return ((
            {"Asegurado": {"id": "5234567890123456789"}, "Estado": "Activo", "Pago_total": "=2+2"},
        ), False)

    def _batch(self, module, fields, ids):
        if module == "Contacts":
            return {"5234567890123456789": {"id": "5234567890123456789", "Tipo_ID": "CC", "N_mero_de_ID": "001234", "Full_Name": "=FORMULA"}}
        return {}


class BranchAndTokenTests(SimpleTestCase):
    def test_five_closed_branches_and_shared_exequial_rule(self):
        self.assertEqual(set(COLLECTIVE_BRANCH_CONFIG), {"91", "86", "28", "83", "40"})
        self.assertEqual(classify_branch("Exequial colectivo").code, "86")
        self.assertIsNone(classify_branch("Exequial individual"))

    def test_branch_classification_accepts_confirmed_zoho_alias_and_typography(self):
        self.assertEqual(classify_branch("VG deudores").code, "83")
        self.assertEqual(classify_branch("  SALUD   COLECTIVO  ").code, "91")
        self.assertIsNone(classify_branch("Hogar"))

    def test_duplicate_branch_value_is_rejected(self):
        duplicate = dict(COLLECTIVE_BRANCH_CONFIG)
        duplicate["91"] = duplicate["86"]
        with self.assertRaises(BranchConfigurationError):
            validate_branch_config(duplicate)

    def test_typed_tokens_prevent_cross_entity_idor(self):
        company = sign_record_id("1234567890123456789", "company")
        self.assertEqual(unsign_record_id(company, "company"), "1234567890123456789")
        with self.assertRaises(ColectivosServiceError):
            unsign_record_id(company, "person")

    def test_excel_is_formula_safe_textual_and_has_three_sheets(self):
        content = build_current_policy_workbook(TOKEN, FakePolicyService())
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Información actual", "Información de póliza", "Metadatos"])
        self.assertEqual(workbook["Metadatos"].sheet_state, "hidden")
        self.assertEqual(workbook["Información actual"]["E2"].number_format, "@")
        self.assertTrue(workbook["Información actual"]["E2"].value.startswith("'="))
        self.assertTrue(workbook["Información actual"]["Q2"].value.startswith("'="))
        self.assertIn("Rol relacionado", tuple(cell.value for cell in workbook["Información actual"][1]))


@override_settings(ZOHO_ACTIVE_PROFILE="sandbox", COLECTIVOS_INTERNAL_PUBLIC_ACCESS=False)
class RequestWorkflowTests(TestCase):
    def setUp(self):
        User = get_user_model()
        self.creator = User.objects.create_superuser("workflow-admin", "workflow@example.test", "Password123!")
        self.owner = User.objects.create_user("workflow-owner", password="Password123!")

    def create_request(self):
        return create_request_from_policy(
            token=TOKEN, source_kind="company", actor=self.creator, assigned_to=self.owner,
            request_type=SolicitudColectivo.RequestType.UPDATE,
            deadline=timezone.localdate() + timedelta(days=5), internal_notes="Nota interna",
            service=FakePolicyService(),
        )

    def create_local_request(self, sequence):
        return SolicitudColectivo.objects.create(
            source_kind="company",
            source_reference_hash=f"{sequence:064x}",
            policy_reference_hash=f"{sequence + 1000:064x}",
            encrypted_policy_token="protected-reference",
            masked_policy_reference=f"Referencia {sequence}",
            client_label=f"Cliente {sequence}",
            branch_code="91",
            branch_name="Salud colectivo",
            request_type=SolicitudColectivo.RequestType.UPDATE,
            assigned_to=self.owner,
            deadline=timezone.localdate() + timedelta(days=5),
            zoho_profile="sandbox",
            encrypted_snapshot="protected-snapshot",
            created_by=self.creator,
        )

    def create_individual_response(self, *, submitted_at, policy_label="IND-9001"):
        context = {
            "policy_token": TOKEN,
            "source_kind": "company",
            "affiliate_key": "affiliate-key",
            "affiliate_label": "Afiliada Individual",
            "branch_slug": "salud",
            "branch_name": "Salud colectivo",
            "schema_version": 1,
            "creator_id": self.creator.pk,
            "policy_label": policy_label,
            "collective_context": "Cliente Individual",
            "requester_email": "original@example.test",
        }
        generated = generate_individual_access(
            context=context, actor=self.creator, recipient="edited@example.test",
        )
        quotation = CotizacionIndividual.objects.create(
            branch_code="91",
            branch_slug="salud",
            schema_version=1,
            encrypted_payload=encrypt(json.dumps({
                "fields": {
                    "requester_name": "Afiliada Individual",
                    "requester_email": "edited@example.test",
                },
                "groups": {"people": []},
                "context": context,
            })),
            payload_checksum="a" * 64,
            context_hash="b" * 64,
            created_by=self.creator,
        )
        CotizacionIndividual.objects.filter(pk=quotation.pk).update(submitted_at=submitted_at)
        quotation.refresh_from_db()
        access = generated.access
        access.quotation = quotation
        access.status = access.Status.USED
        access.used_at = submitted_at
        access.last_access_at = submitted_at
        access.save(update_fields=("quotation", "status", "used_at", "last_access_at"))
        NotificacionCotizacionIndividual.objects.create(
            user=self.creator,
            quotation=quotation,
            message="Respuesta individual recibida.",
            deduplication_key=f"individual:{quotation.public_id}",
        )
        return access, quotation

    def test_request_snapshot_rows_event_without_administrative_notification_or_zoho_write(self):
        item = self.create_request()
        self.assertTrue(item.public_id.startswith("COL-"))
        self.assertNotIn("Nota interna", item.encrypted_internal_notes)
        self.assertEqual(item.records.count(), 1)
        self.assertEqual(EventoSolicitudColectivo.objects.filter(request=item).count(), 1)
        self.assertFalse(
            NotificacionColectivos.objects.filter(request=item, user=self.owner).exists()
        )
        snapshot = request_snapshot(item)
        self.assertEqual(snapshot["policy"]["branch_code"], "91")
        self.assertEqual(snapshot["policy"]["reference"], "083002914855")
        self.assertEqual(snapshot["group"][0]["role"], "Asegurado")

    def test_duplicate_active_request_is_rejected(self):
        self.create_request()
        with self.assertRaisesMessage(ColectivosServiceError, "Ya existe"):
            self.create_request()

    def test_direct_flow_reuses_the_active_request(self):
        existing = self.create_request()
        item, created = create_or_reuse_request_from_policy(
            token=TOKEN,
            source_kind="company",
            actor=self.creator,
            assigned_to=self.owner,
            request_type=SolicitudColectivo.RequestType.UPDATE,
            deadline=timezone.localdate() + timedelta(days=15),
            service=FakePolicyService(),
        )
        self.assertFalse(created)
        self.assertEqual(item.pk, existing.pk)

    def test_state_machine_rejects_jump_and_records_valid_transition(self):
        item = self.create_request()
        with self.assertRaises(ValidationError):
            item.transition_to(SolicitudColectivo.Status.APPROVED)
        item = transition_request(request=item, target=SolicitudColectivo.Status.READY, actor=self.creator)
        self.assertEqual(item.status, SolicitudColectivo.Status.READY)
        self.assertEqual(item.events.filter(event_type="STATUS_CHANGED").count(), 1)

    def test_internal_views_require_permission_and_csrf(self):
        item = self.create_request()
        self.assertEqual(self.client.get(reverse("cotizacion_colectivos:request_list")).status_code, 403)
        self.client.force_login(self.creator)
        list_response = self.client.get(reverse("cotizacion_colectivos:request_list"))
        self.assertEqual(list_response.status_code, 200)
        with self.settings(COLECTIVOS_INTERNAL_PUBLIC_ACCESS=True):
            self.assertEqual(
                Client().get(reverse("cotizacion_colectivos:request_list")).status_code,
                200,
            )
        csrf_client = Client(enforce_csrf_checks=True)
        csrf_client.force_login(self.creator)
        response = csrf_client.post(reverse("cotizacion_colectivos:request_transition", args=[item.public_id]), {"target": SolicitudColectivo.Status.READY})
        self.assertEqual(response.status_code, 403)

    def test_request_list_has_deterministic_visible_order_and_stable_pages(self):
        items = [self.create_local_request(index) for index in range(1, 28)]
        anchor = timezone.now() - timedelta(days=1)
        for index, item in enumerate(items):
            stamp = anchor - timedelta(minutes=index)
            SolicitudColectivo.objects.filter(pk=item.pk).update(
                created_at=stamp,
                updated_at=stamp,
            )

        tied_updated = anchor - timedelta(minutes=1)
        SolicitudColectivo.objects.filter(pk=items[1].pk).update(
            created_at=anchor - timedelta(minutes=1),
            updated_at=tied_updated,
        )
        for item in (items[2], items[3]):
            SolicitudColectivo.objects.filter(pk=item.pk).update(
                created_at=anchor - timedelta(minutes=2),
                updated_at=tied_updated,
            )

        # Igual prioridad y actividad se resuelven por PK descendente para que
        # la paginacion de la bandeja unificada permanezca deterministica.
        expected = [items[0], items[3], items[2], items[1], *items[4:]]
        self.client.force_login(self.creator)
        first_page = self.client.get(reverse("cotizacion_colectivos:request_list"))
        second_page = self.client.get(
            reverse("cotizacion_colectivos:request_list"),
            {"page": 2},
        )
        first_ids = [item.pk for item in first_page.context["page"].object_list]
        second_ids = [item.pk for item in second_page.context["page"].object_list]

        self.assertEqual(first_ids, [item.pk for item in expected[:25]])
        self.assertEqual(second_ids, [item.pk for item in expected[25:]])
        self.assertFalse(set(first_ids) & set(second_ids))
        self.assertEqual(first_ids + second_ids, [item.pk for item in expected])

    def test_old_notification_route_redirects_to_the_canonical_inbox(self):
        item = self.create_local_request(99)
        notifications = [
            NotificacionColectivos.objects.create(
                user=self.creator,
                request=item,
                notification_type="CLIENT_RESPONSE",
                title=f"NotificaciÃ³n {index}",
                message="Mensaje seguro",
                deduplication_key=f"order:{index}",
            )
            for index in range(3)
        ]
        NotificacionColectivos.objects.create(
            user=self.creator,
            request=item,
            notification_type="ASSIGNED",
            title="Solicitud asignada",
            message="Mensaje histórico",
            deduplication_key="order:hidden-administrative",
        )
        self.client.force_login(self.creator)
        response = self.client.get(reverse("cotizacion_colectivos:notification_list"))
        self.assertRedirects(
            response,
            reverse("cotizacion_colectivos:request_list"),
            fetch_redirect_response=False,
        )
        self.assertEqual(len(notifications), 3)

    def test_administrative_notification_returns_to_informational_list_when_locked(self):
        item = self.create_local_request(100)
        notification = NotificacionColectivos.objects.create(
            user=self.creator,
            request=item,
            notification_type="STATUS",
            title="Solicitud actualizada",
            message="Mensaje seguro",
            deduplication_key="locked:notification",
        )
        self.client.force_login(self.creator)
        locked_queryset = MagicMock()
        locked_queryset.update.side_effect = OperationalError("database is locked")
        with patch.object(NotificacionColectivos.objects, "filter", return_value=locked_queryset):
            with self.assertLogs("cotizacion_colectivos", level="WARNING") as captured:
                response = self.client.post(
                    reverse("cotizacion_colectivos:notification_read", args=[notification.pk])
                )
        self.assertRedirects(
            response,
            reverse("cotizacion_colectivos:request_list"),
            fetch_redirect_response=False,
        )
        self.assertEqual(locked_queryset.update.call_count, 2)
        self.assertTrue(any("category=sqlite_locked" in line for line in captured.output))

    def test_inbox_uses_full_policy_search_and_prioritizes_answered_requests(self):
        waiting = self.create_local_request(201)
        answered = self.create_request()
        answered.status = answered.Status.ANSWERED
        answered.save(update_fields=("status", "updated_at"))
        self.client.force_login(self.creator)

        response = self.client.get(reverse("cotizacion_colectivos:request_list"))
        self.assertEqual(response.context["page"].object_list[0].pk, answered.pk)
        self.assertContains(response, "Póliza 083002914855")
        self.assertNotContains(response, "Póliza Referencia terminada en 3456")
        self.assertContains(response, answered.public_id)
        self.assertNotEqual(waiting.pk, response.context["page"].object_list[0].pk)

        searched = self.client.get(
            reverse("cotizacion_colectivos:request_list"), {"query": "083002914855"},
        )
        self.assertEqual([row.pk for row in searched.context["page"].object_list], [answered.pk])

    @patch("cotizacion_colectivos.views.PolicyService")
    def test_inbox_is_one_chronological_stream_for_requests_and_individual_quotes(self, policy_service):
        now = timezone.now()
        opened = self.create_local_request(301)
        opened.status = opened.Status.OPENED
        opened.save(update_fields=("status", "updated_at"))
        SolicitudColectivo.objects.filter(pk=opened.pk).update(updated_at=now)
        answered = self.create_request()
        answered.status = answered.Status.ANSWERED
        answered.save(update_fields=("status", "updated_at"))
        SolicitudColectivo.objects.filter(pk=answered.pk).update(
            updated_at=now - timedelta(minutes=10),
        )
        _access, quotation = self.create_individual_response(
            submitted_at=now - timedelta(minutes=5), policy_label="IND-9001",
        )
        self.client.force_login(self.creator)

        response = self.client.get(reverse("cotizacion_colectivos:request_list"))
        rows = response.context["page"].object_list
        self.assertEqual([row.inbox_kind for row in rows[:3]], [
            "individual", "request", "request",
        ])
        self.assertEqual(rows[0].quotation_id, quotation.pk)
        self.assertContains(response, 'data-inbox-kind="individual"', html=False)
        self.assertContains(response, 'data-inbox-kind="request"', html=False)
        self.assertContains(response, "Póliza IND-9001")
        self.assertNotContains(response, "Enlaces y respuestas")
        policy_service.assert_not_called()

        filtered = self.client.get(
            reverse("cotizacion_colectivos:request_list"),
            {"request_type": SolicitudColectivo.RequestType.QUOTE},
        )
        filtered_rows = filtered.context["page"].object_list
        self.assertEqual(len(filtered_rows), 1)
        self.assertEqual(filtered_rows[0].inbox_kind, "individual")

    def test_individual_response_opens_canonical_operational_expedient(self):
        _access, quotation = self.create_individual_response(
            submitted_at=timezone.now(), policy_label="POLIZA-COMPLETA-123",
        )
        token = sign_receipt(quotation.public_id)
        self.client.force_login(self.creator)

        inbox = self.client.get(reverse("cotizacion_colectivos:request_list"))
        canonical = reverse("cotizacion_colectivos:individual_expedient", args=[token])
        self.assertContains(inbox, f'href="{canonical}"', html=False)
        detail = self.client.get(canonical)
        self.assertContains(detail, "Expediente interno")
        self.assertContains(detail, "Póliza POLIZA-COMPLETA-123")
        self.assertContains(detail, "Cotización Individual")
        self.assertContains(detail, "Afiliada Individual")
        self.assertContains(detail, "edited@example.test")
        self.assertContains(detail, "Respuesta recibida")
        self.assertContains(detail, "Zoho")
        self.assertContains(detail, "Más información")
        self.assertContains(detail, '<details class="workspace-card technical-disclosure">', html=False)
        self.assertNotContains(detail, "<details open", html=False)

        legacy = self.client.get(reverse(
            "cotizacion_colectivos:individual_quotation_detail", args=[token],
        ))
        self.assertRedirects(legacy, canonical, fetch_redirect_response=False)

    def test_request_workspace_centers_human_response_and_collapses_history(self):
        item = self.create_request()
        item.status = item.Status.ANSWERED
        item.save(update_fields=("status", "updated_at"))
        response = RespuestaSolicitudColectivo.objects.create(
            request=item,
            version=1,
            status=RespuestaSolicitudColectivo.Status.SUBMITTED,
            origin=RespuestaSolicitudColectivo.Origin.WEB,
            submitted_at=timezone.now(),
            checksum="e" * 64,
            encrypted_client_observations=encrypt("Retiro solicitado por el cliente."),
        )
        CambioSolicitudColectivo.objects.create(
            response=response,
            policy=item.policies.first(),
            original_record=item.records.first(),
            action=CambioSolicitudColectivo.Action.RETIRE,
            functional_field="fecha_retiro",
            encrypted_previous_value=encrypt(""),
            encrypted_new_value=encrypt("2026-08-31"),
            position=1,
            checksum="f" * 64,
        )
        self.client.force_login(self.creator)
        page = self.client.get(reverse("cotizacion_colectivos:request_detail", args=[item.public_id]))

        self.assertContains(page, "Póliza 083002914855")
        self.assertContains(page, "Respuesta recibida")
        self.assertContains(page, "Fecha de retiro")
        self.assertContains(page, "2026-08-31")
        self.assertContains(page, "identificador técnico")
        self.assertContains(page, '<details class="workspace-card technical-disclosure">', html=False)
        self.assertNotContains(page, '<details open class="workspace-card technical-disclosure">', html=False)
        self.assertContains(page, "No disponible: faltan layout y reglas obligatorias confirmadas.")
