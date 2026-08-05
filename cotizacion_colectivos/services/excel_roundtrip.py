from __future__ import annotations

import hashlib
import io
import json
import zipfile
from collections import OrderedDict
from dataclasses import dataclass

from django.conf import settings
from django.core import signing
from django.core.signing import salted_hmac
from django.core.exceptions import ValidationError
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException

from vault.crypto import decrypt

from ..models import CambioSolicitudColectivo, RespuestaSolicitudColectivo, SolicitudColectivo
from .external import ACTION_TO_ADJUSTMENT, response_checksum
from ..filenames import safe_filename_part

TEMPLATE_VERSION = 3
SUPPORTED_TEMPLATE_VERSIONS = (1, 2, 3)
MAX_ROWS = 5000
LEGACY_HEADERS = ("Acción", "Tipo ID asociado", "ID asociado", "Nombre asociado", "Tipo ID asegurado", "ID asegurado", "Nombre asegurado", "Póliza", "Ramo", "Código de ramo", "Aseguradora", "Estado actual", "Estado solicitado", "Fecha de ingreso", "Fecha de retiro", "Fecha efectiva", "Parentesco", "Plan actual", "Plan solicitado", "Pago mensual con IVA actual", "Pago asegurado sin IVA actual", "Valor anterior", "Valor nuevo", "Tipo de novedad", "Motivo", "Observaciones", "Requiere adjunto", "Referencia de fila")
V2_HEADERS = (
    "Acción", "Tipo ID asociado", "ID asociado", "Nombre asociado",
    "Tipo ID asegurado", "ID asegurado", "Nombre asegurado",
    "Tipo ID beneficiario", "ID beneficiario", "Nombre beneficiario",
    "Póliza", "Ramo", "Código de ramo", "Aseguradora", "Estado actual",
    "Estado solicitado", "Fecha de ingreso", "Fecha de retiro", "Fecha efectiva",
    "Parentesco", "Plan actual", "Plan solicitado", "Pago mensual con IVA actual",
    "Pago asegurado sin IVA actual", "Valor anterior", "Valor nuevo",
    "Tipo de novedad", "Motivo", "Observaciones", "Ciudad", "Dirección",
    "Tipo de uso", "Año de construcción", "Vehículo", "Placa", "Marca",
    "Modelo", "Valor asegurado", "Requiere adjunto",
    "Referencia de fila",
)
HEADERS = V2_HEADERS[:-1] + (
    "Principal", "Rol principal", "Persona relacionada", "Roles",
    "Tipo de relación", "Referencia funcional de fila", "Referencia de fila",
)


def _safe(value):
    if value is None:
        return ""
    text = str(value)
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def _metadata_signature(request: SolicitudColectivo, nonce: str, sheet_map=(), row_map=()) -> str:
    return signing.dumps({
        "request": str(request.uuid), "branch": request.branch_code,
        "snapshot": request.snapshot_revision, "nonce": nonce,
        "sheets": tuple(sheet_map), "rows": tuple(row_map),
    }, salt="colectivos.excel.v1", compress=True)


def _policy_sheets(request: SolicitudColectivo):
    policies = list(request.policies.filter(active=True).order_by("position"))
    if not policies:
        return (("Novedades", None),)
    result = []
    used = set()
    for index, policy in enumerate(policies, 1):
        digits = "".join(character for character in policy.masked_policy_reference if character.isdigit())[-4:]
        base = safe_filename_part(policy.branch_name, fallback=f"Poliza_{index}", limit=20)
        suffix = digits or f"{index:02d}"
        name = f"{base}_{suffix}"[:31]
        if name in used:
            name = f"{name[:27]}_{index:02d}"
        used.add(name)
        result.append((name, policy))
    return tuple(result)


def _style_novelties_sheet(sheet) -> None:
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="155A96")
    sheet.freeze_panes = "A2"


def _record_payload(record) -> dict[str, object]:
    try:
        payload = json.loads(decrypt(record.encrypted_branch_payload))
    except (ValueError, TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _functional_records(records):
    """Agrupa por referencias técnicas HMAC, nunca por datos descriptivos."""
    grouped = OrderedDict()
    prefix_by_role = {"Afiliado": "associate", "Asegurado": "insured", "Beneficiario": "beneficiary"}
    for record in records:
        payload = _record_payload(record)
        prefix = prefix_by_role.get(str(record.role or ""), "")
        key = str(payload.get(f"{prefix}_key") or "") if prefix else ""
        key = key or str(payload.get("risk_key") or "")
        if not key:
            key = salted_hmac(
                "colectivos.excel.functional-row.v1", str(record.public_key)
            ).hexdigest()
        item = grouped.setdefault(key, {
            "key": key, "records": [], "payload": {}, "roles": set(),
        })
        item["records"].append(record)
        item["roles"].add(str(record.role or "Registro relacionado"))
        for field, value in payload.items():
            if value not in (None, "", [], {}, ()) and item["payload"].get(field) in (None, "", [], {}, ()):
                item["payload"][field] = value
    return tuple(grouped.values())


def build_novelties_template(request: SolicitudColectivo) -> bytes:
    nonce = hashlib.sha256(f"{request.uuid}:{timezone.now().isoformat()}".encode()).hexdigest()[:24]
    book = Workbook()
    sheet_specs = _policy_sheets(request)
    functional_row_map = []
    sheet = book.active
    for index, (sheet_name, policy_item) in enumerate(sheet_specs):
        if index:
            sheet = book.create_sheet(sheet_name)
        else:
            sheet.title = sheet_name
        _style_novelties_sheet(sheet)
        records = request.records.order_by("original_position")
        if policy_item is not None:
            records = records.filter(policy=policy_item)
        reference = policy_item.masked_policy_reference if policy_item else request.masked_policy_reference
        branch_name = policy_item.branch_name if policy_item else request.branch_name
        branch_code = policy_item.branch_code if policy_item else request.branch_code
        for functional in _functional_records(records[:MAX_ROWS]):
            record = functional["records"][0]
            current = functional["payload"]
            role = str(record.role or "")
            associate = (
                current.get("associate_id_type", ""), current.get("associate_document", ""), current.get("associate_name", ""),
            )
            insured = (
                current.get("insured_id_type", ""), current.get("insured_document", ""), current.get("insured_name", ""),
            )
            beneficiary = (
                current.get("beneficiary_id_type", ""), current.get("beneficiary_document", ""), current.get("beneficiary_name", ""),
            )
            legacy_person = (current.get("id_type", ""), current.get("document", ""), current.get("display_name", ""))
            if not any(associate) and role == "Afiliado":
                associate = legacy_person
            if not any(insured) and role == "Asegurado":
                insured = legacy_person
            if not any(beneficiary) and role == "Beneficiario":
                beneficiary = legacy_person
            risk = current.get("risk_attributes", {}) if isinstance(current.get("risk_attributes"), dict) else {}
            economics = dict(record.economic_values)
            roles = tuple(sorted(functional["roles"]))
            principal = current.get("associate_name") or current.get("insured_name") or current.get("display_name") or "Información protegida"
            related = current.get("beneficiary_name") or current.get("insured_name") or current.get("display_name") or ""
            functional_reference = functional["key"]
            source_keys = tuple(str(item.public_key) for item in functional["records"])
            functional_row_map.append((sheet_name, functional_reference, source_keys))
            sheet.append(tuple(_safe(value) for value in (
                "SIN_CAMBIOS",
                *associate, *insured, *beneficiary,
                reference, branch_name, branch_code, policy_item.insurer if policy_item else "", record.initial_status, "", record.entry_date, record.exit_date, "", current.get("relationship", ""), record.plan, "",
                economics.get("Pago según forma de pago", ""), economics.get("Pago empleado sin IVA", ""), "", "", "", "", "",
                risk.get("ciudad", ""), risk.get("direccion", ""), risk.get("tipo_uso", ""), risk.get("anio_construccion", ""), risk.get("vehiculo", ""), risk.get("placa", ""), risk.get("marca", ""), risk.get("modelo", ""), economics.get("Valor asegurado", ""),
                "No", principal, roles[0] if roles else role, related,
                ", ".join(roles), current.get("relationship", ""),
                functional_reference, functional_reference,
            )))
        sheet.auto_filter.ref = sheet.dimensions
    policy = book.create_sheet("Póliza")
    policy.append(("Solicitud", request.public_id))
    policy.append(("Fecha límite", request.deadline))
    policy.append(("Posición", "Póliza", "Ramo", "Aseguradora", "Estado", "Modalidad"))
    for index, (_, item) in enumerate(sheet_specs, 1):
        modality = item.get_modality_display() if item and item.modality != "NO_DETERMINADA" else ""
        policy.append(tuple(_safe(value) for value in (index, item.masked_policy_reference if item else request.masked_policy_reference, item.branch_name if item else request.branch_name, item.insurer if item else "", item.policy_status if item else "", modality)))
    instructions = book.create_sheet("Instrucciones")
    instructions.append(("Use únicamente las acciones del catálogo. No agregue fórmulas, macros ni cambie la hoja Metadatos.",))
    catalogs = book.create_sheet("Catálogos")
    catalogs.append(("Acciones", "Tipos de novedad", "Tipos de identificación"))
    for index, action in enumerate(CambioSolicitudColectivo.Action.values, 2):
        catalogs.cell(index, 1, action)
    catalogs.cell(1, 5, "Hoja")
    catalogs.cell(1, 6, "Acciones habilitadas")
    adjustment_actions = {value: key for key, value in ACTION_TO_ADJUSTMENT.items()}
    for index, (sheet_name, item) in enumerate(sheet_specs, 2):
        catalogs.cell(index, 5, sheet_name)
        enabled = item.enabled_adjustments if item else tuple(adjustment_actions)
        catalogs.cell(index, 6, ", ".join(adjustment_actions[value] for value in enabled if value in adjustment_actions))
    metadata = book.create_sheet("Metadatos")
    metadata.sheet_state = "hidden"
    sheet_map = tuple((name, item.position if item else 0) for name, item in sheet_specs)
    signed_rows = tuple((sheet_name, reference, tuple(source_keys)) for sheet_name, reference, source_keys in functional_row_map)
    values = (("version", TEMPLATE_VERSION), ("request", request.public_id), ("branch", request.branch_code), ("snapshot", request.snapshot_revision), ("nonce", nonce), ("checksum", _metadata_signature(request, nonce, sheet_map, signed_rows)), *(tuple((f"sheet:{name}", position) for name, position in sheet_map)), *(tuple((f"row:{sheet_name}:{reference}", ",".join(source_keys)) for sheet_name, reference, source_keys in functional_row_map)))
    for row in values:
        metadata.append(row)
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


@dataclass(frozen=True)
class ExcelPreview:
    rows: tuple[dict[str, str], ...]
    counts: dict[str, int]
    warnings: tuple[str, ...]


def parse_novelties(uploaded, request: SolicitudColectivo) -> ExcelPreview:
    name = str(getattr(uploaded, "name", "")).casefold()
    if not name.endswith(".xlsx") or name.endswith(".xlsm") or getattr(uploaded, "size", 0) > settings.COLECTIVOS_ATTACHMENT_MAX_BYTES:
        raise ValidationError("El archivo XLSX no es válido.")
    raw = uploaded.read()
    uploaded.seek(0)
    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            infos = archive.infolist()
            if len(infos) > 2_000 or sum(item.file_size for item in infos) > 50 * 1024 * 1024:
                raise ValidationError("El archivo XLSX excede los límites de seguridad.")
            names = {item.filename.casefold() for item in infos}
            if any("vbaproject" in item or item.endswith(".bin") or "externallinks" in item for item in names):
                raise ValidationError("El archivo contiene contenido no permitido.")
            for info in infos:
                if info.filename.casefold().endswith(".rels"):
                    relationship_xml = archive.read(info)
                    if b'TargetMode="External"' in relationship_xml or b"TargetMode='External'" in relationship_xml:
                        raise ValidationError("El archivo contiene vínculos externos no permitidos.")
        book = load_workbook(io.BytesIO(raw), read_only=True, data_only=False, keep_links=False)
    except (zipfile.BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValidationError("El archivo XLSX no es válido.") from exc
    required = {"Póliza", "Instrucciones", "Catálogos", "Metadatos"}
    if not required.issubset(book.sheetnames):
        raise ValidationError("Faltan hojas obligatorias.")
    meta = {str(row[0].value): row[1].value for row in book["Metadatos"].iter_rows(min_row=1, max_col=2) if row[0].value}
    try:
        payload = signing.loads(str(meta.get("checksum", "")), salt="colectivos.excel.v1", max_age=60 * 60 * 24 * 30)
    except signing.BadSignature as exc:
        raise ValidationError("Los metadatos no superan la validación.") from exc
    signed_sheets = tuple(tuple(item) for item in payload.get("sheets", ()))
    try:
        metadata_sheets = tuple((str(key).split(":", 1)[1], int(value or 0)) for key, value in meta.items() if str(key).startswith("sheet:"))
        metadata_rows = tuple(
            (
                str(key).split(":", 2)[1],
                str(key).split(":", 2)[2],
                tuple(filter(None, str(value or "").split(","))),
            )
            for key, value in meta.items() if str(key).startswith("row:")
        )
    except (TypeError, ValueError) as exc:
        raise ValidationError("Los metadatos no superan la validación.") from exc
    signed_rows = tuple(
        (str(item[0]), str(item[1]), tuple(str(value) for value in item[2]))
        for item in payload.get("rows", ())
    )
    if payload.get("request") != str(request.uuid) or payload.get("branch") != request.branch_code or payload.get("snapshot") != request.snapshot_revision or meta.get("version") not in SUPPORTED_TEMPLATE_VERSIONS or signed_sheets != metadata_sheets or signed_rows != metadata_rows:
        raise ValidationError("La plantilla no corresponde a esta solicitud.")
    functional_sources = {
        (sheet_name, reference): source_keys
        for sheet_name, reference, source_keys in metadata_rows
    }
    rows, seen, counts = [], set(), {action: 0 for action in CambioSolicitudColectivo.Action.values}
    zero_payments = negative_payments = 0
    total_rows = 0
    novelty_sheets = [str(key).split(":", 1)[1] for key in meta if str(key).startswith("sheet:")]
    if not novelty_sheets and "Novedades" in book.sheetnames:
        novelty_sheets = ["Novedades"]
    if not novelty_sheets or any(name not in book.sheetnames for name in novelty_sheets):
        raise ValidationError("Falta la información de novedades.")
    for sheet_name in novelty_sheets:
        sheet = book[sheet_name]
        headers = tuple(cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)))
        if headers not in {HEADERS, V2_HEADERS, LEGACY_HEADERS}:
            raise ValidationError("Los encabezados fueron alterados.")
        policy_position = int(meta.get(f"sheet:{sheet_name}", 0) or 0)
        policy_item = request.policies.filter(position=policy_position, active=True).first() if policy_position else None
        if request.policies.exists() and policy_item is None:
            raise ValidationError("Una hoja no corresponde a una póliza de la solicitud.")
        for position, cells in enumerate(sheet.iter_rows(min_row=2), 1):
            total_rows += 1
            if position > MAX_ROWS or total_rows > MAX_ROWS:
                raise ValidationError("El archivo supera el máximo de filas.")
            if any(cell.data_type == "f" or getattr(cell, "hyperlink", None) for cell in cells):
                raise ValidationError("No se permiten fórmulas ni hipervínculos.")
            values = {headers[index]: _safe(cell.value).strip() for index, cell in enumerate(cells[:len(headers)])}
            if not any(values.values()):
                continue
            action = values["Acción"].upper()
            if action not in counts:
                raise ValidationError("Existe una acción no permitida.")
            if policy_item is not None and ACTION_TO_ADJUSTMENT[action] not in set(policy_item.enabled_adjustments or ()):
                raise ValidationError("Una acción no está habilitada para la póliza indicada.")
            functional_template = headers == HEADERS
            reference = values.get("Referencia funcional de fila", "") if functional_template else values["Referencia de fila"]
            source_references = functional_sources.get((sheet_name, reference), ()) if functional_template else (reference,)
            if action != "INCLUIR":
                if not source_references:
                    raise ValidationError("Existe una referencia duplicada o inválida.")
                records = request.records.filter(public_key__in=source_references)
                if policy_item is not None:
                    records = records.filter(policy=policy_item)
                found = {str(item.public_key) for item in records.only("public_key")}
                if found != set(source_references) or reference in seen:
                    raise ValidationError("Existe una referencia duplicada o inválida.")
                seen.add(reference)
            rows.append({
                "record": source_references[0] if source_references else "",
                "records": source_references,
                "functional_key": reference if functional_template else "",
                "policy": str(policy_item.pk) if policy_item else "",
                "action": action, "plan": values["Plan solicitado"],
                "parentesco": values["Parentesco"], "fecha_efectiva": values["Fecha efectiva"],
                "fecha_ingreso": values["Fecha de ingreso"], "fecha_retiro": values["Fecha de retiro"],
                "motivo": values["Motivo"], "observaciones": values["Observaciones"],
                "ciudad": values.get("Ciudad", ""), "direccion": values.get("Dirección", ""),
                "tipo_uso": values.get("Tipo de uso", ""), "anio_construccion": values.get("Año de construcción", ""),
                "vehiculo": values.get("Vehículo", ""), "placa": values.get("Placa", ""),
                "marca": values.get("Marca", ""), "modelo": values.get("Modelo", ""),
                "valor_asegurado": values.get("Valor asegurado", ""),
            })
            counts[action] += 1
            for heading in ("Pago mensual con IVA actual", "Pago asegurado sin IVA actual"):
                raw_payment = values[heading].replace(",", "").strip()
                if not raw_payment:
                    continue
                try:
                    number = float(raw_payment)
                except ValueError:
                    continue
                zero_payments += number == 0
                negative_payments += number < 0
    counts.update({"VALIDAS": len(rows), "ERRORES": 0, "ADVERTENCIAS": 0, "DUPLICADOS": 0, "PAGOS_CERO": zero_payments, "PAGOS_NEGATIVOS": negative_payments})
    return ExcelPreview(tuple(rows), counts, tuple())


def build_comparison(response: RespuestaSolicitudColectivo) -> bytes:
    book = Workbook()
    summary = book.active
    summary.title = "Resumen"
    for row in (("Solicitud", response.request.public_id), ("Ramo", response.request.branch_name), ("Versión", response.version), ("Estado", response.status)):
        summary.append(row)
    changes = book.create_sheet("Cambios")
    changes.append(("Registro", "Póliza", "Ramo", "Campo", "Valor inicial", "Valor solicitado", "Decisión", "Valor aprobado", "Observación"))
    for change in response.changes.select_related("original_record", "policy"):
        review = change.reviews.order_by("-version").first()
        changes.append((str(change.original_record.public_key) if change.original_record else "INCLUSIÓN", change.policy.masked_policy_reference if change.policy else response.request.masked_policy_reference, change.policy.branch_name if change.policy else response.request.branch_name, change.functional_field, decrypt(change.encrypted_previous_value), decrypt(change.encrypted_new_value), review.decision if review else "PENDIENTE", decrypt(review.encrypted_approved_value) if review else "", ""))
    book.create_sheet("Novedades")
    book.create_sheet("Advertencias")
    metadata = book.create_sheet("Metadatos")
    metadata.sheet_state = "hidden"
    metadata.append(("checksum", response.checksum))
    stream = io.BytesIO(); book.save(stream); return stream.getvalue()


def build_response_workbook(response: RespuestaSolicitudColectivo) -> bytes:
    book = Workbook()
    summary = book.active
    summary.title = "Resumen"
    for row in (
        ("Solicitud", response.request.public_id),
        ("Ramo", response.request.branch_name),
        ("Origen", response.origin),
        ("Versión", response.version),
        ("Estado", response.status),
        ("Fecha de envío", response.submitted_at or ""),
    ):
        summary.append(tuple(_safe(value) for value in row))
    changes = book.create_sheet("Respuesta")
    changes.append(("Referencia", "Póliza", "Ramo", "Acción", "Campo", "Valor inicial", "Valor solicitado"))
    for change in response.changes.select_related("original_record", "policy"):
        changes.append((
            str(change.original_record.public_key) if change.original_record else "INCLUSIÓN",
            change.policy.masked_policy_reference if change.policy else response.request.masked_policy_reference,
            change.policy.branch_name if change.policy else response.request.branch_name,
            change.action,
            change.functional_field,
            _safe(decrypt(change.encrypted_previous_value)),
            _safe(decrypt(change.encrypted_new_value)),
        ))
    metadata = book.create_sheet("Metadatos")
    metadata.sheet_state = "hidden"
    metadata.append(("checksum", response.checksum))
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def build_approved_consolidated(response: RespuestaSolicitudColectivo) -> bytes:
    if response.status != response.Status.APPROVED or response.request.status != response.request.Status.APPROVED:
        raise ValidationError("La respuesta todavía no está aprobada.")
    changes = list(response.changes.select_related("original_record", "policy").prefetch_related("reviews"))
    latest_reviews = [(change, change.reviews.order_by("-version").first()) for change in changes]
    if not latest_reviews or any(review is None for _, review in latest_reviews):
        raise ValidationError("Existen decisiones pendientes.")
    if any(review.decision not in {"APROBAR", "APROBAR_CON_AJUSTE"} for _, review in latest_reviews):
        raise ValidationError("Existen decisiones no aprobadas.")
    book = Workbook()
    sheet = book.active
    sheet.title = "Consolidado aprobado"
    sheet.append((
        "Referencia", "Póliza", "Ramo", "Tipo de novedad", "Campo",
        "Valor anterior", "Valor aprobado", "Estado de revisión",
        "Responsable", "Solicitud",
    ))
    for change, review in latest_reviews:
        approved = decrypt(review.encrypted_approved_value) or decrypt(change.encrypted_new_value)
        sheet.append((
            str(change.original_record.public_key) if change.original_record else "INCLUSIÓN",
            change.policy.masked_policy_reference if change.policy else response.request.masked_policy_reference,
            change.policy.branch_name if change.policy else response.request.branch_name,
            change.action,
            change.functional_field,
            _safe(decrypt(change.encrypted_previous_value)),
            _safe(approved),
            review.decision,
            review.reviewer.get_username(),
            response.request.public_id,
        ))
    metadata = book.create_sheet("Metadatos")
    metadata.sheet_state = "hidden"
    metadata.append(("checksum", response.checksum))
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()
