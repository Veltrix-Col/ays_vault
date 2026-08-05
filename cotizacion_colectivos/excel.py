from __future__ import annotations

from io import BytesIO
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .services.common import unsign_record_context
from .services.policies import PolicyService

MAX_EXPORT_ROWS = 5000


def _functional_members(members):
    result = OrderedDict()
    for index, item in enumerate(members):
        key = {
            "Afiliado": item.associate_key,
            "Asegurado": item.insured_key,
            "Beneficiario": item.beneficiary_key,
        }.get(item.role, "") or f"record:{index}"
        current = result.get(key)
        if current is None:
            result[key] = {"item": item, "roles": {item.role}}
        else:
            current["roles"].add(item.role)
    return tuple((value["item"], tuple(sorted(value["roles"]))) for value in result.values())


def _safe_cell(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


def build_current_policy_workbook(token: str, service: PolicyService | None = None, *, include_detail: bool = False):
    service = service or PolicyService()
    context = unsign_record_context(token, "policy")
    detail, members = service.group(token, source_kind=context.get("source_kind"))
    if detail.classification != "confirmed" or not detail.branch_code:
        from .services.common import ColectivosServiceError

        raise ColectivosServiceError(
            "invalid_record", "La póliza no tiene una clasificación segura."
        )
    truncated = detail.truncated or len(members) > MAX_EXPORT_ROWS
    members = members[:MAX_EXPORT_ROWS]

    workbook = Workbook()
    current = workbook.active
    current.title = "Información actual"
    headers = (
        "Tipo ID asociado", "ID asociado", "Nombre asociado", "Tipo ID asegurado",
        "ID asegurado", "Nombre asegurado", "Póliza", "Ramo", "Código de ramo",
        "Aseguradora", "Estado asegurado", "Fecha de ingreso", "Fecha de retiro",
        "Parentesco", "Plan", "Prima", "Pago total", "Pago según forma de pago",
        "Pago empleado sin IVA", "Valor asegurado", "Observaciones actuales",
        "Rol relacionado", "Tipo ID beneficiario", "ID beneficiario",
        "Nombre beneficiario", "Correo", "Teléfono", "Móvil",
        "Riesgo", "Ciudad", "Dirección", "Tipo de uso", "Año de construcción",
        "Vehículo", "Placa", "Marca", "Modelo",
    )
    current.append(headers)
    for item, roles in _functional_members(members):
        economics = dict(item.economic_values)
        risk = dict(item.risk_attributes)
        associate = (
            item.associate_id_type, item.associate_document, item.associate_name,
        ) if item.associate_name or item.associate_document else (
            (item.id_type, item.document, item.display_name) if item.role == "Afiliado" else ("", "", "")
        )
        insured = (
            item.insured_id_type, item.insured_document, item.insured_name,
        ) if item.insured_name or item.insured_document else (
            (item.id_type, item.document, item.display_name) if item.role == "Asegurado" else ("", "", "")
        )
        beneficiary = (
            item.beneficiary_id_type, item.beneficiary_document, item.beneficiary_name,
        ) if item.beneficiary_name or item.beneficiary_document else (
            (item.id_type, item.document, item.display_name) if item.role == "Beneficiario" else ("", "", "")
        )
        current.append(tuple(_safe_cell(value) for value in (
            *associate, *insured,
            detail.masked_reference, detail.branch_name, detail.branch_code, detail.insurer,
            item.state, item.entry_date, item.exit_date, item.relationship, item.plan,
            economics.get("Prima", ""), economics.get("Pago total", ""), economics.get("Pago según forma de pago", ""), economics.get("Pago empleado sin IVA", ""),
            economics.get("Valor asegurado", ""), "", " / ".join(roles),
            *beneficiary,
            item.email, item.phone, item.mobile, item.risk_summary,
            risk.get("ciudad", ""), risk.get("direccion", ""), risk.get("tipo_uso", ""),
            risk.get("anio_construccion", ""), risk.get("vehiculo", ""),
            risk.get("placa", ""), risk.get("marca", ""), risk.get("modelo", ""),
        )))

    policy = workbook.create_sheet("Información de póliza")
    policy.append(("Campo", "Valor"))
    for label, value in (
        ("Póliza", detail.full_reference or detail.masked_reference), ("Ramo", detail.branch_name),
        ("Código", detail.branch_code), ("Aseguradora", detail.insurer),
        ("Inicio de vigencia", detail.start_date), ("Fin de vigencia", detail.end_date),
        ("Estado", detail.state), ("Modo de pago", detail.payment_mode),
        ("Frecuencia", detail.frequency), ("Número de cuotas", detail.installments),
        ("Primera cuota", detail.first_installment_date), ("Perfil", service.profile),
        ("Modo", "Solo lectura"), ("Resultado parcial", "Sí" if truncated else "No"),
    ):
        policy.append((_safe_cell(label), _safe_cell(value)))

    metadata = workbook.create_sheet("Metadatos")
    metadata.sheet_state = "hidden"
    metadata.append(("Versión de plantilla", "1"))
    metadata.append(("Ramo", detail.branch_name))
    metadata.append(("Código", detail.branch_code))
    metadata.append(("Perfil", service.profile))
    metadata.append(("Modo", "Solo lectura"))

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0875B8")
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 36)
            sheet.column_dimensions[column[0].column_letter].width = width
        text_columns = {
            cell.column for cell in sheet[1]
            if cell.value in {"ID asociado", "ID asegurado", "ID beneficiario"}
        }
        for row in sheet.iter_rows():
            for cell in row:
                if cell.column in text_columns and sheet.title == "Información actual":
                    cell.number_format = "@"
    output = BytesIO()
    workbook.save(output)
    content = output.getvalue()
    return (content, detail) if include_detail else content
